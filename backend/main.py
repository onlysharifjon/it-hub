import html
import json
import os
from uuid import uuid4
from datetime import datetime, timedelta, date
from decimal import Decimal
from typing import List, Optional
import io

from dotenv import load_dotenv
from fastapi import Depends, FastAPI, File, Header, HTTPException, Query, Request, UploadFile, status
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, HTMLResponse
from fastapi.security import HTTPAuthorizationCredentials, HTTPBearer
from fastapi.staticfiles import StaticFiles
import bcrypt as _bcrypt
from jose import JWTError, jwt
from sqlalchemy import func, extract
from sqlalchemy.orm import Session, selectinload, joinedload

from . import models, schemas, core_calc
from .database import get_db
from .models import UserRole

load_dotenv()

SECRET_KEY = os.getenv("SECRET_KEY", "change-me-in-production")
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = int(os.getenv("ACCESS_TOKEN_EXPIRE_MINUTES", "720"))  # default 12h
CORS_ORIGINS = os.getenv("CORS_ORIGINS", "http://localhost:5173,http://localhost:5174").split(",")

# Admin seed (parol faqat env orqali — kodda default parol yo'q)
ADMIN_USERNAME = os.getenv("ADMIN_USERNAME", "admin")
ADMIN_PASSWORD = os.getenv("ADMIN_PASSWORD")            # bo'lmasa admin yaratilmaydi
SEED_DEMO_USERS = os.getenv("SEED_DEMO_USERS", "false").lower() == "true"

auth_scheme = HTTPBearer(auto_error=False)

UPLOAD_DIR = os.getenv("UPLOAD_DIR", "/app/uploads")
AVATARS_DIR = f"{UPLOAD_DIR}/avatars"
CERTIFICATES_DIR = f"{UPLOAD_DIR}/certificates"
STUDENTS_DIR = f"{UPLOAD_DIR}/students"
STUDENT_PHOTOS_DIR = f"{UPLOAD_DIR}/student_photos"
STAFF_PHOTOS_DIR = f"{UPLOAD_DIR}/staff_photos"
os.makedirs(AVATARS_DIR, exist_ok=True)
os.makedirs(CERTIFICATES_DIR, exist_ok=True)
os.makedirs(STUDENTS_DIR, exist_ok=True)
os.makedirs(STUDENT_PHOTOS_DIR, exist_ok=True)
os.makedirs(STAFF_PHOTOS_DIR, exist_ok=True)

CAMERA_API_KEY = os.getenv("CAMERA_API_KEY", "")

# Standart docs'ni o'chiramiz — nginx `/api/` prefiksi bilan mos kelmaydi va o'rniga
# parol bilan himoyalangan variantini beramiz (openapi.json ham himoyalangan).
# openapi_url=None — FastAPI'ning himoyasiz built-in /openapi.json marshrutini o'chiradi.
DOCS_PASSWORD = os.getenv("DOCS_PASSWORD", "1107")
app = FastAPI(
    title="IT Hub — LMS API", version="3.0.0",
    docs_url=None, redoc_url=None, openapi_url=None,
)

# Public bazaviy yo'l (nginx orqali): https://crm.minaracademy.uz/api
PUBLIC_API_PREFIX = os.getenv("PUBLIC_API_PREFIX", "/api")
# Tashqi (mobil ilova) uchun to'liq URL yasashda ishlatiladi — masalan sertifikat PDF linki
PUBLIC_BASE_URL = os.getenv("PUBLIC_BASE_URL", "https://crm.minaracademy.uz").rstrip("/")


# ── Hujjatlar himoyasi (HTTP Basic — parol: DOCS_PASSWORD, default "1107") ────
# Brauzer parol so'raydi; login (username) e'tiborga olinmaydi, faqat parol tekshiriladi.
# Muvaffaqiyatli kirilgach brauzer Basic ma'lumotlarni /openapi.json'ga ham avtomatik yuboradi.
import secrets as _sec
from fastapi.security import HTTPBasic, HTTPBasicCredentials

_docs_basic = HTTPBasic(auto_error=False)


def require_docs_auth(creds: Optional[HTTPBasicCredentials] = Depends(_docs_basic)) -> bool:
    unauth = HTTPException(
        status_code=status.HTTP_401_UNAUTHORIZED,
        detail="Hujjatlar uchun parol kerak",
        headers={"WWW-Authenticate": 'Basic realm="Minar CRM Docs"'},
    )
    if not creds or not _sec.compare_digest(creds.password, DOCS_PASSWORD):
        raise unauth
    return True


@app.get("/docs", include_in_schema=False)
def custom_swagger_ui(_: bool = Depends(require_docs_auth)):
    from fastapi.openapi.docs import get_swagger_ui_html
    return get_swagger_ui_html(
        openapi_url=f"{PUBLIC_API_PREFIX}/openapi.json",
        title="Minar CRM API — Swagger",
    )


@app.get("/redoc", include_in_schema=False)
def custom_redoc(_: bool = Depends(require_docs_auth)):
    from fastapi.openapi.docs import get_redoc_html
    return get_redoc_html(
        openapi_url=f"{PUBLIC_API_PREFIX}/openapi.json",
        title="Minar CRM API — ReDoc",
    )


@app.get("/openapi.json", include_in_schema=False)
def protected_openapi(_: bool = Depends(require_docs_auth)):
    from fastapi.responses import JSONResponse
    return JSONResponse(app.openapi())

# ── Default users ──────────────────────────────────────────────────────────────

# Demo akkountlar faqat SEED_DEMO_USERS=true bo'lsa (dev muhitda) yaratiladi —
# parollar env orqali beriladi, kodda qattiq kodlangan parol yo'q.
DEMO_USERS = [
    {"username": "metodist",    "role": UserRole.metodist.value,    "full_name": "Metodist"},
    {"username": "teacher1",    "role": UserRole.teacher.value,     "full_name": "Sarvar Toshmatov"},
    {"username": "hunter1",     "role": UserRole.hunter.value,      "full_name": "Hunter 1"},
    {"username": "callcenter1", "role": UserRole.call_center.value, "full_name": "Call Center 1"},
]


# ── Default lead pipeline ────────────────────────────────────────────────────
# slug'lar eski LeadStatus enum qiymatlariga mos — status ustuni bilan sinxron.

DEFAULT_STAGES = [
    {"slug": "new",       "name": "Yangi",             "order": 10, "color": "sky",     "icon": "sparkles",       "kind": "lead"},
    {"slug": "called",    "name": "Qo'ng'iroq qilindi","order": 20, "color": "indigo",  "icon": "phone",          "kind": "lead"},
    {"slug": "callback",  "name": "Qayta ring",        "order": 30, "color": "amber",   "icon": "phone-incoming", "kind": "lead"},
    {"slug": "will_come", "name": "Keladi",            "order": 40, "color": "violet",  "icon": "calendar",       "kind": "lead"},
    {"slug": "enrolled",  "name": "Ro'yxatda",         "order": 50, "color": "emerald", "icon": "graduation-cap", "kind": "won"},
    {"slug": "rejected",  "name": "Rad etildi",        "order": 60, "color": "red",     "icon": "x-circle",       "kind": "lost"},
]

DEFAULT_SOURCES = [
    {"name": "Instagram",  "is_campaign": False},
    {"name": "Telegram",   "is_campaign": False},
    {"name": "Walk-in",    "is_campaign": False, "is_default": True},
    {"name": "Website",    "is_campaign": False},
    {"name": "Referral",   "is_campaign": False},
    {"name": "Qo'ng'iroq", "is_campaign": False},
    {"name": "Boshqa",     "is_campaign": False},
]


@app.on_event("startup")
def ensure_default_users() -> None:
    """Faqat admin akkountini ta'minlaydi (parol env orqali). Mavjud parollarga tegmaydi."""
    from .database import SessionLocal
    db = SessionLocal()
    try:
        # Admin — faqat ADMIN_PASSWORD berilgan bo'lsa va admin hali yo'q bo'lsa
        if ADMIN_PASSWORD:
            exists = db.query(models.User).filter(models.User.username == ADMIN_USERNAME).first()
            if not exists:
                db.add(models.User(
                    username=ADMIN_USERNAME,
                    hashed_password=hash_password(ADMIN_PASSWORD),
                    role=UserRole.admin.value,
                    full_name="Administrator",
                    is_active=True,
                    created_at=datetime.utcnow(),
                ))
        # Demo akkountlar — faqat aniq yoqilgan bo'lsa (dev muhit)
        if SEED_DEMO_USERS:
            demo_pw = os.getenv("DEMO_PASSWORD", "")
            if demo_pw:
                for u in DEMO_USERS:
                    if not db.query(models.User).filter(models.User.username == u["username"]).first():
                        db.add(models.User(
                            username=u["username"],
                            hashed_password=hash_password(demo_pw),
                            role=u["role"], full_name=u["full_name"],
                            is_active=True, created_at=datetime.utcnow(),
                        ))
        db.commit()
    finally:
        db.close()


@app.on_event("startup")
def ensure_default_pipeline() -> None:
    """Standart pipeline bosqichlari va lid manbalarini yaratadi (bir marta)."""
    from .database import SessionLocal
    db = SessionLocal()
    try:
        if db.query(models.LeadStage).count() == 0:
            for s in DEFAULT_STAGES:
                db.add(models.LeadStage(**s, created_at=datetime.utcnow()))
        if db.query(models.LeadSource).count() == 0:
            for s in DEFAULT_SOURCES:
                db.add(models.LeadSource(
                    name=s["name"],
                    is_campaign=s.get("is_campaign", False),
                    is_default=s.get("is_default", False),
                    is_active=True,
                    created_at=datetime.utcnow(),
                ))
        db.commit()
    except Exception:
        db.rollback()
    finally:
        db.close()


# Har bir tizimда doim mavjud bo'ladigan standart tariflar (guruh formasida default tanlov)
DEFAULT_TARIFFS = [
    {"name": "Pro",     "price": 700000},
    {"name": "Starter", "price": 500000},
]


@app.on_event("startup")
def ensure_default_tariffs() -> None:
    """Global 'Pro' va 'Starter' tariflarini ta'minlaydi (nomiga qarab, takrorlamaydi)."""
    from .database import SessionLocal
    db = SessionLocal()
    try:
        for t in DEFAULT_TARIFFS:
            exists = (
                db.query(models.Tariff)
                .filter(func.lower(func.trim(models.Tariff.name)) == t["name"].lower())
                .first()
            )
            if not exists:
                db.add(models.Tariff(name=t["name"], price=t["price"], is_active=True))
        db.commit()
    except Exception:
        db.rollback()
    finally:
        db.close()


_cors_origins = CORS_ORIGINS if CORS_ORIGINS != ["*"] else ["*"]
app.add_middleware(
    CORSMiddleware,
    allow_origins=_cors_origins,
    allow_credentials=_cors_origins != ["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

app.mount("/uploads", StaticFiles(directory=UPLOAD_DIR), name="uploads")

# Ota-ona mobil ilovasi API (/parent/...) — alohida paketda
from .parent.router_parent import router as parent_router
from .parent.notifications import notify_parents_of_student
app.include_router(parent_router)


# ── Rate limiting (in-memory; single-process pm2 fork) ───────────────────────
import time as _time
from collections import defaultdict

_rate_hits: dict = defaultdict(list)  # key -> [timestamps]


def _client_ip(request: Request) -> str:
    # nginx `proxy_set_header X-Real-IP $remote_addr;` bilan haqiqiy IP'ni qo'yadi va uni
    # qayta yozadi — shuning uchun ishonchli. X-Forwarded-For'ning eng chap qiymati mijoz
    # tomonidan soxtalashtiriladi (rate-limit'ni aylanib o'tishga imkon berardi), unga ishonmaymiz.
    real = request.headers.get("x-real-ip")
    if real:
        return real.strip()
    return request.client.host if request.client else "unknown"


def rate_limit(key: str, *, limit: int, window: int):
    """Oddiy sliding-window limiter. Limitdan oshsa 429 qaytaradi."""
    now = _time.time()
    hits = [t for t in _rate_hits[key] if now - t < window]
    if len(hits) >= limit:
        raise HTTPException(
            status_code=status.HTTP_429_TOO_MANY_REQUESTS,
            detail="Juda ko'p urinish. Birozdan so'ng qayta urinib ko'ring.",
        )
    hits.append(now)
    _rate_hits[key] = hits


# ── Auth helpers ──────────────────────────────────────────────────────────────

def hash_password(plain: str) -> str:
    return _bcrypt.hashpw(plain.encode(), _bcrypt.gensalt()).decode()


def verify_password(plain: str, hashed: str) -> bool:
    # Buzuq/bo'sh hash bcrypt'da ValueError beradi — 500 o'rniga "parol xato" (False) qaytaramiz.
    try:
        return _bcrypt.checkpw(plain.encode(), hashed.encode())
    except ValueError:
        return False


def create_access_token(data: dict) -> str:
    to_encode = data.copy()
    expire = datetime.utcnow() + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    to_encode["exp"] = expire
    return jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)


# Yuklab olish (chek/Excel) URL'lariga token query-parametrda beriladi — u nginx logi,
# brauzer tarixi va Referer orqali sizib chiqishi mumkin. Shuning uchun uzoq muddatli
# kirish tokeni emas, `typ=download` bo'lgan qisqa muddatli (2 daqiqa) token ishlatiladi.
DOWNLOAD_TOKEN_TTL_SECONDS = 120


def create_download_token(username: str) -> str:
    now = datetime.utcnow()
    return jwt.encode(
        {"sub": username, "typ": "download",
         "exp": now + timedelta(seconds=DOWNLOAD_TOKEN_TTL_SECONDS)},
        SECRET_KEY, algorithm=ALGORITHM,
    )


def require_auth(
    credentials: HTTPAuthorizationCredentials = Depends(auth_scheme),
    db: Session = Depends(get_db),
) -> models.User:
    exc = HTTPException(
        status_code=status.HTTP_401_UNAUTHORIZED,
        detail="Token xato yoki muddati o'tgan",
        headers={"WWW-Authenticate": "Bearer"},
    )
    if not credentials or credentials.scheme.lower() != "bearer":
        raise exc
    try:
        payload = jwt.decode(credentials.credentials, SECRET_KEY, algorithms=[ALGORITHM])
        username: str = payload.get("sub")
        if not username:
            raise exc
    except JWTError:
        raise exc
    user = db.query(models.User).filter(models.User.username == username).first()
    if not user or not user.is_active:
        raise exc
    # Check expiry on each request
    if user.expires_at and datetime.utcnow() > user.expires_at:
        user.is_active = False
        user.blocked_reason = user.blocked_reason or "Akkount muddati tugadi"
        user.blocked_at = datetime.utcnow()
        db.commit()
        raise exc
    return user


def require_staff(user: models.User = Depends(require_auth)) -> models.User:
    """Har qanday tizimga kirgan xodim (rol cheklovisiz)."""
    return user


def require_metodist(user: models.User = Depends(require_auth)) -> models.User:
    if user.role not in (UserRole.metodist.value, UserRole.admin.value):
        raise HTTPException(status_code=403, detail="Bu amal faqat metodist/admin uchun")
    return user


def require_admin(user: models.User = Depends(require_auth)) -> models.User:
    if user.role != UserRole.admin.value:
        raise HTTPException(status_code=403, detail="Bu amal faqat admin uchun")
    return user


def require_hunter(user: models.User = Depends(require_auth)) -> models.User:
    if user.role not in (UserRole.hunter.value, UserRole.admin.value):
        raise HTTPException(status_code=403, detail="Bu amal faqat hunter/admin uchun")
    return user


def require_call_center(user: models.User = Depends(require_auth)) -> models.User:
    if user.role not in (UserRole.call_center.value, UserRole.admin.value):
        raise HTTPException(status_code=403, detail="Bu amal faqat call_center/admin uchun")
    return user


def require_crm_access(user: models.User = Depends(require_auth)) -> models.User:
    """Hunter, Sales, Call center va admin lidlarga kirishi mumkin."""
    allowed = (UserRole.hunter.value, UserRole.sales.value,
               UserRole.call_center.value, UserRole.admin.value)
    if user.role not in allowed:
        raise HTTPException(status_code=403, detail="Ruxsat yo'q")
    return user


def require_attendance_editor(user: models.User = Depends(require_auth)) -> models.User:
    """Yo'qlama qilish: admin, metodist, hunter va teacher uchun ruxsat.
    Teacher faqat o'z guruhida — bu endpoint ichida tekshiriladi."""
    allowed = (UserRole.admin.value, UserRole.metodist.value,
               UserRole.hunter.value, UserRole.teacher.value)
    if user.role not in allowed:
        raise HTTPException(status_code=403, detail="Yo'qlama qilishga ruxsat yo'q")
    return user


def require_lms_write(user: models.User = Depends(require_auth)) -> models.User:
    """Metodist, Hunter, Sales, Call center va admin — talabalar va guruhlar bilan ishlash."""
    allowed = (UserRole.metodist.value, UserRole.hunter.value,
               UserRole.sales.value, UserRole.call_center.value, UserRole.admin.value)
    if user.role not in allowed:
        raise HTTPException(status_code=403, detail="Ruxsat yo'q")
    return user


def require_feedback_viewer(user: models.User = Depends(require_auth)) -> models.User:
    """Izohlar ro'yxati: admin, metodist, teacher, hunter va call_center."""
    allowed = (UserRole.admin.value, UserRole.metodist.value, UserRole.teacher.value,
               UserRole.hunter.value, UserRole.call_center.value)
    if user.role not in allowed:
        raise HTTPException(status_code=403, detail="Ruxsat yo'q")
    return user


def require_feedback_status(user: models.User = Depends(require_auth)) -> models.User:
    """Izoh statusini o'zgartirish (hal qilindi / javob bermadi): hunter, call_center va admin."""
    allowed = (UserRole.hunter.value, UserRole.call_center.value, UserRole.admin.value)
    if user.role not in allowed:
        raise HTTPException(status_code=403, detail="Bu amal faqat hunter/call_center/admin uchun")
    return user


def require_audit(user: models.User = Depends(require_auth)) -> models.User:
    """Xodimlarga ogohlantirish berish: audit va admin."""
    allowed = (UserRole.audit.value, UserRole.admin.value)
    if user.role not in allowed:
        raise HTTPException(status_code=403, detail="Bu amal faqat audit/admin uchun")
    return user


def _resolve_token(raw_token: str, db: Session, *, expect_download: bool) -> models.User:
    """Decode a raw JWT string and return the authenticated admin user.

    expect_download=True — token query-parametrda kelgan; faqat `typ=download` bo'lgan
    qisqa muddatli token qabul qilinadi (uzoq muddatli kirish tokeni URL'da yuborilmasin).
    """
    exc = HTTPException(
        status_code=status.HTTP_401_UNAUTHORIZED,
        detail="Token xato yoki muddati o'tgan",
        headers={"WWW-Authenticate": "Bearer"},
    )
    try:
        payload = jwt.decode(raw_token, SECRET_KEY, algorithms=[ALGORITHM])
        username: str = payload.get("sub")
        if not username:
            raise exc
    except JWTError:
        raise exc
    is_download = payload.get("typ") == "download"
    if expect_download and not is_download:
        raise exc
    if not expect_download and is_download:
        # Yuklab-olish tokeni oddiy API chaqiruvlari uchun ishlatilmasin.
        raise exc
    user = db.query(models.User).filter(models.User.username == username).first()
    if not user or not user.is_active:
        raise exc
    if user.role != UserRole.admin.value:
        raise HTTPException(status_code=403, detail="Bu amal faqat admin uchun")
    return user


def require_admin_download(
    _token: Optional[str] = Query(None),
    credentials: HTTPAuthorizationCredentials = Depends(auth_scheme),
    db: Session = Depends(get_db),
) -> models.User:
    """File-download auth: Bearer header (oddiy token) YOKI ?_token= (qisqa download tokeni)."""
    if credentials and credentials.scheme.lower() == "bearer":
        return _resolve_token(credentials.credentials, db, expect_download=False)
    if _token:
        return _resolve_token(_token, db, expect_download=True)
    raise HTTPException(
        status_code=status.HTTP_401_UNAUTHORIZED,
        detail="Token xato yoki muddati o'tgan",
        headers={"WWW-Authenticate": "Bearer"},
    )


# ── Audit log helper ──────────────────────────────────────────────────────────

def write_audit(db, *, entity_type, entity_id, action, changed_by_id, old_value=None, new_value=None):
    log = models.AuditLog(
        entity_type=entity_type,
        entity_id=entity_id,
        action=action,
        changed_by_id=changed_by_id,
        changed_at=datetime.utcnow(),
        old_value=json.dumps(old_value, ensure_ascii=False, default=str) if old_value is not None else None,
        new_value=json.dumps(new_value, ensure_ascii=False, default=str) if new_value is not None else None,
    )
    db.add(log)


# ── Auth endpoints ────────────────────────────────────────────────────────────

@app.post("/auth/login", response_model=schemas.TokenResponse)
def login(payload: schemas.LoginRequest, request: Request, db: Session = Depends(get_db)):
    ip = _client_ip(request)
    # Brute-force himoyasi: IP bo'yicha 10/min, login bo'yicha 5/min
    rate_limit(f"login-ip:{ip}", limit=10, window=60)
    rate_limit(f"login-user:{payload.username.lower()}", limit=5, window=60)

    user = db.query(models.User).filter(models.User.username == payload.username).first()
    if not user or not verify_password(payload.password, user.hashed_password):
        raise HTTPException(status_code=401, detail="Login yoki parol xato")

    # Muddati tugagan — avtomatik bloklash
    if user.expires_at and datetime.utcnow() > user.expires_at:
        if user.is_active:
            user.is_active = False
            user.blocked_reason = "Akkount muddati tugadi"
            user.blocked_at = datetime.utcnow()
            db.commit()
        raise HTTPException(status_code=403, detail={
            "code": "expired",
            "reason": user.blocked_reason or "Akkount muddati tugadi",
            "contact": user.blocked_contact or "",
        })

    # Bloklangan
    if not user.is_active:
        raise HTTPException(status_code=403, detail={
            "code": "blocked",
            "reason": user.blocked_reason or "Akkount faol emas",
            "contact": user.blocked_contact or "",
        })

    token = create_access_token({"sub": user.username, "role": user.role})
    return schemas.TokenResponse(access_token=token)


@app.get("/auth/me", response_model=schemas.UserRead)
def me(user: models.User = Depends(require_auth)):
    return user


@app.get("/auth/download-token")
def download_token(user: models.User = Depends(require_admin)):
    """Chek/Excel URL'lari uchun qisqa muddatli (2 daqiqa) token beradi.
    Uzoq muddatli kirish tokeni URL query-parametrida yuborilmasligi uchun."""
    return {"token": create_download_token(user.username), "expires_in": DOWNLOAD_TOKEN_TTL_SECONDS}


@app.put("/me", response_model=schemas.UserRead)
def update_me(payload: schemas.ProfileUpdate, db: Session = Depends(get_db), user: models.User = Depends(require_auth)):
    """Har qanday foydalanuvchi o'z profil ma'lumotlarini (ism, parol) o'zgartiradi."""
    if payload.full_name is not None:
        user.full_name = payload.full_name.strip() or None
    if payload.password is not None:
        # Parol o'zgartirilsa — joriy parol tasdiqlanishi shart
        if not payload.current_password or not verify_password(payload.current_password, user.hashed_password):
            raise HTTPException(status_code=400, detail="Joriy parol noto'g'ri")
        user.hashed_password = hash_password(payload.password)
    db.commit()
    db.refresh(user)
    return user


ALLOWED_IMAGE_TYPES = {"image/jpeg", "image/png", "image/webp", "image/gif"}
MAX_AVATAR_BYTES = 5 * 1024 * 1024  # 5 MB


@app.post("/me/avatar", response_model=schemas.UserRead)
async def upload_avatar(
    file: UploadFile = File(...),
    user: models.User = Depends(require_auth),
    db: Session = Depends(get_db),
):
    if file.content_type not in ALLOWED_IMAGE_TYPES:
        raise HTTPException(400, "Faqat JPEG, PNG, WebP yoki GIF rasm yuklash mumkin")

    contents = await file.read()
    if len(contents) > MAX_AVATAR_BYTES:
        raise HTTPException(400, "Fayl hajmi 5 MB dan oshmasin")

    # Kengaytmani content-type'dan aniqlaymiz — foydalanuvchi fayl nomiga ISHONMAYMIZ
    # (aks holda .html/.svg yuklab stored-XSS qilish mumkin edi)
    ext = {
        "image/jpeg": "jpg", "image/png": "png",
        "image/webp": "webp", "image/gif": "gif",
    }[file.content_type]
    filename = f"{user.id}.{ext}"
    filepath = os.path.join(AVATARS_DIR, filename)

    with open(filepath, "wb") as f:
        f.write(contents)

    user.avatar = f"avatars/{filename}"
    db.commit()
    db.refresh(user)
    return user


# ── Users endpoints ───────────────────────────────────────────────────────────

@app.get("/users", response_model=List[schemas.UserRead])
def list_users(
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    search: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    _: models.User = Depends(require_metodist),
):
    q = db.query(models.User)
    if search:
        q = q.filter(models.User.username.ilike(f"%{search}%") | models.User.full_name.ilike(f"%{search}%"))
    return q.order_by(models.User.id).offset((page - 1) * page_size).limit(page_size).all()


@app.get("/teachers", response_model=List[schemas.UserRead])
def list_teachers(db: Session = Depends(get_db), _: models.User = Depends(require_lms_write)):
    """Guruhga biriktirish uchun o'qituvchilar ro'yxati — metodist, hunter va admin uchun."""
    return (
        db.query(models.User)
        .filter(
            models.User.role.in_([UserRole.teacher.value, UserRole.metodist.value]),
            models.User.is_active == True,  # noqa: E712
        )
        .order_by(models.User.full_name, models.User.username)
        .all()
    )


@app.post("/users", response_model=schemas.UserRead, status_code=201)
def create_user(payload: schemas.UserCreate, db: Session = Depends(get_db), actor: models.User = Depends(require_admin)):
    if db.query(models.User).filter(models.User.username == payload.username).first():
        raise HTTPException(status_code=400, detail="Bu username allaqachon mavjud")
    user = models.User(
        username=payload.username,
        hashed_password=hash_password(payload.password),
        role=payload.role.value,
        full_name=payload.full_name,
        expires_at=payload.expires_at,
        telegram_chat_id=payload.telegram_chat_id,
    )
    db.add(user)
    db.flush()
    write_audit(db, entity_type="user", entity_id=user.id, action="create",
                changed_by_id=actor.id, new_value={"username": user.username, "role": user.role})
    db.commit()
    db.refresh(user)
    return user


@app.put("/users/{user_id}", response_model=schemas.UserRead)
def update_user(user_id: int, payload: schemas.UserUpdate, db: Session = Depends(get_db), actor: models.User = Depends(require_admin)):
    user = db.query(models.User).filter(models.User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="Foydalanuvchi topilmadi")
    old = {"role": user.role, "is_active": user.is_active}
    if payload.password is not None:
        user.hashed_password = hash_password(payload.password)
    if payload.role is not None:
        user.role = payload.role.value
    if payload.is_active is not None:
        user.is_active = payload.is_active
    if payload.full_name is not None:
        user.full_name = payload.full_name
    if payload.expires_at is not None:
        user.expires_at = payload.expires_at
    if payload.blocked_reason is not None:
        user.blocked_reason = payload.blocked_reason
    if payload.blocked_contact is not None:
        user.blocked_contact = payload.blocked_contact
    if payload.telegram_chat_id is not None:
        user.telegram_chat_id = payload.telegram_chat_id
    write_audit(db, entity_type="user", entity_id=user.id, action="update",
                changed_by_id=actor.id, old_value=old, new_value={"role": user.role, "is_active": user.is_active})
    db.commit()
    db.refresh(user)
    return user


@app.post("/users/{user_id}/block", response_model=schemas.UserRead)
def block_user(user_id: int, payload: schemas.BlockUserRequest, db: Session = Depends(get_db), actor: models.User = Depends(require_admin)):
    user = db.query(models.User).filter(models.User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="Foydalanuvchi topilmadi")
    if user.id == actor.id:
        raise HTTPException(status_code=400, detail="O'zingizni bloklolmaysiz")
    user.is_active = False
    user.blocked_reason = payload.reason
    user.blocked_contact = payload.contact
    user.blocked_at = datetime.utcnow()
    write_audit(db, entity_type="user", entity_id=user.id, action="block",
                changed_by_id=actor.id, new_value={"reason": payload.reason, "contact": payload.contact})
    db.commit()
    db.refresh(user)
    return user


@app.post("/users/{user_id}/unblock", response_model=schemas.UserRead)
def unblock_user(user_id: int, db: Session = Depends(get_db), actor: models.User = Depends(require_admin)):
    user = db.query(models.User).filter(models.User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="Foydalanuvchi topilmadi")
    user.is_active = True
    user.blocked_reason = None
    user.blocked_contact = None
    user.blocked_at = None
    write_audit(db, entity_type="user", entity_id=user.id, action="unblock",
                changed_by_id=actor.id, new_value={"is_active": True})
    db.commit()
    db.refresh(user)
    return user


@app.delete("/users/{user_id}/permanent")
def delete_user_permanent(user_id: int, db: Session = Depends(get_db), actor: models.User = Depends(require_admin)):
    """Foydalanuvchini bazadan butunlay o'chirish.

    O'chirishdan oldin userning o'zi va unga bog'liq barcha jadvallardagi
    yozuvlar (to'lov, lid, chegirma va h.k. — kim qilganligi) JSON backup
    faylga saqlanadi, so'ng user qatori butunlay o'chiriladi. Bog'liq
    yozuvlarning o'zi o'chirilmaydi — faqat "kim qildi" bog'lanishi
    tarixiy backup faylda qoladi (users.id qayta ishlatilmaydi, chunki
    autoincrement).
    """
    if user_id == actor.id:
        raise HTTPException(status_code=400, detail="O'zingizni butunlay o'chirolmaysiz")
    user = db.query(models.User).filter(models.User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="Foydalanuvchi topilmadi")
    if user.role == UserRole.admin.value:
        remaining_admins = db.query(models.User).filter(
            models.User.role == UserRole.admin.value, models.User.id != user_id,
        ).count()
        if remaining_admins == 0:
            raise HTTPException(status_code=400, detail="Oxirgi admin akkauntini o'chirib bo'lmaydi")

    user_snapshot = {c.name: getattr(user, c.name) for c in models.User.__table__.columns}

    related_records = {}
    for mapper in models.Base.registry.mappers:
        table = mapper.local_table
        if table is None or table.name == "users":
            continue
        for col in table.columns:
            for fk in col.foreign_keys:
                if fk.column.table.name == "users" and fk.column.name == "id":
                    rows = db.query(table).filter(col == user_id).all()
                    if rows:
                        related_records.setdefault(table.name, []).extend(
                            dict(row._mapping) for row in rows
                        )

    backup_dir = os.path.join(os.path.dirname(__file__), "deleted_users_backup")
    os.makedirs(backup_dir, exist_ok=True)
    backup_filename = f"user_{user_id}_{user.username}_{datetime.utcnow().strftime('%Y%m%dT%H%M%S')}.json"
    with open(os.path.join(backup_dir, backup_filename), "w", encoding="utf-8") as f:
        json.dump(
            {"user": user_snapshot, "related_records": related_records,
             "deleted_by": actor.username, "deleted_at": datetime.utcnow().isoformat()},
            f, ensure_ascii=False, indent=2, default=str,
        )

    write_audit(db, entity_type="user", entity_id=user_id, action="permanent_delete",
                changed_by_id=actor.id,
                old_value={"username": user.username, "role": user.role, "backup_file": backup_filename})
    db.query(models.User).filter(models.User.id == user_id).delete(synchronize_session=False)
    db.commit()
    return {"deleted": True, "backup_file": backup_filename}


# ── Lessons endpoints ─────────────────────────────────────────────────────────

@app.get("/lessons", response_model=List[schemas.LessonRead])
def list_lessons(
    category: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    _: models.User = Depends(require_auth),
):
    q = db.query(models.Lesson)
    if category:
        q = q.filter(models.Lesson.category == category.lower())
    return q.order_by(models.Lesson.lesson_number).all()


@app.get("/lessons/{lesson_id}", response_model=schemas.LessonRead)
def get_lesson(lesson_id: int, db: Session = Depends(get_db), _: models.User = Depends(require_auth)):
    lesson = db.query(models.Lesson).filter(models.Lesson.id == lesson_id).first()
    if not lesson:
        raise HTTPException(status_code=404, detail="Dars topilmadi")
    return lesson


@app.post("/lessons", response_model=schemas.LessonRead, status_code=201)
def create_lesson(payload: schemas.LessonCreate, db: Session = Depends(get_db), actor: models.User = Depends(require_metodist)):
    if payload.category not in schemas.LESSON_CATEGORIES:
        raise HTTPException(status_code=400, detail="Noto'g'ri kategoriya")
    if db.query(models.Lesson).filter(
        models.Lesson.category == payload.category,
        models.Lesson.lesson_number == payload.lesson_number
    ).first():
        raise HTTPException(status_code=400, detail="Bu kategoriyada ushbu dars raqami allaqachon mavjud")
    lesson = models.Lesson(**payload.dict(), updated_at=datetime.utcnow(), updated_by_id=actor.id)
    db.add(lesson)
    db.flush()
    write_audit(db, entity_type="lesson", entity_id=lesson.id, action="create",
                changed_by_id=actor.id, new_value=payload.dict())
    db.commit()
    db.refresh(lesson)
    return lesson


@app.put("/lessons/reorder", response_model=List[schemas.LessonRead])
def reorder_lessons(payload: schemas.ReorderRequest, db: Session = Depends(get_db), actor: models.User = Depends(require_metodist)):
    old_order = {}
    lessons_map = {}
    for item in payload.items:
        lesson = db.query(models.Lesson).filter(models.Lesson.id == item.id).first()
        if not lesson:
            raise HTTPException(status_code=404, detail=f"Dars {item.id} topilmadi")
        old_order[item.id] = lesson.lesson_number
        lessons_map[item.id] = (lesson, item.lesson_number)
    OFFSET = 100_000
    now = datetime.utcnow()
    for lesson, _ in lessons_map.values():
        lesson.lesson_number = lesson.lesson_number + OFFSET
    db.flush()
    for lesson, new_num in lessons_map.values():
        lesson.lesson_number = new_num
        lesson.updated_at = now
        lesson.updated_by_id = actor.id
    write_audit(db, entity_type="lesson", entity_id=None, action="reorder",
                changed_by_id=actor.id,
                old_value={str(k): v for k, v in old_order.items()},
                new_value={str(i.id): i.lesson_number for i in payload.items})
    db.commit()
    updated = [lessons_map[item.id][0] for item in payload.items]
    for lesson in updated:
        db.refresh(lesson)
    return updated


@app.put("/lessons/{lesson_id}", response_model=schemas.LessonRead)
def update_lesson(lesson_id: int, payload: schemas.LessonUpdate, db: Session = Depends(get_db), actor: models.User = Depends(require_metodist)):
    lesson = db.query(models.Lesson).filter(models.Lesson.id == lesson_id).first()
    if not lesson:
        raise HTTPException(status_code=404, detail="Dars topilmadi")
    changes = payload.dict(exclude_unset=True)
    old_value = {k: getattr(lesson, k) for k in changes}
    for field, value in changes.items():
        setattr(lesson, field, value)
    lesson.updated_at = datetime.utcnow()
    lesson.updated_by_id = actor.id
    write_audit(db, entity_type="lesson", entity_id=lesson.id, action="update",
                changed_by_id=actor.id, old_value=old_value, new_value=changes)
    db.commit()
    db.refresh(lesson)
    return lesson


@app.delete("/lessons/{lesson_id}", status_code=204)
def delete_lesson(lesson_id: int, db: Session = Depends(get_db), actor: models.User = Depends(require_metodist)):
    lesson = db.query(models.Lesson).filter(models.Lesson.id == lesson_id).first()
    if not lesson:
        raise HTTPException(status_code=404, detail="Dars topilmadi")
    write_audit(db, entity_type="lesson", entity_id=lesson_id, action="delete",
                changed_by_id=actor.id, old_value={"title": lesson.title})
    db.delete(lesson)
    db.commit()


# ── Audit log endpoints ───────────────────────────────────────────────────────

@app.get("/audit-logs")
def list_audit_logs(
    page: int = Query(1, ge=1),
    page_size: int = Query(25, ge=1, le=100),
    date_from: Optional[date] = Query(None),
    date_to: Optional[date] = Query(None),
    db: Session = Depends(get_db),
    _: models.User = Depends(require_metodist),
):
    q = db.query(models.AuditLog).order_by(models.AuditLog.changed_at.desc())
    if date_from:
        q = q.filter(models.AuditLog.changed_at >= datetime(date_from.year, date_from.month, date_from.day))
    if date_to:
        q = q.filter(models.AuditLog.changed_at < datetime(date_to.year, date_to.month, date_to.day) + timedelta(days=1))
    total = q.count()
    logs = q.offset((page - 1) * page_size).limit(page_size).all()
    return {
        "items": [schemas.AuditLogRead.from_orm(l) for l in logs],
        "meta": {
            "total": total,
            "page": page,
            "page_size": page_size,
            "total_pages": max(1, (total + page_size - 1) // page_size),
        }
    }


@app.get("/audit-logs/lesson/{lesson_id}", response_model=List[schemas.AuditLogRead])
def lesson_audit_logs(lesson_id: int, db: Session = Depends(get_db), _: models.User = Depends(require_metodist)):
    return (db.query(models.AuditLog)
            .filter(models.AuditLog.entity_type == "lesson", models.AuditLog.entity_id == lesson_id)
            .order_by(models.AuditLog.changed_at.desc()).all())


# ── Tariffs endpoints ─────────────────────────────────────────────────────────

@app.get("/tariffs", response_model=List[schemas.TariffRead])
def list_tariffs(db: Session = Depends(get_db), _: models.User = Depends(require_auth)):
    return db.query(models.Tariff).order_by(models.Tariff.name).all()


@app.post("/tariffs", response_model=schemas.TariffRead, status_code=201)
def create_tariff(payload: schemas.TariffCreate, db: Session = Depends(get_db), actor: models.User = Depends(require_hunter)):
    t = models.Tariff(**payload.dict())
    db.add(t)
    db.commit()
    db.refresh(t)
    return t


@app.put("/tariffs/{tariff_id}", response_model=schemas.TariffRead)
def update_tariff(tariff_id: int, payload: schemas.TariffUpdate, db: Session = Depends(get_db), actor: models.User = Depends(require_hunter)):
    t = db.query(models.Tariff).filter(models.Tariff.id == tariff_id).first()
    if not t:
        raise HTTPException(status_code=404, detail="Tarif topilmadi")
    for k, v in payload.dict(exclude_unset=True).items():
        setattr(t, k, v)
    db.commit()
    db.refresh(t)
    return t


@app.delete("/tariffs/{tariff_id}", status_code=204)
def delete_tariff(tariff_id: int, db: Session = Depends(get_db), actor: models.User = Depends(require_hunter)):
    t = db.query(models.Tariff).filter(models.Tariff.id == tariff_id).first()
    if not t:
        raise HTTPException(status_code=404, detail="Tarif topilmadi")
    db.delete(t)
    db.commit()


# ── Courses endpoints ─────────────────────────────────────────────────────────

@app.get("/courses", response_model=List[schemas.CourseRead])
def list_courses(db: Session = Depends(get_db), _: models.User = Depends(require_auth)):
    return db.query(models.Course).order_by(models.Course.name).all()


@app.post("/courses", response_model=schemas.CourseRead, status_code=201)
def create_course(payload: schemas.CourseCreate, db: Session = Depends(get_db), _: models.User = Depends(require_metodist)):
    c = models.Course(**payload.dict())
    db.add(c)
    db.commit()
    db.refresh(c)
    return c


@app.put("/courses/{course_id}", response_model=schemas.CourseRead)
def update_course(course_id: int, payload: schemas.CourseUpdate, db: Session = Depends(get_db), _: models.User = Depends(require_metodist)):
    c = db.query(models.Course).filter(models.Course.id == course_id).first()
    if not c:
        raise HTTPException(status_code=404, detail="Kurs topilmadi")
    for k, v in payload.dict(exclude_unset=True).items():
        setattr(c, k, v)
    db.commit()
    db.refresh(c)
    return c


@app.delete("/courses/{course_id}", status_code=204)
def delete_course(course_id: int, db: Session = Depends(get_db), _: models.User = Depends(require_metodist)):
    c = db.query(models.Course).filter(models.Course.id == course_id).first()
    if not c:
        raise HTTPException(status_code=404, detail="Kurs topilmadi")
    if db.query(models.Group).filter(models.Group.course_id == course_id).first():
        raise HTTPException(status_code=400, detail="Bu kursga bog'liq guruhlar mavjud")
    db.delete(c)
    db.commit()


# ── Students endpoints ────────────────────────────────────────────────────────

def _student_owed_total(s: models.Student, month: int = None, year: int = None,
                        discounts: list = None) -> Decimal:
    """Talabaning oylik to'liq summasi — faol guruhlar bo'yicha (tarif yoki guruh narxi).
    _student_month_owed bilan bir xil mantiq — Special chegirmalar qo'llanadi."""
    total = Decimal(0)
    for m in s.group_memberships:
        g = m.group
        if not g or not g.is_active:
            continue
        if m.tariff and m.tariff.price:
            price = Decimal(str(m.tariff.price))
        elif g.course_price and Decimal(str(g.course_price)) > 0:
            price = Decimal(str(g.course_price))
        else:
            continue
        if discounts:
            price = core_calc.apply_special_discounts(price, discounts, g.id, month, year)
        total += price
    return total.quantize(Decimal('1'))


def _students_payment_map(db: Session, students, month: int, year: int) -> dict:
    """Har talaba uchun (owed, paid, debt, status) — bir marta agregat so'rov bilan."""
    ids = [s.id for s in students]
    paid_by = {}
    discounts_by = {}
    if ids:
        rows = (
            db.query(models.Payment.student_id, func.sum(models.Payment.amount))
            .filter(models.Payment.student_id.in_(ids),
                    models.Payment.month == month, models.Payment.year == year)
            .group_by(models.Payment.student_id).all()
        )
        paid_by = {sid: Decimal(str(amt or 0)) for sid, amt in rows}
        # Special chegirmalar — bitta so'rovda
        for d in db.query(models.SpecialDiscount).filter(
            models.SpecialDiscount.student_id.in_(ids),
            models.SpecialDiscount.is_active == True,
        ).all():
            discounts_by.setdefault(d.student_id, []).append(d)
    out = {}
    for s in students:
        owed = _student_owed_total(s, month, year, discounts_by.get(s.id))
        paid = paid_by.get(s.id, Decimal(0))
        advance = Decimal(str(s.advance_balance or 0))
        # Avans balansi qarzni kamaytiradi
        advance_applied = min(advance, max(Decimal(0), owed - paid))
        covered = paid + advance
        debt = max(Decimal(0), owed - covered)
        if owed <= 0:
            status = "none"
        elif covered >= owed:
            status = "paid"
        elif covered > 0:
            status = "partial"
        else:
            status = "debtor"
        out[s.id] = (owed, paid, debt, status, advance_applied)
    return out


def _student_read(s: models.Student, pay: Optional[tuple] = None) -> schemas.StudentRead:
    owed = paid = debt = pstatus = advance_applied = None
    if pay is not None:
        owed, paid, debt, pstatus, advance_applied = pay
    return schemas.StudentRead(
        id=s.id, full_name=s.full_name, phone1=s.phone1,
        father_name=s.father_name, father_phone=s.father_phone,
        mother_name=s.mother_name, mother_phone=s.mother_phone,
        telegram_id=s.telegram_id, telegram_user_id=s.telegram_user_id,
        photo=s.photo, notes=s.notes, is_active=s.is_active,
        is_archived=s.is_archived, advance_balance=s.advance_balance,
        created_at=s.created_at, updated_at=s.updated_at, group_count=len(s.group_memberships),
        group_names=[m.group.name for m in s.group_memberships if m.group],
        owed_month=owed, paid_month=paid, debt=debt, payment_status=pstatus,
        advance_applied=advance_applied,
        photo_url=f"/uploads/{s.photo_path}" if s.photo_path else None,
    )


@app.get("/students")
def list_students(
    search: Optional[str] = Query(None),
    is_active: Optional[bool] = Query(None),
    is_archived: Optional[bool] = Query(None),
    date_from: Optional[date] = Query(None),
    date_to: Optional[date] = Query(None),
    payment: Optional[str] = Query(None, description="debtor | partial | paid | unpaid"),
    month: Optional[int] = Query(None, ge=1, le=12),
    year: Optional[int] = Query(None, ge=2020),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
    _: models.User = Depends(require_lms_write),
):
    now = datetime.utcnow()
    month = month or now.month
    year = year or now.year

    q = db.query(models.Student).options(
        joinedload(models.Student.group_memberships).joinedload(models.GroupStudent.tariff),
        joinedload(models.Student.group_memberships).joinedload(models.GroupStudent.group),
    )
    # by default exclude archived unless explicitly requested
    if is_archived is None:
        q = q.filter(models.Student.is_archived == False)
    else:
        q = q.filter(models.Student.is_archived == is_archived)
    if is_active is not None:
        q = q.filter(models.Student.is_active == is_active)
    if search:
        q = q.filter(
            models.Student.full_name.ilike(f"%{search}%") |
            models.Student.phone1.ilike(f"%{search}%") |
            models.Student.father_phone.ilike(f"%{search}%") |
            models.Student.mother_phone.ilike(f"%{search}%")
        )
    if date_from:
        q = q.filter(models.Student.created_at >= datetime(date_from.year, date_from.month, date_from.day))
    if date_to:
        q = q.filter(models.Student.created_at < datetime(date_to.year, date_to.month, date_to.day) + timedelta(days=1))
    q = q.order_by(models.Student.full_name)

    if payment:
        # To'lov holati bo'yicha filtr — butun to'plamni hisoblab, keyin sahifalash
        alls = q.all()
        pmap = _students_payment_map(db, alls, month, year)
        want = {"unpaid": {"debtor", "partial"}}.get(payment, {payment})
        filtered = [s for s in alls if pmap[s.id][3] in want]
        total = len(filtered)
        students = filtered[(page - 1) * page_size: page * page_size]
        page_map = pmap
    else:
        total = q.count()
        students = q.offset((page - 1) * page_size).limit(page_size).all()
        page_map = _students_payment_map(db, students, month, year)

    return {
        "items": [_student_read(s, page_map.get(s.id)) for s in students],
        "meta": {
            "total": total,
            "page": page,
            "page_size": page_size,
            "total_pages": max(1, (total + page_size - 1) // page_size),
            "month": month,
            "year": year,
        }
    }


@app.get("/students/{student_id}", response_model=schemas.StudentRead)
def get_student(student_id: int, db: Session = Depends(get_db), _: models.User = Depends(require_lms_write)):
    s = db.query(models.Student).filter(models.Student.id == student_id).first()
    if not s:
        raise HTTPException(status_code=404, detail="Talaba topilmadi")
    return _student_read(s)


@app.post("/students", response_model=schemas.StudentRead, status_code=201)
def create_student(payload: schemas.StudentCreate, db: Session = Depends(get_db), actor: models.User = Depends(require_lms_write)):
    data = payload.dict()
    advance = data.pop("advance", 0) or 0
    s = models.Student(**data, advance_balance=advance)
    db.add(s)
    db.flush()
    write_audit(db, entity_type="student", entity_id=s.id, action="create",
                changed_by_id=actor.id, new_value=payload.dict())
    db.commit()
    db.refresh(s)
    return _student_read(s)


@app.put("/students/{student_id}", response_model=schemas.StudentRead)
def update_student(student_id: int, payload: schemas.StudentUpdate, db: Session = Depends(get_db), actor: models.User = Depends(require_lms_write)):
    s = db.query(models.Student).filter(models.Student.id == student_id).first()
    if not s:
        raise HTTPException(status_code=404, detail="Talaba topilmadi")
    changes = payload.dict(exclude_unset=True)
    for k, v in changes.items():
        setattr(s, k, v)
    db.commit()
    db.refresh(s)
    return _student_read(s)


@app.post("/students/{student_id}/archive", response_model=schemas.StudentRead)
def archive_student(student_id: int, db: Session = Depends(get_db), actor: models.User = Depends(require_lms_write)):
    s = db.query(models.Student).filter(models.Student.id == student_id).first()
    if not s:
        raise HTTPException(status_code=404, detail="Talaba topilmadi")
    s.is_archived = True
    s.is_active = False
    db.commit()
    db.refresh(s)
    write_audit(db, entity_type="student", entity_id=s.id, action="archive",
                changed_by_id=actor.id, new_value={"is_archived": True})
    db.commit()
    return _student_read(s)


@app.post("/students/{student_id}/unarchive", response_model=schemas.StudentRead)
def unarchive_student(student_id: int, db: Session = Depends(get_db), actor: models.User = Depends(require_lms_write)):
    s = db.query(models.Student).filter(models.Student.id == student_id).first()
    if not s:
        raise HTTPException(status_code=404, detail="Talaba topilmadi")
    s.is_archived = False
    s.is_active = True
    db.commit()
    db.refresh(s)
    write_audit(db, entity_type="student", entity_id=s.id, action="unarchive",
                changed_by_id=actor.id, new_value={"is_archived": False})
    db.commit()
    return _student_read(s)


# ── Groups endpoints ──────────────────────────────────────────────────────────

def _group_read(g: models.Group, db: Session = None) -> schemas.GroupRead:
    stage = g.stage or 'foundation'
    total = g.course.total_lessons if g.course else schemas.STAGE_TOTAL_LESSONS.get(stage, 24)
    completed = 0
    if db is not None:
        completed = db.query(func.count(func.distinct(models.Attendance.lesson_date))).filter(
            models.Attendance.group_id == g.id
        ).scalar() or 0
    remaining = max(0, total - completed)
    pct = round(completed / total * 100, 1) if total > 0 else 0.0
    return schemas.GroupRead(
        id=g.id, name=g.name, stage=stage,
        course_id=g.course_id, course_name=g.course.name if g.course else None,
        teacher_id=g.teacher_id,
        teacher_name=g.teacher.full_name or g.teacher.username if g.teacher else None,
        course_price=g.course_price, schedule=g.schedule, lesson_time=g.lesson_time,
        start_date=g.start_date, is_active=g.is_active,
        telegram_chat_id=g.telegram_chat_id,
        created_at=g.created_at, student_count=len(g.members),
        total_lessons=total, completed_lessons=completed,
        remaining_lessons=remaining, progress_pct=pct,
    )


def _group_detail(g: models.Group, db: Session = None) -> schemas.GroupDetail:
    members = [
        schemas.GroupStudentRead(
            id=m.id, student_id=m.student_id,
            student_name=m.student.full_name,
            student_phone=m.student.phone1,
            joined_at=m.joined_at,
            tariff_id=m.tariff_id,
            tariff_name=m.tariff.name if m.tariff else None,
            tariff_price=m.tariff.price if m.tariff else None,
        ) for m in g.members
    ]
    base = _group_read(g, db=db)
    return schemas.GroupDetail(**base.dict(), members=members)


@app.get("/groups")
def list_groups(
    is_active: Optional[bool] = Query(None),
    search: Optional[str] = Query(None),
    date_from: Optional[date] = Query(None),
    date_to: Optional[date] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
    _: models.User = Depends(require_lms_write),
):
    q = db.query(models.Group)
    if is_active is not None:
        q = q.filter(models.Group.is_active == is_active)
    if search:
        q = q.filter(models.Group.name.ilike(f"%{search}%"))
    if date_from:
        q = q.filter(models.Group.start_date >= datetime(date_from.year, date_from.month, date_from.day))
    if date_to:
        q = q.filter(models.Group.start_date < datetime(date_to.year, date_to.month, date_to.day) + timedelta(days=1))
    total = q.count()
    groups = (
        q.options(
            joinedload(models.Group.teacher),
            joinedload(models.Group.course),
            selectinload(models.Group.members),
        )
        .order_by(models.Group.name)
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    return {
        "items": [_group_read(g, db=db) for g in groups],
        "meta": {
            "total": total,
            "page": page,
            "page_size": page_size,
            "total_pages": max(1, (total + page_size - 1) // page_size),
        }
    }


@app.get("/groups/{group_id}", response_model=schemas.GroupDetail)
def get_group(group_id: int, db: Session = Depends(get_db), actor: models.User = Depends(require_auth)):
    g = (
        db.query(models.Group)
        .options(
            joinedload(models.Group.teacher),
            joinedload(models.Group.course),
            selectinload(models.Group.members).joinedload(models.GroupStudent.student),
            selectinload(models.Group.members).joinedload(models.GroupStudent.tariff),
        )
        .filter(models.Group.id == group_id)
        .first()
    )
    if not g:
        raise HTTPException(status_code=404, detail="Guruh topilmadi")
    # Teacher faqat o'z guruhini ko'ra oladi
    if actor.role == UserRole.teacher.value and g.teacher_id != actor.id:
        raise HTTPException(status_code=403, detail="Bu guruh sizga tegishli emas")
    return _group_detail(g, db=db)


@app.post("/groups", response_model=schemas.GroupRead, status_code=201)
def create_group(payload: schemas.GroupCreate, db: Session = Depends(get_db), actor: models.User = Depends(require_lms_write)):
    data = payload.dict()
    data.setdefault('stage', 'foundation')
    g = models.Group(**data)
    db.add(g)
    db.flush()
    write_audit(db, entity_type="group", entity_id=g.id, action="create",
                changed_by_id=actor.id, new_value=data)
    db.commit()
    db.refresh(g)
    return _group_read(g, db=db)


@app.put("/groups/{group_id}", response_model=schemas.GroupRead)
def update_group(group_id: int, payload: schemas.GroupUpdate, db: Session = Depends(get_db), actor: models.User = Depends(require_lms_write)):
    g = db.query(models.Group).filter(models.Group.id == group_id).first()
    if not g:
        raise HTTPException(status_code=404, detail="Guruh topilmadi")
    changes = payload.dict(exclude_unset=True)
    old = {k: str(getattr(g, k)) for k in changes}
    for k, v in changes.items():
        setattr(g, k, v)
    write_audit(db, entity_type="group", entity_id=g.id, action="update",
                changed_by_id=actor.id, old_value=old,
                new_value={k: str(v) for k, v in changes.items()})
    db.commit()
    db.refresh(g)
    return _group_read(g, db=db)


@app.delete("/groups/{group_id}", status_code=204)
def delete_group(group_id: int, db: Session = Depends(get_db), actor: models.User = Depends(require_metodist)):
    g = db.query(models.Group).filter(models.Group.id == group_id).first()
    if not g:
        raise HTTPException(status_code=404, detail="Guruh topilmadi")
    db.delete(g)
    db.commit()


@app.post("/groups/{group_id}/students", status_code=201)
def add_student_to_group(group_id: int, payload: schemas.AddStudentToGroup, db: Session = Depends(get_db), actor: models.User = Depends(require_lms_write)):
    g = db.query(models.Group).filter(models.Group.id == group_id).first()
    if not g:
        raise HTTPException(status_code=404, detail="Guruh topilmadi")
    s = db.query(models.Student).filter(models.Student.id == payload.student_id).first()
    if not s:
        raise HTTPException(status_code=404, detail="Talaba topilmadi")
    exists = db.query(models.GroupStudent).filter(
        models.GroupStudent.group_id == group_id,
        models.GroupStudent.student_id == payload.student_id
    ).first()
    if exists:
        raise HTTPException(status_code=400, detail="Talaba bu guruhda allaqachon mavjud")
    tariff_id = payload.tariff_id
    if tariff_id:
        tariff = db.query(models.Tariff).filter(models.Tariff.id == tariff_id).first()
        if not tariff:
            raise HTTPException(status_code=404, detail="Tarif topilmadi")
    elif g.course_price and g.course_price > 0:
        # Auto-assign a tariff matching the group's course_price
        matched = db.query(models.Tariff).filter(
            models.Tariff.price == g.course_price,
            models.Tariff.is_active == True,
        ).first()
        tariff_id = matched.id if matched else None
    gs = models.GroupStudent(group_id=group_id, student_id=payload.student_id, tariff_id=tariff_id)
    db.add(gs)
    db.commit()
    return {"message": "Talaba guruhga qo'shildi"}


@app.delete("/groups/{group_id}/students/{student_id}", status_code=204)
def remove_student_from_group(group_id: int, student_id: int, db: Session = Depends(get_db), actor: models.User = Depends(require_lms_write)):
    gs = db.query(models.GroupStudent).filter(
        models.GroupStudent.group_id == group_id,
        models.GroupStudent.student_id == student_id
    ).first()
    if not gs:
        raise HTTPException(status_code=404, detail="Topilmadi")
    db.delete(gs)
    db.commit()


# ── Payments endpoints ────────────────────────────────────────────────────────

# To'lov hisobi core_calc.py da (yagona manba) — bu yerda faqat delegatsiya.
_student_month_owed = core_calc.student_month_owed
_student_month_paid = core_calc.student_month_paid


def _payment_read(p: models.Payment, db: Optional[Session] = None) -> schemas.PaymentRead:
    expected = paid_total = remaining = None
    status = None
    if db is not None:
        expected = _student_month_owed(db, p.student_id, p.group_id, p.month, p.year)
        paid_total = _student_month_paid(db, p.student_id, p.group_id, p.month, p.year)
        if expected and expected > 0:
            remaining = max(Decimal(0), expected - paid_total)
            status = "paid" if paid_total >= expected else "partial"
    return schemas.PaymentRead(
        id=p.id, student_id=p.student_id,
        student_name=p.student.full_name if p.student else None,
        group_id=p.group_id,
        group_name=p.group.name if p.group else None,
        amount=p.amount, month=p.month, year=p.year,
        paid_at=p.paid_at, notes=p.notes,
        recorded_by_name=(p.recorded_by.full_name or p.recorded_by.username) if p.recorded_by else None,
        expected=expected, paid_total=paid_total, remaining=remaining, status=status,
    )


@app.get("/payments/expected")
def payment_expected(
    student_id: int = Query(...),
    group_id: int = Query(...),
    month: int = Query(..., ge=1, le=12),
    year: int = Query(..., ge=2020),
    db: Session = Depends(get_db),
    _: models.User = Depends(require_crm_access),
):
    """Talabaning shu oy uchun to'liq summasi, to'langani va qolgani."""
    expected = _student_month_owed(db, student_id, group_id, month, year)
    paid = _student_month_paid(db, student_id, group_id, month, year)
    remaining = max(Decimal(0), expected - paid) if expected > 0 else Decimal(0)
    return {
        "expected": float(expected),
        "paid": float(paid),
        "remaining": float(remaining),
    }


@app.get("/students/{student_id}/payments/summary", response_model=schemas.StudentPaymentSummary)
def student_payment_summary(
    student_id: int,
    month: Optional[int] = Query(None, ge=1, le=12),
    year: Optional[int] = Query(None, ge=2020),
    db: Session = Depends(get_db),
    _: models.User = Depends(require_crm_access),
):
    """Talabaning joriy oy uchun to'lov xulosasi — guruhlar kesimida qarz, avans va so'nggi to'lovlar."""
    now = datetime.utcnow()
    month, year = month or now.month, year or now.year

    s = (
        db.query(models.Student)
        .options(
            joinedload(models.Student.group_memberships).joinedload(models.GroupStudent.group),
            joinedload(models.Student.group_memberships).joinedload(models.GroupStudent.tariff),
        )
        .filter(models.Student.id == student_id).first()
    )
    if not s:
        raise HTTPException(404, "Talaba topilmadi")

    groups: list[schemas.StudentGroupFee] = []
    total_owed = total_paid = Decimal(0)
    for m in s.group_memberships:
        g = m.group
        if not g or not g.is_active:
            continue
        owed = _student_month_owed(db, s.id, g.id, month, year)
        paid = _student_month_paid(db, s.id, g.id, month, year)
        total_owed += owed
        total_paid += paid
        groups.append(schemas.StudentGroupFee(
            group_id=g.id, group_name=g.name, owed=owed, paid=paid,
            remaining=max(Decimal(0), owed - paid),
        ))

    advance = Decimal(str(s.advance_balance or 0))
    covered = total_paid + advance
    debt = max(Decimal(0), total_owed - covered)
    if total_owed <= 0:
        status = "none"
    elif covered >= total_owed:
        status = "paid"
    elif covered > 0:
        status = "partial"
    else:
        status = "debtor"

    recent = (
        db.query(models.Payment)
        .options(
            joinedload(models.Payment.group),
            joinedload(models.Payment.student),
            joinedload(models.Payment.recorded_by),
        )
        .filter(models.Payment.student_id == student_id)
        .order_by(models.Payment.paid_at.desc()).limit(10).all()
    )
    return schemas.StudentPaymentSummary(
        student_id=s.id, student_name=s.full_name, month=month, year=year,
        advance_balance=advance, total_owed=total_owed, total_paid=total_paid,
        debt=debt, payment_status=status, groups=groups,
        recent_payments=[_payment_read(p, db) for p in recent],
    )


@app.get("/payments")
def list_payments(
    student_id: Optional[int] = Query(None),
    group_id: Optional[int] = Query(None),
    month: Optional[int] = Query(None),
    year: Optional[int] = Query(None),
    date_from: Optional[date] = Query(None),
    date_to: Optional[date] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(25, ge=1, le=100),
    db: Session = Depends(get_db),
    _: models.User = Depends(require_crm_access),
):
    q = db.query(models.Payment)
    if student_id:
        q = q.filter(models.Payment.student_id == student_id)
    if group_id:
        q = q.filter(models.Payment.group_id == group_id)
    if month:
        q = q.filter(models.Payment.month == month)
    if year:
        q = q.filter(models.Payment.year == year)
    if date_from:
        q = q.filter(models.Payment.paid_at >= datetime(date_from.year, date_from.month, date_from.day))
    if date_to:
        q = q.filter(models.Payment.paid_at < datetime(date_to.year, date_to.month, date_to.day) + timedelta(days=1))
    total = q.count()
    payments = q.order_by(models.Payment.paid_at.desc()).offset((page - 1) * page_size).limit(page_size).all()
    return {
        "items": [_payment_read(p, db) for p in payments],
        "meta": {
            "total": total,
            "page": page,
            "page_size": page_size,
            "total_pages": max(1, (total + page_size - 1) // page_size),
        }
    }


@app.post("/payments", response_model=schemas.PaymentRead, status_code=201)
def create_payment(payload: schemas.PaymentCreate, db: Session = Depends(get_db), actor: models.User = Depends(require_crm_access)):
    p = models.Payment(**payload.dict(), paid_at=datetime.utcnow(), recorded_by_id=actor.id)
    db.add(p)
    db.flush()
    write_audit(db, entity_type="payment", entity_id=p.id, action="create",
                changed_by_id=actor.id, new_value=payload.dict())
    # Ota-onalarga bildirishnoma
    _amount = int(Decimal(str(payload.amount)))
    notify_parents_of_student(
        db, student_id=payload.student_id, ntype="payment",
        title="To'lov qabul qilindi",
        body=f"{_amount:,} so'm to'lov qabul qilindi".replace(",", " "),
    )
    db.commit()
    db.refresh(p)
    return _payment_read(p, db)


@app.put("/payments/{payment_id}", response_model=schemas.PaymentRead)
def update_payment(payment_id: int, payload: schemas.PaymentUpdate, db: Session = Depends(get_db), actor: models.User = Depends(require_admin)):
    p = db.query(models.Payment).filter(models.Payment.id == payment_id).first()
    if not p:
        raise HTTPException(status_code=404, detail="To'lov topilmadi")
    changes = payload.dict(exclude_unset=True)
    old = {k: str(getattr(p, k)) for k in changes}
    for k, v in changes.items():
        setattr(p, k, v)
    write_audit(db, entity_type="payment", entity_id=p.id, action="update",
                changed_by_id=actor.id, old_value=old,
                new_value={k: str(v) for k, v in changes.items()})
    db.commit()
    db.refresh(p)
    return _payment_read(p, db)


@app.delete("/payments/{payment_id}", status_code=204)
def delete_payment(payment_id: int, db: Session = Depends(get_db), actor: models.User = Depends(require_admin)):
    p = db.query(models.Payment).filter(models.Payment.id == payment_id).first()
    if not p:
        raise HTTPException(status_code=404, detail="To'lov topilmadi")
    write_audit(db, entity_type="payment", entity_id=p.id, action="delete",
                changed_by_id=actor.id,
                old_value={"student_id": p.student_id, "group_id": p.group_id,
                           "amount": str(p.amount), "month": p.month, "year": p.year})
    db.delete(p)
    db.commit()


# ── Salary helper ─────────────────────────────────────────────────────────────

def _group_salary(db: Session, group: models.Group, month: int, year: int):
    """
    O'qituvchi maoshini hisoblaydi — 2 query (N+1 o'rniga).
    Formula: teacher_pay_per_student × talaba_kelgan_darslar_soni
    """
    pay_per = Decimal(str(group.teacher_pay_per_student or 0))
    if pay_per == 0:
        return Decimal(0), []

    # 1-query: guruh a'zolari va ismlari
    member_rows = (
        db.query(models.GroupStudent.student_id, models.Student.full_name)
        .join(models.Student, models.Student.id == models.GroupStudent.student_id)
        .filter(models.GroupStudent.group_id == group.id)
        .all()
    )
    if not member_rows:
        return Decimal(0), []

    student_ids = [r.student_id for r in member_rows]
    name_map    = {r.student_id: r.full_name for r in member_rows}

    # 2-query: bir oyda har talaba uchun kelgan darslar soni (GROUP BY)
    att_rows = (
        db.query(
            models.Attendance.student_id,
            func.count(models.Attendance.id).label('cnt'),
        )
        .filter(
            models.Attendance.group_id  == group.id,
            models.Attendance.is_present == True,
            extract('month', models.Attendance.lesson_date) == month,
            extract('year',  models.Attendance.lesson_date) == year,
            models.Attendance.student_id.in_(student_ids),
        )
        .group_by(models.Attendance.student_id)
        .all()
    )
    attended_map = {r.student_id: r.cnt for r in att_rows}

    total    = Decimal(0)
    students = []
    for sid, sname in name_map.items():
        attended = attended_map.get(sid, 0)
        share    = (pay_per * Decimal(str(attended))).quantize(Decimal('1'))
        total   += share
        students.append({
            "student_id":   sid,
            "student_name": sname,
            "attended":     attended,
            "salary_share": float(share),
        })
    return total, students


def _batch_salary_by_month(db: Session, groups: list, year: int) -> dict:
    """
    Barcha guruhlar uchun 12 oylik maoshni 1 query da hisoblaydi.
    Qaytaradi: {month: total_salary}
    """
    active = [g for g in groups if (g.teacher_pay_per_student or 0) > 0]
    result = {m: Decimal(0) for m in range(1, 13)}
    if not active:
        return result

    group_pay   = {g.id: Decimal(str(g.teacher_pay_per_student)) for g in active}
    group_ids   = list(group_pay.keys())

    rows = (
        db.query(
            extract('month', models.Attendance.lesson_date).label('mon'),
            models.Attendance.group_id,
            models.Attendance.student_id,
            func.count(models.Attendance.id).label('cnt'),
        )
        .filter(
            models.Attendance.group_id.in_(group_ids),
            models.Attendance.is_present == True,
            extract('year', models.Attendance.lesson_date) == year,
        )
        .group_by(
            extract('month', models.Attendance.lesson_date),
            models.Attendance.group_id,
            models.Attendance.student_id,
        )
        .all()
    )
    for row in rows:
        pay = group_pay.get(row.group_id, Decimal(0))
        result[int(row.mon)] += (pay * Decimal(str(row.cnt))).quantize(Decimal('1'))
    return result


# ── Statistics endpoints ──────────────────────────────────────────────────────

@app.get("/stats/overview", response_model=schemas.StatsOverview)
def stats_overview(
    year: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    _: models.User = Depends(require_admin),
):
    now = datetime.utcnow()
    cur_month, cur_year = now.month, now.year
    sel_year = year or cur_year

    prev_month = cur_month - 1 if cur_month > 1 else 12
    prev_year = cur_year if cur_month > 1 else cur_year - 1

    total_students = db.query(func.count(models.Student.id)).scalar()
    active_students = db.query(func.count(models.Student.id)).filter(models.Student.is_active == True).scalar()
    total_groups = db.query(func.count(models.Group.id)).scalar()
    active_groups_count = db.query(func.count(models.Group.id)).filter(models.Group.is_active == True).scalar()

    # Load active groups (1 query)
    active_grps = db.query(models.Group).filter(models.Group.is_active == True).all()

    # Batch: barcha to'lovlar yil bo'yicha (1 query)
    pay_rows = (
        db.query(models.Payment.month, models.Payment.year,
                 func.sum(models.Payment.amount).label('total'),
                 func.count(models.Payment.id).label('cnt'))
        .filter(models.Payment.year.in_([sel_year, cur_year, prev_year]))
        .group_by(models.Payment.month, models.Payment.year)
        .all()
    )
    pay_map = {(r.month, r.year): (r.total or Decimal(0), r.cnt) for r in pay_rows}

    # Batch: barcha xarajatlar yil bo'yicha (1 query)
    exp_rows = (
        db.query(models.Expense.month, models.Expense.year,
                 func.sum(models.Expense.amount).label('total'))
        .filter(models.Expense.year.in_([sel_year, cur_year, prev_year]))
        .group_by(models.Expense.month, models.Expense.year)
        .all()
    )
    exp_map = {(r.month, r.year): r.total or Decimal(0) for r in exp_rows}

    # Batch: 12 oylik maosh (1 query)
    salary_cur_year = _batch_salary_by_month(db, active_grps, cur_year)
    salary_sel_year = salary_cur_year if sel_year == cur_year else _batch_salary_by_month(db, active_grps, sel_year)

    this_income = pay_map.get((cur_month, cur_year), (Decimal(0), 0))[0]
    last_income = pay_map.get((prev_month, prev_year), (Decimal(0), 0))[0]
    change_pct  = float((this_income - last_income) / last_income * 100) if last_income else 0.0

    cur_teacher_salary = salary_cur_year[cur_month]
    cur_external       = exp_map.get((cur_month, cur_year), Decimal(0))
    cur_total_exp      = cur_teacher_salary + cur_external
    cur_net_profit     = this_income - cur_total_exp

    history = []
    for m in range(1, 13):
        income    = pay_map.get((m, sel_year), (Decimal(0), 0))[0]
        pcount    = pay_map.get((m, sel_year), (Decimal(0), 0))[1]
        ext_exp   = exp_map.get((m, sel_year), Decimal(0))
        t_salary  = salary_sel_year[m]
        total_exp = t_salary + ext_exp
        history.append(schemas.MonthlyStats(
            year=sel_year, month=m, total_income=income,
            payment_count=pcount,
            active_students=active_students,
            active_groups=active_groups_count,
            teacher_salary=t_salary,
            external_expenses=ext_exp,
            total_expenses=total_exp,
            net_profit=income - total_exp,
        ))

    return schemas.StatsOverview(
        total_students=total_students,
        active_students=active_students,
        total_groups=total_groups,
        active_groups=active_groups_count,
        this_month_income=this_income,
        last_month_income=last_income,
        income_change_pct=change_pct,
        teacher_salary=cur_teacher_salary,
        external_expenses=cur_external,
        total_expenses=cur_total_exp,
        net_profit=cur_net_profit,
        monthly_history=history,
    )


# ── Teacher salaries breakdown ────────────────────────────────────────────────

@app.get("/stats/teacher-salaries")
def teacher_salaries_breakdown(
    month: int = Query(..., ge=1, le=12),
    year: int = Query(..., ge=2020),
    db: Session = Depends(get_db),
    _: models.User = Depends(require_admin),
):
    """
    Per-teacher salary breakdown based on attendance.
    Formula: teacher_pay_per_student × attended_lessons  (per-lesson rate per student)
    """
    groups = db.query(models.Group).filter(models.Group.is_active == True).all()
    teachers: dict = {}
    grand_total = Decimal(0)

    for g in groups:
        group_salary, student_details = _group_salary(db, g, month, year)
        if group_salary == 0 and not student_details:
            continue

        grand_total += group_salary
        tid = g.teacher_id  # may be None
        if tid not in teachers:
            if tid is not None and g.teacher:
                tname = g.teacher.full_name or g.teacher.username
            else:
                tname = "O'qituvchi tayinlanmagan"
            teachers[tid] = {
                "teacher_id": tid,
                "teacher_name": tname,
                "groups": [],
                "total_salary": Decimal(0),
            }
        teachers[tid]["groups"].append({
            "group_id": g.id,
            "group_name": g.name,
            "stage": g.stage or "foundation",
            "teacher_pay_per_student": float(pay_per),
            "students": student_details,
            "total_attended": sum(s["attended"] for s in student_details),
            "group_salary": float(group_salary),
        })
        teachers[tid]["total_salary"] += group_salary

    result = sorted(
        [
            {**t, "total_salary": float(t["total_salary"])}
            for t in teachers.values()
        ],
        key=lambda t: t["total_salary"],
        reverse=True,
    )
    return {
        "month": month,
        "year": year,
        "total_teacher_salary": float(grand_total),
        "teachers": result,
    }


# ── Teacher: My dashboard ─────────────────────────────────────────────────────

@app.get("/teacher/dashboard")
def teacher_my_dashboard(
    month: int = Query(None, ge=1, le=12),
    year:  int = Query(None, ge=2020),
    db: Session = Depends(get_db),
    actor: models.User = Depends(require_auth),
):
    """
    Teacher sees their own groups + estimated salary for the given month.
    Admin/metodist can also call this for any teacher_id (optional query param).
    """
    sel_month = month or date.today().month
    sel_year  = year  or date.today().year

    groups = (
        db.query(models.Group)
        .options(selectinload(models.Group.members))
        .filter(models.Group.teacher_id == actor.id, models.Group.is_active == True)
        .order_by(models.Group.name)
        .all()
    )
    if not groups:
        return {"teacher_id": actor.id, "teacher_name": actor.full_name or actor.username,
                "month": sel_month, "year": sel_year, "total_groups": 0,
                "total_students": 0, "total_salary": 0.0, "groups": []}

    group_ids = [g.id for g in groups]

    # Batch 1: all-time completed lessons per group (1 query)
    comp_rows = (
        db.query(models.Attendance.group_id,
                 func.count(func.distinct(models.Attendance.lesson_date)).label('cnt'))
        .filter(models.Attendance.group_id.in_(group_ids))
        .group_by(models.Attendance.group_id)
        .all()
    )
    completed_map = {r.group_id: r.cnt for r in comp_rows}

    # Batch 2: this-month lessons per group (1 query)
    mless_rows = (
        db.query(models.Attendance.group_id,
                 func.count(func.distinct(models.Attendance.lesson_date)).label('cnt'))
        .filter(
            models.Attendance.group_id.in_(group_ids),
            extract('month', models.Attendance.lesson_date) == sel_month,
            extract('year',  models.Attendance.lesson_date) == sel_year,
        )
        .group_by(models.Attendance.group_id)
        .all()
    )
    month_lessons_map = {r.group_id: r.cnt for r in mless_rows}

    # Batch 3: all attendance for salary (1 query)
    att_rows = (
        db.query(models.Attendance.group_id, models.Attendance.student_id,
                 func.count(models.Attendance.id).label('cnt'))
        .filter(
            models.Attendance.group_id.in_(group_ids),
            models.Attendance.is_present == True,
            extract('month', models.Attendance.lesson_date) == sel_month,
            extract('year',  models.Attendance.lesson_date) == sel_year,
        )
        .group_by(models.Attendance.group_id, models.Attendance.student_id)
        .all()
    )
    # {group_id: {student_id: attended_count}}
    att_map: dict = {}
    for r in att_rows:
        att_map.setdefault(r.group_id, {})[r.student_id] = r.cnt

    # Batch 4: student names (1 query)
    all_student_ids = [m.student_id for g in groups for m in g.members]
    name_rows = (
        db.query(models.Student.id, models.Student.full_name)
        .filter(models.Student.id.in_(all_student_ids))
        .all()
    )
    student_name_map = {r.id: r.full_name for r in name_rows}

    total_salary   = Decimal(0)
    total_students = 0
    groups_out     = []

    for g in groups:
        stage     = g.stage or 'foundation'
        total_less = schemas.STAGE_TOTAL_LESSONS.get(stage, 24)
        completed  = completed_map.get(g.id, 0)
        pct        = round(completed / total_less * 100, 1) if total_less > 0 else 0.0
        pay_per    = Decimal(str(g.teacher_pay_per_student or 0))

        group_salary    = Decimal(0)
        student_salaries = []
        g_att = att_map.get(g.id, {})

        for mem in g.members:
            attended = g_att.get(mem.student_id, 0)
            share    = (pay_per * Decimal(str(attended))).quantize(Decimal('1'))
            group_salary += share
            student_salaries.append({
                "student_id":   mem.student_id,
                "student_name": student_name_map.get(mem.student_id, ""),
                "attended":     attended,
                "salary_share": float(share),
            })

        total_salary   += group_salary
        total_students += len(g.members)

        groups_out.append({
            "id":                      g.id,
            "name":                    g.name,
            "stage":                   stage,
            "schedule":                g.schedule or "",
            "start_date":              str(g.start_date) if g.start_date else None,
            "student_count":           len(g.members),
            "total_lessons":           total_less,
            "completed_lessons":       completed,
            "progress_pct":            pct,
            "teacher_pay_per_student": float(pay_per),
            "month_salary":            float(group_salary),
            "month_lessons_held":      month_lessons_map.get(g.id, 0),
            "student_salaries":        student_salaries,
        })

    return {
        "teacher_id":    actor.id,
        "teacher_name":  actor.full_name or actor.username,
        "month":         sel_month,
        "year":          sel_year,
        "total_groups":  len(groups_out),
        "total_students": total_students,
        "total_salary":  float(total_salary),
        "groups":        groups_out,
    }


# ── Expenses endpoints ────────────────────────────────────────────────────────

@app.get("/expenses", response_model=List[schemas.ExpenseRead])
def list_expenses(
    month: Optional[int] = Query(None),
    year: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    _: models.User = Depends(require_admin),
):
    q = db.query(models.Expense)
    if month is not None:
        q = q.filter(models.Expense.month == month)
    if year is not None:
        q = q.filter(models.Expense.year == year)
    return q.order_by(models.Expense.year.desc(), models.Expense.month.desc(), models.Expense.id.desc()).all()


@app.post("/expenses", response_model=schemas.ExpenseRead, status_code=201)
def create_expense(payload: schemas.ExpenseCreate, db: Session = Depends(get_db), _: models.User = Depends(require_hunter)):
    exp = models.Expense(**payload.dict())
    db.add(exp)
    db.commit()
    db.refresh(exp)
    return exp


@app.put("/expenses/{expense_id}", response_model=schemas.ExpenseRead)
def update_expense(expense_id: int, payload: schemas.ExpenseUpdate, db: Session = Depends(get_db), _: models.User = Depends(require_admin)):
    exp = db.query(models.Expense).filter(models.Expense.id == expense_id).first()
    if not exp:
        raise HTTPException(status_code=404, detail="Xarajat topilmadi")
    for k, v in payload.dict(exclude_unset=True).items():
        setattr(exp, k, v)
    db.commit()
    db.refresh(exp)
    return exp


@app.delete("/expenses/{expense_id}", status_code=204)
def delete_expense(expense_id: int, db: Session = Depends(get_db), _: models.User = Depends(require_admin)):
    exp = db.query(models.Expense).filter(models.Expense.id == expense_id).first()
    if not exp:
        raise HTTPException(status_code=404, detail="Xarajat topilmadi")
    db.delete(exp)
    db.commit()


MONTHS_UZ = ['','Yanvar','Fevral','Mart','Aprel','May','Iyun',
             'Iyul','Avgust','Sentyabr','Oktyabr','Noyabr','Dekabr']

@app.get("/payments/{payment_id}/receipt", response_class=HTMLResponse)
def payment_receipt(
    payment_id: int,
    _token: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    _: models.User = Depends(require_admin_download),
):
    p = db.query(models.Payment).filter(models.Payment.id == payment_id).first()
    if not p:
        raise HTTPException(status_code=404, detail="To'lov topilmadi")

    student = p.student
    group   = p.group
    month_name = MONTHS_UZ[p.month] if 1 <= p.month <= 12 else str(p.month)
    paid_date  = p.paid_at.strftime('%d.%m.%Y %H:%M') if p.paid_at else '—'
    amount_fmt = f"{int(p.amount):,}".replace(',', ' ')

    # Foydalanuvchi kiritgan matnlar HTML kontekstiga qo'yilishidan oldin escape qilinadi
    # (aks holda talaba ismi / izoh orqali saqlanadigan XSS bo'lardi — chek admin brauzerida ochiladi).
    student_name = html.escape(student.full_name) if student and student.full_name else '—'
    group_name   = html.escape(group.name) if group and group.name else '—'
    notes_html   = html.escape(p.notes) if p.notes else ''

    page = f"""<!DOCTYPE html>
<html lang="uz">
<head>
<meta charset="UTF-8">
<title>Chek #{p.id:05d}</title>
<style>
  * {{ margin:0; padding:0; box-sizing:border-box; }}
  body {{ font-family: 'Segoe UI', Arial, sans-serif; background:#f5f5f5; display:flex;
         justify-content:center; align-items:flex-start; min-height:100vh; padding:20px; }}
  .receipt {{ background:#fff; width:340px; padding:28px 24px 32px;
              border-radius:12px; box-shadow:0 4px 24px rgba(0,0,0,.12); }}
  .brand {{ text-align:center; margin-bottom:20px; }}
  .brand-name {{ font-size:22px; font-weight:800; color:#1d4ed8; letter-spacing:-.5px; }}
  .brand-sub  {{ font-size:11px; color:#737373; margin-top:2px; }}
  .divider {{ border:none; border-top:1px dashed #d4d4d4; margin:16px 0; }}
  .receipt-title {{ text-align:center; font-size:13px; color:#737373; margin-bottom:16px; }}
  .receipt-num {{ font-size:11px; color:#a3a3a3; text-align:center; margin-top:2px; }}
  .row {{ display:flex; justify-content:space-between; align-items:flex-start;
          margin-bottom:10px; gap:8px; }}
  .row .label {{ font-size:12px; color:#737373; white-space:nowrap; }}
  .row .val   {{ font-size:12px; color:#0a0a0a; font-weight:500; text-align:right; }}
  .amount-box {{ background:#eff6ff; border-radius:8px; padding:14px 16px;
                 text-align:center; margin:16px 0; }}
  .amount-box .amt {{ font-size:26px; font-weight:800; color:#1d4ed8; letter-spacing:-1px; }}
  .amount-box .cur {{ font-size:13px; color:#3b82f6; font-weight:600; margin-left:4px; }}
  .status {{ display:inline-flex; align-items:center; gap:6px; background:#dcfce7;
             color:#16a34a; border-radius:20px; padding:4px 14px;
             font-size:12px; font-weight:700; margin:0 auto; display:block; text-align:center; }}
  .footer {{ text-align:center; font-size:10px; color:#a3a3a3; margin-top:20px; line-height:1.5; }}
  @media print {{
    body {{ background:#fff; padding:0; }}
    .receipt {{ box-shadow:none; border-radius:0; }}
    .no-print {{ display:none !important; }}
  }}
  .print-btn {{ display:block; width:100%; margin-top:20px; padding:10px;
                background:#1d4ed8; color:#fff; border:none; border-radius:8px;
                font-size:14px; font-weight:600; cursor:pointer; }}
  .print-btn:hover {{ background:#1e40af; }}
</style>
</head>
<body>
<div class="receipt">
  <div class="brand">
    <div class="brand-name">IT Hub</div>
    <div class="brand-sub">O'quv markazi</div>
  </div>

  <hr class="divider">
  <div class="receipt-title">TO'LOV CHEKI</div>
  <div class="receipt-num"># {p.id:05d}</div>
  <hr class="divider">

  <div class="row">
    <span class="label">O'quvchi</span>
    <span class="val">{student_name}</span>
  </div>
  <div class="row">
    <span class="label">Guruh</span>
    <span class="val">{group_name}</span>
  </div>
  <div class="row">
    <span class="label">Oy</span>
    <span class="val">{month_name} {p.year}</span>
  </div>
  <div class="row">
    <span class="label">Sana</span>
    <span class="val">{paid_date}</span>
  </div>
  {f'<div class="row"><span class="label">Izoh</span><span class="val">{notes_html}</span></div>' if notes_html else ''}

  <hr class="divider">

  <div class="amount-box">
    <span class="amt">{amount_fmt}</span>
    <span class="cur">so'm</span>
  </div>

  <div class="status">&#10003; TO'LANGAN</div>

  <div class="footer">
    Ushbu chek IT Hub o'quv markazi tomonidan<br>
    rasmiy to'lov tasdiqi sifatida berilgan.
  </div>

  <button class="print-btn no-print" onclick="window.print()">&#128438; Chop etish</button>
</div>
</body>
</html>"""
    return HTMLResponse(content=page)


@app.get("/stats/export/excel")
def export_payments_excel(
    month: Optional[int] = Query(None),
    year: Optional[int] = Query(None),
    _token: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    _: models.User = Depends(require_admin_download),
):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl kutubxonasi o'rnatilmagan")

    q = db.query(models.Payment)
    if month:
        q = q.filter(models.Payment.month == month)
    if year:
        q = q.filter(models.Payment.year == year)
    payments = q.order_by(models.Payment.paid_at.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "To'lovlar"

    headers = ["#", "Talaba", "Guruh", "Miqdor (so'm)", "Oy", "Yil", "To'langan sana", "Izoh"]
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    month_names = {1:"Yanvar",2:"Fevral",3:"Mart",4:"Aprel",5:"May",6:"Iyun",
                   7:"Iyul",8:"Avgust",9:"Sentyabr",10:"Oktyabr",11:"Noyabr",12:"Dekabr"}

    for i, p in enumerate(payments, 1):
        ws.append([
            i,
            p.student.full_name if p.student else "",
            p.group.name if p.group else "",
            float(p.amount),
            month_names.get(p.month, p.month),
            p.year,
            p.paid_at.strftime("%d.%m.%Y %H:%M") if p.paid_at else "",
            p.notes or "",
        ])

    # Column widths
    col_widths = [5, 25, 20, 18, 12, 8, 20, 25]
    for col, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w

    # Total row
    total = sum(float(p.amount) for p in payments)
    last_row = len(payments) + 2
    ws.cell(row=last_row, column=1, value="JAMI:")
    ws.cell(row=last_row, column=1).font = Font(bold=True)
    ws.cell(row=last_row, column=4, value=total)
    ws.cell(row=last_row, column=4).font = Font(bold=True)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"payments_{year or 'all'}_{month or 'all'}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@app.get("/health")
def health_check():
    return {"status": "ok", "version": "3.0.0"}


# ── Teacher: Today's groups for quick attendance ───────────────────────────────

# Uzbek day abbreviations → Python weekday (Mon=0)
_DAY_WORDS = {
    'du': 0, 'dush': 0,
    'se': 1, 'sesh': 1,
    'chor': 2,
    'pay': 3,
    'ju': 4, 'jum': 4,
    'shan': 5,
    'yak': 6,
}


def _schedule_has_today(schedule: str, today_wd: int) -> bool:
    if not schedule:
        return False
    import re
    tokens = re.split(r'[\s,\-/]+', schedule.lower())
    return any(_DAY_WORDS.get(t) == today_wd for t in tokens)


@app.get("/attendance/today")
def today_groups(
    target_date: Optional[date] = Query(None),  # admin can pick any date
    db: Session = Depends(get_db),
    actor: models.User = Depends(require_auth),
):
    """
    Groups for quick attendance.
    - Admin/metodist/hunter: ALL active groups, any date (default today).
    - Teacher: only their groups scheduled for today.
    """
    ref_date = target_date or date.today()
    ref_wd = ref_date.weekday()
    is_admin_or_metodist = actor.role in (UserRole.admin.value, UserRole.metodist.value,
                                          UserRole.hunter.value, UserRole.sales.value)

    q = (
        db.query(models.Group)
        .options(joinedload(models.Group.teacher), selectinload(models.Group.members))
        .filter(models.Group.is_active == True)
    )
    if not is_admin_or_metodist:
        q = q.filter(models.Group.teacher_id == actor.id)
    groups = q.order_by(models.Group.name).all()

    # Batch: barcha guruhlar uchun shu kunda davomat soni (1 query)
    group_ids = [g.id for g in groups]
    taken_rows = (
        db.query(models.Attendance.group_id,
                 func.count(models.Attendance.id).label('cnt'))
        .filter(
            models.Attendance.group_id.in_(group_ids),
            models.Attendance.lesson_date == ref_date,
        )
        .group_by(models.Attendance.group_id)
        .all()
    )
    taken_map = {r.group_id: r.cnt for r in taken_rows}

    result = []
    for g in groups:
        if not is_admin_or_metodist and not _schedule_has_today(g.schedule, ref_wd):
            continue
        taken = taken_map.get(g.id, 0)
        result.append({
            "id": g.id,
            "name": g.name,
            "stage": g.stage or 'foundation',
            "schedule": g.schedule,
            "lesson_time": g.lesson_time,
            "teacher_name": g.teacher.full_name or g.teacher.username if g.teacher else None,
            "student_count": len(g.members),
            "attendance_taken": taken > 0,
            "attendance_count": taken,
            "date": str(ref_date),
        })

    return result


# ── Finance: monthly summary ───────────────────────────────────────────────────

@app.get("/finance/monthly")
def finance_monthly(
    month: int = Query(..., ge=1, le=12),
    year: int = Query(..., ge=2020),
    db: Session = Depends(get_db),
    _: models.User = Depends(require_admin),
):
    """
    Monthly finance summary.
    Income formula: owed = tariff_price (full monthly price, attendance-independent).
    Students pay for the month regardless of how many lessons they attended.
    Teacher salary is attendance-based (handled separately).
    """
    groups = db.query(models.Group).filter(models.Group.is_active == True).all()
    result = []
    total_expected = Decimal(0)
    total_actual = Decimal(0)

    for g in groups:
        # Actual payments for this group this month
        actual = db.query(func.sum(models.Payment.amount)).filter(
            models.Payment.group_id == g.id,
            models.Payment.month == month,
            models.Payment.year == year,
        ).scalar() or Decimal(0)

        # Total lessons held for this group this month (informational)
        total_lessons_held = db.query(
            func.count(func.distinct(models.Attendance.lesson_date))
        ).filter(
            models.Attendance.group_id == g.id,
            extract('month', models.Attendance.lesson_date) == month,
            extract('year', models.Attendance.lesson_date) == year,
        ).scalar() or 0

        expected = Decimal(0)
        student_details = []

        for m in g.members:
            # Count lessons student attended this month (informational only)
            attended = db.query(func.count(models.Attendance.id)).filter(
                models.Attendance.group_id == g.id,
                models.Attendance.student_id == m.student_id,
                models.Attendance.is_present == True,
                extract('month', models.Attendance.lesson_date) == month,
                extract('year', models.Attendance.lesson_date) == year,
            ).scalar() or 0

            if m.tariff:
                price = Decimal(str(m.tariff.price))
                tariff_name = m.tariff.name
            elif g.course_price and Decimal(str(g.course_price)) > 0:
                # Fallback: use group's course_price if student has no individual tariff
                price = Decimal(str(g.course_price))
                tariff_name = "Guruh narxi"
            else:
                price = Decimal(0)
                tariff_name = None

            if price > 0:
                price = core_calc.apply_special_discounts(
                    price, core_calc.active_special_discounts(db, m.student_id),
                    g.id, month, year,
                )

            if price > 0:
                owed = price.quantize(Decimal('1'))
                tariff_price = float(price)
                paid_amount = db.query(func.sum(models.Payment.amount)).filter(
                    models.Payment.group_id == g.id,
                    models.Payment.student_id == m.student_id,
                    models.Payment.month == month,
                    models.Payment.year == year,
                ).scalar() or Decimal(0)
                is_paid = paid_amount >= owed
                expected += owed
            else:
                owed = Decimal(0)
                tariff_price = 0.0
                is_paid = None  # no obligation

            student_details.append({
                "student_id": m.student_id,
                "student_name": m.student.full_name,
                "phone": m.student.phone1,
                "tariff_name": tariff_name,
                "tariff_price": tariff_price,
                "attended": attended,
                "total_lessons_held": total_lessons_held,
                "owed": float(owed),
                "is_paid": is_paid,
            })

        # Unpaid = has tariff, hasn't fully paid
        unpaid_students = [s for s in student_details if s["is_paid"] is False]

        total_expected += expected
        total_actual += actual
        result.append({
            "group_id": g.id,
            "group_name": g.name,
            "student_count": len(g.members),
            "expected": float(expected),
            "actual": float(actual),
            "deficit": float(expected - actual),
            "unpaid_count": len(unpaid_students),
            "unpaid_students": unpaid_students,
            "all_students": student_details,
        })

    return {
        "month": month,
        "year": year,
        "total_expected": float(total_expected),
        "total_actual": float(total_actual),
        "total_deficit": float(total_expected - total_actual),
        "groups": result,
    }


# ── Attendance ─────────────────────────────────────────────────────────────────

@app.get("/groups/{group_id}/camera-attendance")
def group_camera_attendance(
    group_id: int,
    days: int = Query(7, ge=1, le=90),
    db: Session = Depends(get_db),
    _: models.User = Depends(require_staff),
):
    """Guruh o'quvchilarining kamera keldi/ketdi tarixi (so'nggi N kun)."""
    from datetime import timedelta
    since = datetime.utcnow() - timedelta(days=days)

    member_ids = [
        gs.student_id for gs in
        db.query(models.GroupStudent.student_id)
        .filter(models.GroupStudent.group_id == group_id)
        .all()
    ]
    if not member_ids:
        return []

    rows = (
        db.query(models.CameraAttendance, models.Student.full_name)
        .join(models.Student, models.Student.id == models.CameraAttendance.student_id)
        .filter(
            models.CameraAttendance.student_id.in_(member_ids),
            models.CameraAttendance.detected_at >= since,
        )
        .order_by(models.CameraAttendance.detected_at.desc())
        .limit(500)
        .all()
    )
    return [
        {
            "id":           r.CameraAttendance.id,
            "student_id":   r.CameraAttendance.student_id,
            "student_name": r.full_name,
            "event_type":   r.CameraAttendance.event_type,
            "detected_at":  r.CameraAttendance.detected_at.isoformat(),
        }
        for r in rows
    ]


@app.get("/groups/{group_id}/attendance")
def get_attendance(
    group_id: int,
    month: int = Query(..., ge=1, le=12),
    year: int = Query(..., ge=2020),
    db: Session = Depends(get_db),
    actor: models.User = Depends(require_auth),
):
    """Return all attendance records for a group in given month/year."""
    group = db.query(models.Group).filter(models.Group.id == group_id).first()
    if not group:
        raise HTTPException(status_code=404, detail="Guruh topilmadi")

    # Teacher faqat o'z guruhini ko'ra oladi
    if actor.role == UserRole.teacher.value and group.teacher_id != actor.id:
        raise HTTPException(status_code=403, detail="Bu guruh sizga tegishli emas")

    members = (
        db.query(models.GroupStudent)
        .filter(models.GroupStudent.group_id == group_id)
        .all()
    )
    student_ids = [m.student_id for m in members]

    records = (
        db.query(models.Attendance)
        .filter(
            models.Attendance.group_id == group_id,
            extract("month", models.Attendance.lesson_date) == month,
            extract("year", models.Attendance.lesson_date) == year,
        )
        .all()
    )

    # Unique lesson dates this month
    dates = sorted(set(r.lesson_date for r in records))

    # Build lookup: {student_id: {date: is_present}}
    lookup = {}
    for r in records:
        lookup.setdefault(r.student_id, {})[r.lesson_date] = r.is_present

    students_data = []
    for m in members:
        s = m.student
        row = {
            "student_id": s.id,
            "student_name": s.full_name,
            "phone": s.phone1,
            "joined_at": m.joined_at.isoformat(),
            "dates": {str(d): lookup.get(s.id, {}).get(d) for d in dates},
            "present_count": sum(1 for d in dates if lookup.get(s.id, {}).get(d) is True),
            "absent_count": sum(1 for d in dates if lookup.get(s.id, {}).get(d) is False),
            "total_lessons": len(dates),
        }
        students_data.append(row)

    return {
        "group_id": group_id,
        "group_name": group.name,
        "teacher_name": group.teacher.full_name or group.teacher.username if group.teacher else None,
        "schedule": group.schedule,
        "month": month,
        "year": year,
        "dates": [str(d) for d in dates],
        "students": students_data,
    }


@app.post("/groups/{group_id}/attendance/{lesson_date}")
def save_attendance(
    group_id: int,
    lesson_date: str,
    payload: List[dict],
    db: Session = Depends(get_db),
    actor: models.User = Depends(require_attendance_editor),
):
    """Save attendance for a specific date. payload: [{student_id, is_present}]"""
    try:
        d = date.fromisoformat(lesson_date)
    except ValueError:
        raise HTTPException(status_code=400, detail="Sana formati xato (YYYY-MM-DD)")

    group = db.query(models.Group).filter(models.Group.id == group_id).first()
    if not group:
        raise HTTPException(status_code=404, detail="Guruh topilmadi")

    # Teacher faqat o'z guruhida yo'qlama qila oladi
    if actor.role == UserRole.teacher.value and group.teacher_id != actor.id:
        raise HTTPException(status_code=403, detail="Bu guruh sizga tegishli emas")

    newly_absent: list[int] = []
    newly_present: list[int] = []
    for item in payload:
        sid = item.get("student_id")
        present = item.get("is_present", True)
        existing = (
            db.query(models.Attendance)
            .filter(
                models.Attendance.group_id == group_id,
                models.Attendance.student_id == sid,
                models.Attendance.lesson_date == d,
            )
            .first()
        )
        if existing:
            if existing.is_present and not present:
                newly_absent.append(sid)
            elif not existing.is_present and present:
                newly_present.append(sid)
            existing.is_present = present
        else:
            if not present:
                newly_absent.append(sid)
            else:
                newly_present.append(sid)
            db.add(models.Attendance(
                group_id=group_id,
                student_id=sid,
                lesson_date=d,
                is_present=present,
            ))

    # Yo'q deb belgilangan bolalarning ota-onalariga bildirishnoma
    for sid in newly_absent:
        notify_parents_of_student(
            db, student_id=sid, ntype="attendance",
            title="Bola darsga kelmadi",
            body=f"{d:%d.%m.%Y} — {group.name} darsida qatnashmadi",
        )
    # "Keldi" belgilanganlarga — Telegram (biriktirilgan user ID bo'lsa)
    if newly_present:
        students_map = {
            s.id: s for s in db.query(models.Student)
            .filter(models.Student.id.in_(newly_present)).all()
        }
        for sid in newly_present:
            s = students_map.get(sid)
            if s and s.telegram_user_id:
                notify_student_telegram(
                    s,
                    f"📚 <b>{s.full_name}</b> bugun darsda qatnashdi\n"
                    f"👥 {group.name}\n📅 {d.strftime('%d.%m.%Y')}",
                )
    db.commit()
    return {"saved": len(payload), "date": lesson_date}


@app.delete("/groups/{group_id}/attendance/{lesson_date}")
def delete_attendance_date(
    group_id: int,
    lesson_date: str,
    db: Session = Depends(get_db),
    actor: models.User = Depends(require_attendance_editor),
):
    """Delete all attendance records for a specific date."""
    try:
        d = date.fromisoformat(lesson_date)
    except ValueError:
        raise HTTPException(status_code=400, detail="Sana formati xato")

    # Teacher faqat o'z guruhida
    if actor.role == UserRole.teacher.value:
        group = db.query(models.Group).filter(models.Group.id == group_id).first()
        if not group:
            raise HTTPException(status_code=404, detail="Guruh topilmadi")
        if group.teacher_id != actor.id:
            raise HTTPException(status_code=403, detail="Bu guruh sizga tegishli emas")

    db.query(models.Attendance).filter(
        models.Attendance.group_id == group_id,
        models.Attendance.lesson_date == d,
    ).delete()
    db.commit()
    return {"deleted": True, "date": lesson_date}


# ── Student visits — Notifications (keldi/ketdi) ─────────────────────────────

def _visit_read(v: models.StudentVisit) -> schemas.StudentVisitRead:
    return schemas.StudentVisitRead(
        id=v.id, student_id=v.student_id,
        student_name=v.student.full_name if v.student else None,
        student_photo=v.student.photo if v.student else None,
        kind=v.kind,
        noted_by_name=(v.noted_by.full_name or v.noted_by.username) if v.noted_by else None,
        telegram_sent=v.telegram_sent, telegram_error=v.telegram_error,
        created_at=v.created_at,
    )


@app.get("/visits", response_model=List[schemas.StudentVisitRead])
def list_visits(
    visit_date: Optional[date] = Query(None),
    student_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    _: models.User = Depends(require_crm_access),
):
    q = db.query(models.StudentVisit).options(
        joinedload(models.StudentVisit.student),
        joinedload(models.StudentVisit.noted_by),
    )
    if student_id:
        q = q.filter(models.StudentVisit.student_id == student_id)
    else:
        d = visit_date or date.today()
        start = datetime(d.year, d.month, d.day)
        q = q.filter(models.StudentVisit.created_at >= start,
                     models.StudentVisit.created_at < start + timedelta(days=1))
    return [_visit_read(v) for v in q.order_by(models.StudentVisit.created_at.desc()).limit(200).all()]


@app.post("/visits", response_model=schemas.StudentVisitRead, status_code=201)
def create_visit(
    payload: schemas.StudentVisitCreate,
    db: Session = Depends(get_db),
    actor: models.User = Depends(require_crm_access),
):
    """Keldi/ketdi belgilash + biriktirilgan Telegram ID'ga xabar (rasmi bo'lsa rasm bilan)."""
    s = db.query(models.Student).filter(models.Student.id == payload.student_id).first()
    if not s:
        raise HTTPException(status_code=404, detail="Talaba topilmadi")

    v = models.StudentVisit(student_id=s.id, kind=payload.kind, noted_by_id=actor.id)

    time_str = (datetime.utcnow() + timedelta(hours=5)).strftime("%H:%M")  # Toshkent vaqti
    if payload.kind == "arrived":
        caption = f"✅ <b>{s.full_name}</b> o'quv markazga keldi\n🕐 {time_str}"
    else:
        caption = f"🏠 <b>{s.full_name}</b> o'quv markazdan ketdi\n🕐 {time_str}"
    ok, err = notify_student_telegram(s, caption)
    v.telegram_sent = ok
    v.telegram_error = err

    db.add(v)
    db.flush()
    write_audit(db, entity_type="student_visit", entity_id=v.id, action="create",
                changed_by_id=actor.id,
                new_value={"student_id": s.id, "kind": payload.kind, "telegram_sent": ok})
    db.commit()
    db.refresh(v)
    return _visit_read(v)


@app.post("/students/{student_id}/photo", response_model=schemas.StudentRead)
async def upload_student_photo(
    student_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    actor: models.User = Depends(require_crm_access),
):
    s = db.query(models.Student).filter(models.Student.id == student_id).first()
    if not s:
        raise HTTPException(status_code=404, detail="Talaba topilmadi")
    if file.content_type not in ALLOWED_IMAGE_TYPES:
        raise HTTPException(400, "Faqat JPEG, PNG, WebP yoki GIF rasm yuklash mumkin")
    contents = await file.read()
    if len(contents) > MAX_AVATAR_BYTES:
        raise HTTPException(400, "Fayl hajmi 5 MB dan oshmasin")
    ext = {
        "image/jpeg": "jpg", "image/png": "png",
        "image/webp": "webp", "image/gif": "gif",
    }[file.content_type]
    filename = f"{student_id}.{ext}"
    with open(os.path.join(STUDENTS_DIR, filename), "wb") as f:
        f.write(contents)
    s.photo = f"students/{filename}"
    db.commit()
    db.refresh(s)
    return _student_read(s)


# ── Special chegirmalar ───────────────────────────────────────────────────────

def _special_read(d: models.SpecialDiscount) -> schemas.SpecialDiscountRead:
    return schemas.SpecialDiscountRead(
        id=d.id, student_id=d.student_id,
        student_name=d.student.full_name if d.student else None,
        group_id=d.group_id,
        group_name=d.group.name if d.group else None,
        kind=d.kind, amount=d.amount, month=d.month, year=d.year,
        reason=d.reason, is_active=d.is_active,
        created_by_name=(d.created_by.full_name or d.created_by.username) if d.created_by else None,
        created_at=d.created_at,
    )


@app.get("/special-discounts", response_model=List[schemas.SpecialDiscountRead])
def list_special_discounts(
    student_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    _: models.User = Depends(require_hunter),
):
    q = db.query(models.SpecialDiscount).options(
        joinedload(models.SpecialDiscount.student),
        joinedload(models.SpecialDiscount.group),
        joinedload(models.SpecialDiscount.created_by),
    )
    if student_id:
        q = q.filter(models.SpecialDiscount.student_id == student_id)
    return [_special_read(d) for d in q.order_by(models.SpecialDiscount.created_at.desc()).all()]


@app.post("/special-discounts", response_model=schemas.SpecialDiscountRead, status_code=201)
def create_special_discount(
    payload: schemas.SpecialDiscountCreate,
    db: Session = Depends(get_db),
    actor: models.User = Depends(require_hunter),
):
    s = db.query(models.Student).filter(models.Student.id == payload.student_id).first()
    if not s:
        raise HTTPException(status_code=404, detail="Talaba topilmadi")
    if payload.group_id:
        if not db.query(models.Group).filter(models.Group.id == payload.group_id).first():
            raise HTTPException(status_code=404, detail="Guruh topilmadi")
    if payload.kind == "free_month":
        if not payload.month or not payload.year:
            raise HTTPException(status_code=400, detail="Bepul oy uchun oy va yil kiritilishi kerak")
    if payload.kind == "monthly" and not payload.amount:
        raise HTTPException(status_code=400, detail="Oylik chegirma uchun summa kiritilishi kerak")

    d = models.SpecialDiscount(
        student_id=payload.student_id,
        group_id=payload.group_id,
        kind=payload.kind,
        amount=payload.amount if payload.kind == "monthly" else None,
        month=payload.month if payload.kind == "free_month" else None,
        year=payload.year if payload.kind == "free_month" else None,
        reason=payload.reason,
        created_by_id=actor.id,
    )
    db.add(d)
    db.flush()
    write_audit(db, entity_type="special_discount", entity_id=d.id, action="create",
                changed_by_id=actor.id,
                new_value={"student_id": d.student_id, "group_id": d.group_id, "kind": d.kind,
                           "amount": str(d.amount) if d.amount else None,
                           "month": d.month, "year": d.year, "reason": d.reason})
    db.commit()
    db.refresh(d)
    return _special_read(d)


@app.patch("/special-discounts/{discount_id}", response_model=schemas.SpecialDiscountRead)
def update_special_discount(
    discount_id: int,
    payload: schemas.SpecialDiscountUpdate,
    db: Session = Depends(get_db),
    actor: models.User = Depends(require_hunter),
):
    d = db.query(models.SpecialDiscount).filter(models.SpecialDiscount.id == discount_id).first()
    if not d:
        raise HTTPException(status_code=404, detail="Chegirma topilmadi")
    changes = payload.dict(exclude_unset=True)
    old = {k: str(getattr(d, k)) for k in changes}
    for k, v in changes.items():
        setattr(d, k, v)
    write_audit(db, entity_type="special_discount", entity_id=d.id, action="update",
                changed_by_id=actor.id, old_value=old,
                new_value={k: str(v) for k, v in changes.items()})
    db.commit()
    db.refresh(d)
    return _special_read(d)


@app.delete("/special-discounts/{discount_id}", status_code=204)
def delete_special_discount(discount_id: int, db: Session = Depends(get_db), actor: models.User = Depends(require_hunter)):
    d = db.query(models.SpecialDiscount).filter(models.SpecialDiscount.id == discount_id).first()
    if not d:
        raise HTTPException(status_code=404, detail="Chegirma topilmadi")
    write_audit(db, entity_type="special_discount", entity_id=d.id, action="delete",
                changed_by_id=actor.id,
                old_value={"student_id": d.student_id, "kind": d.kind,
                           "amount": str(d.amount) if d.amount else None,
                           "month": d.month, "year": d.year})
    db.delete(d)
    db.commit()


# ── Holidays (dam olish kunlari) ──────────────────────────────────────────────

def _holiday_read(h: models.Holiday) -> schemas.HolidayRead:
    return schemas.HolidayRead(
        id=h.id, name=h.name, start_date=h.start_date, end_date=h.end_date,
        created_by_name=(h.created_by.full_name or h.created_by.username) if h.created_by else None,
        created_at=h.created_at,
    )


@app.get("/holidays", response_model=List[schemas.HolidayRead])
def list_holidays(
    year: Optional[int] = Query(None, ge=2020),
    db: Session = Depends(get_db),
    _: models.User = Depends(require_auth),
):
    q = db.query(models.Holiday).options(joinedload(models.Holiday.created_by))
    if year:
        q = q.filter(models.Holiday.end_date >= date(year, 1, 1),
                     models.Holiday.start_date <= date(year, 12, 31))
    return [_holiday_read(h) for h in q.order_by(models.Holiday.start_date.desc()).all()]


@app.post("/holidays", response_model=schemas.HolidayRead, status_code=201)
def create_holiday(payload: schemas.HolidayCreate, db: Session = Depends(get_db), actor: models.User = Depends(require_admin)):
    if payload.end_date < payload.start_date:
        raise HTTPException(status_code=400, detail="Tugash sanasi boshlanish sanasidan oldin bo'lishi mumkin emas")
    h = models.Holiday(
        name=payload.name, start_date=payload.start_date, end_date=payload.end_date,
        created_by_id=actor.id,
    )
    db.add(h)
    db.flush()
    write_audit(db, entity_type="holiday", entity_id=h.id, action="create",
                changed_by_id=actor.id,
                new_value={"name": h.name, "start_date": str(h.start_date), "end_date": str(h.end_date)})
    db.commit()
    db.refresh(h)
    return _holiday_read(h)


@app.delete("/holidays/{holiday_id}", status_code=204)
def delete_holiday(holiday_id: int, db: Session = Depends(get_db), actor: models.User = Depends(require_admin)):
    h = db.query(models.Holiday).filter(models.Holiday.id == holiday_id).first()
    if not h:
        raise HTTPException(status_code=404, detail="Dam olish kuni topilmadi")
    write_audit(db, entity_type="holiday", entity_id=h.id, action="delete",
                changed_by_id=actor.id,
                old_value={"name": h.name, "start_date": str(h.start_date), "end_date": str(h.end_date)})
    db.delete(h)
    db.commit()


# ── Homework (uy vazifasi) + Telegram ─────────────────────────────────────────

TELEGRAM_BOT_TOKEN = os.getenv("TELEGRAM_BOT_TOKEN", "")


def send_telegram_message(chat_id: str, text: str):
    """Guruh Telegram chatiga xabar yuboradi. (ok, error) qaytaradi."""
    if not TELEGRAM_BOT_TOKEN:
        return False, "TELEGRAM_BOT_TOKEN sozlanmagan (.env)"
    try:
        import httpx
        r = httpx.post(
            f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendMessage",
            json={"chat_id": chat_id, "text": text, "parse_mode": "HTML"},
            timeout=10,
        )
        body = r.json()
        if r.status_code == 200 and body.get("ok"):
            return True, None
        return False, str(body.get("description", r.text))[:400]
    except Exception as e:
        return False, str(e)[:400]


def send_telegram_photo(chat_id: str, photo_path: str, caption: str):
    """Rasm + izoh yuboradi. (ok, error) qaytaradi."""
    if not TELEGRAM_BOT_TOKEN:
        return False, "TELEGRAM_BOT_TOKEN sozlanmagan (.env)"
    try:
        import httpx
        with open(photo_path, "rb") as f:
            r = httpx.post(
                f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendPhoto",
                data={"chat_id": chat_id, "caption": caption, "parse_mode": "HTML"},
                files={"photo": f},
                timeout=15,
            )
        body = r.json()
        if r.status_code == 200 and body.get("ok"):
            return True, None
        return False, str(body.get("description", r.text))[:400]
    except Exception as e:
        return False, str(e)[:400]


def notify_student_telegram(student: models.Student, caption: str):
    """Talabaga biriktirilgan Telegram ID'ga xabar: rasmi bo'lsa rasm bilan, bo'lmasa matn."""
    if not student.telegram_user_id:
        return False, "Talabaga Telegram user ID biriktirilmagan"
    if student.photo:
        path = os.path.join(UPLOAD_DIR, student.photo)
        if os.path.exists(path):
            return send_telegram_photo(student.telegram_user_id, path, caption)
    return send_telegram_message(student.telegram_user_id, caption)


def _homework_read(hw: models.Homework) -> schemas.HomeworkRead:
    return schemas.HomeworkRead(
        id=hw.id, group_id=hw.group_id,
        group_name=hw.group.name if hw.group else None,
        lesson_id=hw.lesson_id, lesson_number=hw.lesson_number,
        lesson_title=hw.lesson_title, text=hw.text, lesson_date=hw.lesson_date,
        created_by_name=(hw.created_by.full_name or hw.created_by.username) if hw.created_by else None,
        created_at=hw.created_at,
        telegram_sent=hw.telegram_sent, telegram_error=hw.telegram_error,
    )


def _check_group_access(db: Session, group_id: int, actor: models.User) -> models.Group:
    group = db.query(models.Group).filter(models.Group.id == group_id).first()
    if not group:
        raise HTTPException(status_code=404, detail="Guruh topilmadi")
    if actor.role == UserRole.teacher.value and group.teacher_id != actor.id:
        raise HTTPException(status_code=403, detail="Bu guruh sizga tegishli emas")
    return group


@app.get("/groups/{group_id}/next-lesson", response_model=schemas.NextLessonInfo)
def group_next_lesson(group_id: int, db: Session = Depends(get_db), actor: models.User = Depends(require_attendance_editor)):
    """Metodika bo'yicha navbatdagi dars: raqami, nomi va uy vazifasi (lessons jadvalidan)."""
    group = _check_group_access(db, group_id, actor)
    category = group.stage or 'foundation'
    total = group.course.total_lessons if group.course else schemas.STAGE_TOTAL_LESSONS.get(category, 24)
    completed = db.query(func.count(func.distinct(models.Attendance.lesson_date))).filter(
        models.Attendance.group_id == group_id
    ).scalar() or 0
    next_num = completed + 1
    lesson = db.query(models.Lesson).filter(
        models.Lesson.category == category,
        models.Lesson.lesson_number == next_num,
    ).first()
    return schemas.NextLessonInfo(
        lesson_id=lesson.id if lesson else None,
        lesson_number=next_num,
        lesson_title=lesson.title if lesson else None,
        homework=lesson.homework if lesson else None,
        category=category,
        completed_lessons=completed,
        total_lessons=total,
    )


@app.get("/groups/{group_id}/homeworks", response_model=List[schemas.HomeworkRead])
def list_homeworks(group_id: int, db: Session = Depends(get_db), actor: models.User = Depends(require_attendance_editor)):
    _check_group_access(db, group_id, actor)
    hws = (
        db.query(models.Homework)
        .options(joinedload(models.Homework.group), joinedload(models.Homework.created_by))
        .filter(models.Homework.group_id == group_id)
        .order_by(models.Homework.lesson_date.desc(), models.Homework.id.desc())
        .limit(30).all()
    )
    return [_homework_read(hw) for hw in hws]


@app.post("/groups/{group_id}/homeworks", response_model=schemas.HomeworkRead, status_code=201)
def create_homework(
    group_id: int,
    payload: schemas.HomeworkCreate,
    db: Session = Depends(get_db),
    actor: models.User = Depends(require_attendance_editor),
):
    """Uy vazifasini saqlaydi va guruh Telegram chatiga yuboradi."""
    group = _check_group_access(db, group_id, actor)
    lesson_date = payload.lesson_date or date.today()

    hw = models.Homework(
        group_id=group_id,
        lesson_id=payload.lesson_id,
        lesson_number=payload.lesson_number,
        lesson_title=payload.lesson_title,
        text=payload.text,
        lesson_date=lesson_date,
        created_by_id=actor.id,
    )

    # Telegram xabari
    lines = [f"📚 <b>{group.name}</b>"]
    if payload.lesson_number:
        title_part = f" — {payload.lesson_title}" if payload.lesson_title else ""
        lines.append(f"📖 {payload.lesson_number}-dars{title_part}")
    elif payload.lesson_title:
        lines.append(f"📖 {payload.lesson_title}")
    lines.append(f"📅 {lesson_date.strftime('%d.%m.%Y')}")
    lines.append("")
    lines.append("📝 <b>Uy vazifasi:</b>")
    lines.append(payload.text)
    message = "\n".join(lines)

    if group.telegram_chat_id:
        ok, err = send_telegram_message(group.telegram_chat_id, message)
        hw.telegram_sent = ok
        hw.telegram_error = err
    else:
        hw.telegram_sent = False
        hw.telegram_error = "Guruhga Telegram chat ID biriktirilmagan"

    db.add(hw)
    db.flush()
    write_audit(db, entity_type="homework", entity_id=hw.id, action="create",
                changed_by_id=actor.id,
                new_value={"group_id": group_id, "lesson_number": payload.lesson_number,
                           "lesson_date": str(lesson_date), "telegram_sent": hw.telegram_sent})
    db.commit()
    db.refresh(hw)
    return _homework_read(hw)


# ── Akademik: baholar / izohlar / sertifikatlar / tadbirlar ──────────────────

def _teacher_group_ids(db: Session, actor: models.User) -> set:
    return {gid for (gid,) in db.query(models.Group.id).filter(models.Group.teacher_id == actor.id).all()}


def _check_academic_target(db: Session, actor: models.User, student_id: int, group_id: Optional[int]) -> None:
    """Talaba/guruh mavjudligini va teacher faqat o'z guruhida ishlashini tekshiradi."""
    if not db.query(models.Student).filter(models.Student.id == student_id).first():
        raise HTTPException(status_code=404, detail="Talaba topilmadi")
    if group_id and not db.query(models.Group).filter(models.Group.id == group_id).first():
        raise HTTPException(status_code=404, detail="Guruh topilmadi")
    if actor.role == UserRole.teacher.value:
        if not group_id:
            raise HTTPException(status_code=400, detail="O'qituvchi uchun guruh tanlanishi shart")
        if group_id not in _teacher_group_ids(db, actor):
            raise HTTPException(status_code=403, detail="Bu guruh sizga tegishli emas")
        enrolled = db.query(models.GroupStudent).filter(
            models.GroupStudent.group_id == group_id,
            models.GroupStudent.student_id == student_id,
        ).first()
        if not enrolled:
            raise HTTPException(status_code=404, detail="Talaba bu guruhda emas")


def _check_academic_owner(db: Session, actor: models.User, group_id: Optional[int]) -> None:
    """Teacher faqat o'z guruhidagi yozuvni o'zgartira oladi/o'chira oladi."""
    if actor.role == UserRole.teacher.value:
        if not group_id or group_id not in _teacher_group_ids(db, actor):
            raise HTTPException(status_code=403, detail="Bu yozuv sizga tegishli emas")


@app.get("/academic/options", response_model=List[schemas.AcademicGroupOption])
def academic_options(db: Session = Depends(get_db), actor: models.User = Depends(require_attendance_editor)):
    """Baho/izoh kiritish uchun guruh→talabalar ro'yxati (teacher — faqat o'z guruhlari)."""
    q = db.query(models.Group).filter(models.Group.is_active.is_(True))
    if actor.role == UserRole.teacher.value:
        q = q.filter(models.Group.teacher_id == actor.id)
    out = []
    for g in q.order_by(models.Group.name.asc()).all():
        students = sorted(
            (m.student for m in g.members if m.student and m.student.is_active and not m.student.is_archived),
            key=lambda s: s.full_name,
        )
        out.append(schemas.AcademicGroupOption(
            group_id=g.id, group_name=g.name,
            students=[schemas.StudentBrief(id=s.id, full_name=s.full_name) for s in students],
        ))
    return out


# ── Baholar ──

def _grade_read(g: models.Grade) -> schemas.GradeRead:
    return schemas.GradeRead(
        id=g.id, student_id=g.student_id,
        student_name=g.student.full_name if g.student else None,
        group_id=g.group_id, group_name=g.group.name if g.group else None,
        subject=g.subject, score=g.score, max_score=g.max_score,
        exam_type=g.exam_type, exam_date=g.exam_date, comment=g.comment,
        created_by_name=(g.created_by.full_name or g.created_by.username) if g.created_by else None,
        created_at=g.created_at,
    )


@app.get("/grades", response_model=List[schemas.GradeRead])
def list_grades(
    student_id: Optional[int] = Query(None),
    group_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    actor: models.User = Depends(require_attendance_editor),
):
    q = db.query(models.Grade).options(
        joinedload(models.Grade.student), joinedload(models.Grade.group),
        joinedload(models.Grade.created_by),
    )
    if actor.role == UserRole.teacher.value:
        q = q.filter(models.Grade.group_id.in_(_teacher_group_ids(db, actor) or {0}))
    if student_id:
        q = q.filter(models.Grade.student_id == student_id)
    if group_id:
        q = q.filter(models.Grade.group_id == group_id)
    return [_grade_read(g) for g in q.order_by(models.Grade.exam_date.desc(), models.Grade.id.desc()).limit(300).all()]


@app.post("/grades", response_model=schemas.GradeRead, status_code=201)
def create_grade(payload: schemas.GradeCreate, db: Session = Depends(get_db), actor: models.User = Depends(require_attendance_editor)):
    _check_academic_target(db, actor, payload.student_id, payload.group_id)
    if payload.score > payload.max_score:
        raise HTTPException(status_code=400, detail="Ball maksimal balldan katta bo'lishi mumkin emas")
    g = models.Grade(
        student_id=payload.student_id, group_id=payload.group_id,
        subject=payload.subject, score=payload.score, max_score=payload.max_score,
        exam_type=payload.exam_type, exam_date=payload.exam_date or date.today(),
        comment=payload.comment, created_by_id=actor.id,
    )
    db.add(g)
    db.flush()
    write_audit(db, entity_type="grade", entity_id=g.id, action="create", changed_by_id=actor.id,
                new_value={"student_id": g.student_id, "subject": g.subject,
                           "score": g.score, "max_score": g.max_score, "exam_date": str(g.exam_date)})
    notify_parents_of_student(
        db, student_id=g.student_id, ntype="announcement",
        title=f"Yangi baho: {g.subject}",
        body=f"Natija: {g.score}/{g.max_score}" + (f" — {g.comment}" if g.comment else ""),
    )
    db.commit()
    db.refresh(g)
    return _grade_read(g)


@app.patch("/grades/{grade_id}", response_model=schemas.GradeRead)
def update_grade(grade_id: int, payload: schemas.GradeUpdate, db: Session = Depends(get_db), actor: models.User = Depends(require_attendance_editor)):
    g = db.query(models.Grade).filter(models.Grade.id == grade_id).first()
    if not g:
        raise HTTPException(status_code=404, detail="Baho topilmadi")
    _check_academic_owner(db, actor, g.group_id)
    changes = payload.dict(exclude_unset=True)
    old = {k: str(getattr(g, k)) for k in changes}
    for k, v in changes.items():
        setattr(g, k, v)
    if g.score > g.max_score:
        raise HTTPException(status_code=400, detail="Ball maksimal balldan katta bo'lishi mumkin emas")
    write_audit(db, entity_type="grade", entity_id=g.id, action="update", changed_by_id=actor.id,
                old_value=old, new_value={k: str(v) for k, v in changes.items()})
    db.commit()
    db.refresh(g)
    return _grade_read(g)


@app.delete("/grades/{grade_id}", status_code=204)
def delete_grade(grade_id: int, db: Session = Depends(get_db), actor: models.User = Depends(require_attendance_editor)):
    g = db.query(models.Grade).filter(models.Grade.id == grade_id).first()
    if not g:
        raise HTTPException(status_code=404, detail="Baho topilmadi")
    _check_academic_owner(db, actor, g.group_id)
    write_audit(db, entity_type="grade", entity_id=g.id, action="delete", changed_by_id=actor.id,
                old_value={"student_id": g.student_id, "subject": g.subject, "score": g.score})
    db.delete(g)
    db.commit()


# ── O'qituvchi izohlari ──

def _feedback_read(f: models.TeacherFeedback) -> schemas.TeacherFeedbackRead:
    return schemas.TeacherFeedbackRead(
        id=f.id, student_id=f.student_id,
        student_name=f.student.full_name if f.student else None,
        group_id=f.group_id, group_name=f.group.name if f.group else None,
        teacher_name=(f.teacher.full_name or f.teacher.username) if f.teacher else None,
        comment=f.comment, created_at=f.created_at,
        status=f.status or "new",
        status_updated_by_name=(f.status_updated_by.full_name or f.status_updated_by.username) if f.status_updated_by else None,
        status_updated_at=f.status_updated_at,
    )


@app.get("/teacher-feedbacks", response_model=List[schemas.TeacherFeedbackRead])
def list_teacher_feedbacks(
    student_id: Optional[int] = Query(None),
    status: Optional[str] = Query(None),     # new | resolved | no_answer
    db: Session = Depends(get_db),
    actor: models.User = Depends(require_feedback_viewer),
):
    q = db.query(models.TeacherFeedback).options(
        joinedload(models.TeacherFeedback.student), joinedload(models.TeacherFeedback.group),
        joinedload(models.TeacherFeedback.teacher), joinedload(models.TeacherFeedback.status_updated_by),
    )
    if actor.role == UserRole.teacher.value:
        q = q.filter(models.TeacherFeedback.group_id.in_(_teacher_group_ids(db, actor) or {0}))
    if student_id:
        q = q.filter(models.TeacherFeedback.student_id == student_id)
    if status:
        q = q.filter(models.TeacherFeedback.status == status)
    return [_feedback_read(f) for f in q.order_by(models.TeacherFeedback.created_at.desc()).limit(300).all()]


@app.get("/teacher-feedbacks/new-count")
def teacher_feedbacks_new_count(
    _: models.User = Depends(require_feedback_status),
    db: Session = Depends(get_db),
):
    """Yangi (hal qilinmagan) izohlar soni — sidebar'dagi qizil badge uchun."""
    n = (
        db.query(func.count(models.TeacherFeedback.id))
        .filter(models.TeacherFeedback.status == "new")
        .scalar()
    )
    return {"count": n or 0}


@app.post("/teacher-feedbacks", response_model=schemas.TeacherFeedbackRead, status_code=201)
def create_teacher_feedback(payload: schemas.TeacherFeedbackCreate, db: Session = Depends(get_db), actor: models.User = Depends(require_attendance_editor)):
    _check_academic_target(db, actor, payload.student_id, payload.group_id)
    f = models.TeacherFeedback(
        student_id=payload.student_id, group_id=payload.group_id,
        teacher_id=actor.id, comment=payload.comment,
    )
    db.add(f)
    db.flush()
    write_audit(db, entity_type="teacher_feedback", entity_id=f.id, action="create", changed_by_id=actor.id,
                new_value={"student_id": f.student_id, "comment": f.comment[:200]})
    notify_parents_of_student(
        db, student_id=f.student_id, ntype="announcement",
        title="O'qituvchi izohi", body=f.comment[:300],
    )
    # Hunter va Call center yangi izohni kuzatib borishi kerak — qizil badge + notification
    student_name = db.query(models.Student.full_name).filter(models.Student.id == f.student_id).scalar()
    _notify(
        db, user_ids=[u[0] for u in db.query(models.User.id).filter(
            models.User.is_active == True,  # noqa: E712
            models.User.role.in_([UserRole.hunter.value, UserRole.call_center.value]),
            models.User.id != actor.id,
        ).all()],
        title="Yangi o'quvchi izohi",
        body=f"{student_name or 'Talaba'}: {f.comment[:250]}",
        link="/feedbacks", ntype="feedback",
    )
    db.commit()
    db.refresh(f)
    return _feedback_read(f)


@app.patch("/teacher-feedbacks/{feedback_id}", response_model=schemas.TeacherFeedbackRead)
def update_teacher_feedback(feedback_id: int, payload: schemas.TeacherFeedbackUpdate, db: Session = Depends(get_db), actor: models.User = Depends(require_attendance_editor)):
    f = db.query(models.TeacherFeedback).filter(models.TeacherFeedback.id == feedback_id).first()
    if not f:
        raise HTTPException(status_code=404, detail="Izoh topilmadi")
    if actor.role == UserRole.teacher.value and f.teacher_id != actor.id:
        raise HTTPException(status_code=403, detail="Bu izoh sizga tegishli emas")
    old = f.comment
    f.comment = payload.comment
    write_audit(db, entity_type="teacher_feedback", entity_id=f.id, action="update", changed_by_id=actor.id,
                old_value={"comment": old[:200]}, new_value={"comment": f.comment[:200]})
    db.commit()
    db.refresh(f)
    return _feedback_read(f)


_FEEDBACK_STATUS_LABELS = {
    "new": "Yangi",
    "resolved": "Hal qilindi",
    "no_answer": "Telefonga ota-onasi javob bermadi",
}


@app.patch("/teacher-feedbacks/{feedback_id}/status", response_model=schemas.TeacherFeedbackRead)
def update_teacher_feedback_status(
    feedback_id: int,
    payload: schemas.TeacherFeedbackStatusUpdate,
    db: Session = Depends(get_db),
    actor: models.User = Depends(require_feedback_status),
):
    """Izoh kuzatuv statusi: hal qilindi / telefonga ota-onasi javob bermadi."""
    f = db.query(models.TeacherFeedback).filter(models.TeacherFeedback.id == feedback_id).first()
    if not f:
        raise HTTPException(status_code=404, detail="Izoh topilmadi")
    old = f.status or "new"
    f.status = payload.status
    f.status_updated_by_id = actor.id
    f.status_updated_at = datetime.utcnow()
    write_audit(db, entity_type="teacher_feedback", entity_id=f.id, action="status", changed_by_id=actor.id,
                old_value={"status": old}, new_value={"status": f.status})
    # Izoh egasi (o'qituvchi) natijadan xabardor bo'ladi
    if f.teacher_id and f.teacher_id != actor.id:
        _notify(
            db, user_ids=[f.teacher_id],
            title="Izoh statusi yangilandi",
            body=f"{_FEEDBACK_STATUS_LABELS.get(f.status, f.status)} — {f.comment[:200]}",
            link="/feedbacks", ntype="feedback",
        )
    db.commit()
    db.refresh(f)
    db.refresh(f, attribute_names=["student", "group", "teacher", "status_updated_by"])
    return _feedback_read(f)


@app.delete("/teacher-feedbacks/{feedback_id}", status_code=204)
def delete_teacher_feedback(feedback_id: int, db: Session = Depends(get_db), actor: models.User = Depends(require_attendance_editor)):
    f = db.query(models.TeacherFeedback).filter(models.TeacherFeedback.id == feedback_id).first()
    if not f:
        raise HTTPException(status_code=404, detail="Izoh topilmadi")
    if actor.role == UserRole.teacher.value and f.teacher_id != actor.id:
        raise HTTPException(status_code=403, detail="Bu izoh sizga tegishli emas")
    write_audit(db, entity_type="teacher_feedback", entity_id=f.id, action="delete", changed_by_id=actor.id,
                old_value={"student_id": f.student_id, "comment": f.comment[:200]})
    db.delete(f)
    db.commit()


# ── Sertifikatlar ──

MAX_CERT_PDF_BYTES = 10 * 1024 * 1024  # 10 MB
_CERT_URL_MARKER = "/uploads/certificates/"


def _delete_cert_file_if_local(file_url: str) -> None:
    """Bizning serverga yuklangan PDF bo'lsa — faylni ham o'chiramiz (best-effort)."""
    if _CERT_URL_MARKER not in (file_url or ""):
        return
    filename = file_url.rsplit("/", 1)[-1]
    # Path traversal himoyasi: faqat bizning uuid.pdf formatimiz
    if not filename.endswith(".pdf") or "/" in filename or ".." in filename:
        return
    try:
        os.remove(os.path.join(CERTIFICATES_DIR, filename))
    except OSError:
        pass


@app.post("/certificates/upload")
async def upload_certificate_pdf(
    file: UploadFile = File(...),
    _: models.User = Depends(require_lms_write),
):
    """PDF faylni serverga yuklaydi va public file_url qaytaradi.

    Fayl nomi tasodifiy (uuid) — taxmin qilib bo'lmaydi, shu sababli
    /uploads statikasi authsiz bo'lsa ham sertifikatlar sanab chiqilmaydi.
    """
    contents = await file.read()
    if len(contents) > MAX_CERT_PDF_BYTES:
        raise HTTPException(400, "Fayl hajmi 10 MB dan oshmasin")
    # Content-type'ga ham, fayl nomiga ham ishonmaymiz — magic bytes tekshiramiz
    if not contents.startswith(b"%PDF-"):
        raise HTTPException(400, "Faqat PDF fayl yuklash mumkin")

    filename = f"{uuid4().hex}.pdf"
    with open(os.path.join(CERTIFICATES_DIR, filename), "wb") as f:
        f.write(contents)

    return {"file_url": f"{PUBLIC_BASE_URL}{PUBLIC_API_PREFIX}{_CERT_URL_MARKER}{filename}"}


def _certificate_read(c: models.Certificate) -> schemas.CertificateRead:
    return schemas.CertificateRead(
        id=c.id, student_id=c.student_id,
        student_name=c.student.full_name if c.student else None,
        title=c.title, file_url=c.file_url, issued_at=c.issued_at,
        created_by_name=(c.created_by.full_name or c.created_by.username) if c.created_by else None,
        created_at=c.created_at,
    )


@app.get("/certificates", response_model=List[schemas.CertificateRead])
def list_certificates(
    student_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    _: models.User = Depends(require_lms_write),
):
    q = db.query(models.Certificate).options(
        joinedload(models.Certificate.student), joinedload(models.Certificate.created_by),
    )
    if student_id:
        q = q.filter(models.Certificate.student_id == student_id)
    return [_certificate_read(c) for c in q.order_by(models.Certificate.created_at.desc()).limit(300).all()]


@app.post("/certificates", response_model=schemas.CertificateRead, status_code=201)
def create_certificate(payload: schemas.CertificateCreate, db: Session = Depends(get_db), actor: models.User = Depends(require_lms_write)):
    if not db.query(models.Student).filter(models.Student.id == payload.student_id).first():
        raise HTTPException(status_code=404, detail="Talaba topilmadi")
    c = models.Certificate(
        student_id=payload.student_id, title=payload.title,
        file_url=payload.file_url, issued_at=payload.issued_at,
        created_by_id=actor.id,
    )
    db.add(c)
    db.flush()
    write_audit(db, entity_type="certificate", entity_id=c.id, action="create", changed_by_id=actor.id,
                new_value={"student_id": c.student_id, "title": c.title, "file_url": c.file_url})
    notify_parents_of_student(
        db, student_id=c.student_id, ntype="announcement",
        title="Yangi sertifikat", body=c.title,
    )
    db.commit()
    db.refresh(c)
    return _certificate_read(c)


@app.patch("/certificates/{certificate_id}", response_model=schemas.CertificateRead)
def update_certificate(certificate_id: int, payload: schemas.CertificateUpdate, db: Session = Depends(get_db), actor: models.User = Depends(require_lms_write)):
    c = db.query(models.Certificate).filter(models.Certificate.id == certificate_id).first()
    if not c:
        raise HTTPException(status_code=404, detail="Sertifikat topilmadi")
    changes = payload.dict(exclude_unset=True)
    old = {k: str(getattr(c, k)) for k in changes}
    old_file_url = c.file_url
    for k, v in changes.items():
        setattr(c, k, v)
    write_audit(db, entity_type="certificate", entity_id=c.id, action="update", changed_by_id=actor.id,
                old_value=old, new_value={k: str(v) for k, v in changes.items()})
    db.commit()
    db.refresh(c)
    if "file_url" in changes and changes["file_url"] != old_file_url:
        _delete_cert_file_if_local(old_file_url)
    return _certificate_read(c)


@app.delete("/certificates/{certificate_id}", status_code=204)
def delete_certificate(certificate_id: int, db: Session = Depends(get_db), actor: models.User = Depends(require_lms_write)):
    c = db.query(models.Certificate).filter(models.Certificate.id == certificate_id).first()
    if not c:
        raise HTTPException(status_code=404, detail="Sertifikat topilmadi")
    write_audit(db, entity_type="certificate", entity_id=c.id, action="delete", changed_by_id=actor.id,
                old_value={"student_id": c.student_id, "title": c.title})
    file_url = c.file_url
    db.delete(c)
    db.commit()
    _delete_cert_file_if_local(file_url)


# ── Tadbirlar ──

def _event_read(e: models.Event) -> schemas.EventRead:
    return schemas.EventRead(
        id=e.id, title=e.title, description=e.description,
        event_date=e.event_date, location=e.location, is_active=e.is_active,
        created_by_name=(e.created_by.full_name or e.created_by.username) if e.created_by else None,
        created_at=e.created_at,
    )


@app.get("/events", response_model=List[schemas.EventRead])
def list_events(db: Session = Depends(get_db), _: models.User = Depends(require_auth)):
    rows = (
        db.query(models.Event).options(joinedload(models.Event.created_by))
        .order_by(models.Event.event_date.desc()).limit(200).all()
    )
    return [_event_read(e) for e in rows]


@app.post("/events", response_model=schemas.EventRead, status_code=201)
def create_event(payload: schemas.EventCreate, db: Session = Depends(get_db), actor: models.User = Depends(require_admin)):
    e = models.Event(
        title=payload.title, description=payload.description,
        event_date=payload.event_date, location=payload.location,
        created_by_id=actor.id,
    )
    db.add(e)
    db.flush()
    write_audit(db, entity_type="event", entity_id=e.id, action="create", changed_by_id=actor.id,
                new_value={"title": e.title, "event_date": str(e.event_date), "location": e.location})
    db.commit()
    db.refresh(e)
    return _event_read(e)


@app.patch("/events/{event_id}", response_model=schemas.EventRead)
def update_event(event_id: int, payload: schemas.EventUpdate, db: Session = Depends(get_db), actor: models.User = Depends(require_admin)):
    e = db.query(models.Event).filter(models.Event.id == event_id).first()
    if not e:
        raise HTTPException(status_code=404, detail="Tadbir topilmadi")
    changes = payload.dict(exclude_unset=True)
    old = {k: str(getattr(e, k)) for k in changes}
    for k, v in changes.items():
        setattr(e, k, v)
    write_audit(db, entity_type="event", entity_id=e.id, action="update", changed_by_id=actor.id,
                old_value=old, new_value={k: str(v) for k, v in changes.items()})
    db.commit()
    db.refresh(e)
    return _event_read(e)


@app.delete("/events/{event_id}", status_code=204)
def delete_event(event_id: int, db: Session = Depends(get_db), actor: models.User = Depends(require_admin)):
    e = db.query(models.Event).filter(models.Event.id == event_id).first()
    if not e:
        raise HTTPException(status_code=404, detail="Tadbir topilmadi")
    write_audit(db, entity_type="event", entity_id=e.id, action="delete", changed_by_id=actor.id,
                old_value={"title": e.title, "event_date": str(e.event_date)})
    db.delete(e)
    db.commit()


# ── Coinlar (o'qituvchi rag'bati) ────────────────────────────────────────────
# Budjet alohida saqlanmaydi: oylik budjet = COINS_PER_STUDENT × o'qituvchining
# faol guruhlaridagi faol talabalar soni; sarf esa joriy kalendar oyda
# berilganlar yig'indisi. Shu sabab har oyning 1-kunida qoldiq o'z-o'zidan
# to'liq budjetga qaytadi — cron kerak emas.

COINS_PER_STUDENT = int(os.getenv("COINS_PER_STUDENT", "50"))


def _teacher_coin_budget(db: Session, teacher_id: int) -> int:
    """Oylik budjet: 50 coin × o'qituvchi guruhlaridagi har bir faol talaba."""
    student_count = (
        db.query(func.count(func.distinct(models.GroupStudent.student_id)))
        .join(models.Group, models.Group.id == models.GroupStudent.group_id)
        .join(models.Student, models.Student.id == models.GroupStudent.student_id)
        .filter(
            models.Group.teacher_id == teacher_id,
            models.Group.is_active.is_(True),
            models.Student.is_active.is_(True),
            models.Student.is_archived.is_(False),
        )
        .scalar()
    ) or 0
    return COINS_PER_STUDENT * student_count


def _coins_spent_this_month(db: Session, teacher_id: int, now: Optional[datetime] = None) -> int:
    now = now or datetime.utcnow()
    return int(
        db.query(func.coalesce(func.sum(models.CoinTransaction.amount), 0))
        .filter(
            models.CoinTransaction.teacher_id == teacher_id,
            extract("month", models.CoinTransaction.created_at) == now.month,
            extract("year", models.CoinTransaction.created_at) == now.year,
        )
        .scalar()
    )


def _coin_tx_read(t: models.CoinTransaction) -> schemas.CoinTransactionRead:
    return schemas.CoinTransactionRead(
        id=t.id, student_id=t.student_id,
        student_name=t.student.full_name if t.student else None,
        group_name=t.group.name if t.group else None,
        teacher_name=(t.teacher.full_name or t.teacher.username) if t.teacher else None,
        amount=t.amount, reason=t.reason, created_at=t.created_at,
    )


@app.get("/coins/summary", response_model=schemas.CoinSummary)
def coin_summary(db: Session = Depends(get_db), actor: models.User = Depends(require_attendance_editor)):
    now = datetime.utcnow()
    spent = _coins_spent_this_month(db, actor.id, now)
    if actor.role == UserRole.teacher.value:
        budget = _teacher_coin_budget(db, actor.id)
        return schemas.CoinSummary(month=now.month, year=now.year, budget=budget,
                                   spent=spent, remaining=max(0, budget - spent))
    return schemas.CoinSummary(month=now.month, year=now.year, budget=None, spent=spent, remaining=None)


@app.post("/coins/give", response_model=schemas.CoinTransactionRead, status_code=201)
def give_coins(payload: schemas.CoinGive, db: Session = Depends(get_db), actor: models.User = Depends(require_attendance_editor)):
    _check_academic_target(db, actor, payload.student_id, payload.group_id)

    # Teacher uchun oylik budjet nazorati (boshqa rollar cheklanmagan)
    if actor.role == UserRole.teacher.value:
        budget = _teacher_coin_budget(db, actor.id)
        remaining = budget - _coins_spent_this_month(db, actor.id)
        if payload.amount > remaining:
            raise HTTPException(
                status_code=400,
                detail=f"Coin yetarli emas: qoldiq {max(0, remaining)} (har oy 1-sanada {budget} ga to'ladi)",
            )

    t = models.CoinTransaction(
        student_id=payload.student_id, teacher_id=actor.id, group_id=payload.group_id,
        amount=payload.amount, reason=payload.reason, created_at=datetime.utcnow(),
    )
    db.add(t)
    db.flush()
    write_audit(db, entity_type="coin", entity_id=t.id, action="create", changed_by_id=actor.id,
                new_value={"student_id": t.student_id, "amount": t.amount, "reason": t.reason})
    notify_parents_of_student(
        db, student_id=t.student_id, ntype="announcement",
        title=f"Farzandingiz {t.amount} coin oldi 🎉",
        body=t.reason or "O'qituvchi rag'bati",
    )
    db.commit()
    db.refresh(t)
    return _coin_tx_read(t)


def _student_coin_total(db: Session, student_id: int) -> int:
    return int(
        db.query(func.coalesce(func.sum(models.CoinTransaction.amount), 0))
        .filter(models.CoinTransaction.student_id == student_id)
        .scalar()
    )


@app.post("/coins/deduct", response_model=schemas.CoinTransactionRead, status_code=201)
def deduct_coins(payload: schemas.CoinDeduct, db: Session = Depends(get_db), actor: models.User = Depends(require_admin)):
    """Talabadan coin yechish (jarima/tuzatish) — faqat admin. Manfiy yozuv sifatida saqlanadi."""
    student = db.query(models.Student).filter(models.Student.id == payload.student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="Talaba topilmadi")

    total = _student_coin_total(db, payload.student_id)
    if payload.amount > total:
        raise HTTPException(status_code=400, detail=f"Talabada faqat {max(0, total)} coin bor — balans manfiy bo'lolmaydi")

    t = models.CoinTransaction(
        student_id=payload.student_id, teacher_id=actor.id, group_id=None,
        amount=-payload.amount, reason=payload.reason, created_at=datetime.utcnow(),
    )
    db.add(t)
    db.flush()
    write_audit(db, entity_type="coin", entity_id=t.id, action="deduct", changed_by_id=actor.id,
                new_value={"student_id": t.student_id, "amount": t.amount, "reason": t.reason})
    notify_parents_of_student(
        db, student_id=t.student_id, ntype="announcement",
        title=f"Farzandingizdan {payload.amount} coin yechildi",
        body=t.reason or "Administratsiya qarori",
    )
    db.commit()
    db.refresh(t)
    return _coin_tx_read(t)


@app.get("/coins/transactions", response_model=List[schemas.CoinTransactionRead])
def list_coin_transactions(
    student_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    actor: models.User = Depends(require_attendance_editor),
):
    q = db.query(models.CoinTransaction).options(
        joinedload(models.CoinTransaction.student),
        joinedload(models.CoinTransaction.group),
        joinedload(models.CoinTransaction.teacher),
    )
    if actor.role == UserRole.teacher.value:
        q = q.filter(models.CoinTransaction.teacher_id == actor.id)
    if student_id:
        q = q.filter(models.CoinTransaction.student_id == student_id)
    return [_coin_tx_read(t) for t in q.order_by(models.CoinTransaction.created_at.desc()).limit(300).all()]


@app.get("/coins/totals", response_model=List[schemas.StudentCoinTotal])
def coin_totals(db: Session = Depends(get_db), actor: models.User = Depends(require_attendance_editor)):
    """Talabalar bo'yicha umumiy coin reytingi (teacher — faqat o'z guruhlari talabalari)."""
    q = (
        db.query(models.Student.id, models.Student.full_name,
                 func.coalesce(func.sum(models.CoinTransaction.amount), 0).label("total"))
        .join(models.CoinTransaction, models.CoinTransaction.student_id == models.Student.id)
        .group_by(models.Student.id)
    )
    if actor.role == UserRole.teacher.value:
        gids = _teacher_group_ids(db, actor) or {0}
        q = q.join(models.GroupStudent, models.GroupStudent.student_id == models.Student.id) \
             .filter(models.GroupStudent.group_id.in_(gids)).group_by(models.Student.id)
    rows = q.order_by(func.sum(models.CoinTransaction.amount).desc()).limit(100).all()
    return [schemas.StudentCoinTotal(student_id=r[0], student_name=r[1], total=int(r[2])) for r in rows]


# ═══════════════════════════════════════════════════════
#  LEADS  (Hunter & Call Center CRM)
# ═══════════════════════════════════════════════════════

def _lead_read(lead: models.Lead) -> schemas.LeadRead:
    st = lead.stage
    src = lead.source
    return schemas.LeadRead(
        id=lead.id,
        full_name=lead.full_name,
        phone=lead.phone,
        course_interest=lead.course_interest,
        status=lead.status,
        stage_id=lead.stage_id,
        stage_name=st.name if st else None,
        stage_color=st.color if st else None,
        stage_icon=st.icon if st else None,
        stage_kind=st.kind if st else None,
        source_id=lead.source_id,
        source_name=src.name if src else None,
        callback_at=lead.callback_at,
        notes=lead.notes,
        date_of_birth=lead.date_of_birth,
        parent_phone=lead.parent_phone,
        parent2_phone=lead.parent2_phone,
        interested_group_id=lead.interested_group_id,
        interested_group_name=lead.interested_group.name if lead.interested_group else None,
        is_shared=lead.is_shared,
        claimed_by_id=lead.claimed_by_id,
        claimed_by_name=(lead.claimed_by.full_name or lead.claimed_by.username) if lead.claimed_by else None,
        created_by_id=lead.created_by_id,
        created_by_name=lead.created_by.full_name or lead.created_by.username if lead.created_by else None,
        updated_by_name=lead.updated_by.full_name or lead.updated_by.username if lead.updated_by else None,
        created_at=lead.created_at,
        updated_at=lead.updated_at,
    )


def _log_lead_activity(db, *, lead_id, action, description, author_id=None, meta=None):
    db.add(models.LeadActivity(
        lead_id=lead_id,
        action=action,
        description=description,
        author_id=author_id,
        meta_json=json.dumps(meta, ensure_ascii=False, default=str) if meta else None,
        created_at=datetime.utcnow(),
    ))


def _notify(db, *, user_ids, title, body=None, link=None, ntype="new_lead"):
    now = datetime.utcnow()
    for uid in set(user_ids):
        db.add(models.Notification(
            user_id=uid, notification_type=ntype, title=title,
            body=body, link=link, is_read=False, created_at=now,
        ))


def _crm_recipient_ids(db, exclude_id=None):
    """Yangi lid haqida xabardor qilinadigan foydalanuvchilar (admin + call_center)."""
    rows = (
        db.query(models.User.id)
        .filter(
            models.User.is_active == True,  # noqa: E712
            models.User.role.in_([UserRole.admin.value, UserRole.call_center.value]),
        ).all()
    )
    return [r[0] for r in rows if r[0] != exclude_id]


def _slugify(name: str) -> str:
    base = "".join(c if c.isalnum() else "-" for c in name.lower()).strip("-")
    base = "-".join(filter(None, base.split("-"))) or "stage"
    return base[:70]


def _default_stage(db) -> Optional[models.LeadStage]:
    return (
        db.query(models.LeadStage)
        .filter(models.LeadStage.is_archived == False, models.LeadStage.kind == "lead")  # noqa: E712
        .order_by(models.LeadStage.order.asc())
        .first()
    )


_LEAD_LOAD = (
    joinedload(models.Lead.created_by),
    joinedload(models.Lead.updated_by),
    joinedload(models.Lead.claimed_by),
    joinedload(models.Lead.stage),
    joinedload(models.Lead.source),
    joinedload(models.Lead.interested_group),
)


@app.get("/leads", response_model=List[schemas.LeadRead])
def list_leads(
    status: Optional[str] = Query(None),
    stage: Optional[str] = Query(None),      # stage.slug bo'yicha filtr
    source_id: Optional[int] = Query(None),
    pool: bool = Query(False),               # faqat umumiy havza (band qilinmagan)
    search: Optional[str] = Query(None),
    actor: models.User = Depends(require_crm_access),
    db: Session = Depends(get_db),
):
    q = db.query(models.Lead).options(*_LEAD_LOAD)
    if pool:
        # Umumiy havza: ulashilgan va hali band qilinmagan lidlar
        q = q.filter(models.Lead.is_shared == True, models.Lead.claimed_by_id.is_(None))  # noqa: E712
    elif actor.role in (UserRole.hunter.value, UserRole.sales.value):
        # Hunter/Sales: o'zi yaratgan yoki band qilgan + havzadagi bo'sh lidlar
        q = q.filter(
            (models.Lead.created_by_id == actor.id) |
            (models.Lead.claimed_by_id == actor.id) |
            ((models.Lead.is_shared == True) & (models.Lead.claimed_by_id.is_(None)))  # noqa: E712
        )
    if stage:
        q = q.join(models.LeadStage, models.Lead.stage_id == models.LeadStage.id).filter(models.LeadStage.slug == stage)
    elif status:
        q = q.filter(models.Lead.status == status)
    if source_id:
        q = q.filter(models.Lead.source_id == source_id)
    if search:
        q = q.filter(
            models.Lead.full_name.ilike(f"%{search}%") |
            models.Lead.phone.ilike(f"%{search}%")
        )
    leads = q.order_by(models.Lead.created_at.desc()).all()
    return [_lead_read(l) for l in leads]


@app.get("/leads/stats", response_model=schemas.LeadStatsRead)
def lead_stats(
    actor: models.User = Depends(require_crm_access),
    db: Session = Depends(get_db),
):
    lead_q = db.query(models.Lead)
    if actor.role in (UserRole.hunter.value, UserRole.sales.value):
        lead_q = lead_q.filter(models.Lead.created_by_id == actor.id)
    total = lead_q.count()

    counts = dict(
        lead_q.with_entities(models.Lead.stage_id, func.count(models.Lead.id))
        .group_by(models.Lead.stage_id).all()
    )
    stages = (
        db.query(models.LeadStage)
        .filter(models.LeadStage.is_archived == False)  # noqa: E712
        .order_by(models.LeadStage.order.asc()).all()
    )
    return schemas.LeadStatsRead(
        total=total,
        stages=[
            schemas.LeadStageStat(
                slug=s.slug, name=s.name, color=s.color, icon=s.icon,
                kind=s.kind, order=s.order, count=counts.get(s.id, 0),
            ) for s in stages
        ],
    )


@app.post("/leads", response_model=schemas.LeadRead, status_code=201)
def create_lead(
    payload: schemas.LeadCreate,
    actor: models.User = Depends(require_crm_access),
    db: Session = Depends(get_db),
):
    stage = _default_stage(db)
    lead = models.Lead(
        **payload.dict(),
        status=stage.slug if stage else models.LeadStatus.new.value,
        stage_id=stage.id if stage else None,
        created_by_id=actor.id,
        created_at=datetime.utcnow(),
    )
    db.add(lead)
    db.flush()
    _log_lead_activity(
        db, lead_id=lead.id, action="created",
        description=f"Lid qo'shildi: {lead.full_name}", author_id=actor.id,
    )
    _notify(
        db, user_ids=_crm_recipient_ids(db, exclude_id=actor.id),
        title="Yangi lid qo'shildi",
        body=f"{lead.full_name} · {lead.source.name if lead.source else 'manba yo‘q'}",
        link=f"/leads?lead={lead.id}", ntype="new_lead",
    )
    db.commit()
    db.refresh(lead)
    db.refresh(lead, attribute_names=["created_by", "updated_by", "stage", "source"])
    return _lead_read(lead)


@app.patch("/leads/{lead_id}/status", response_model=schemas.LeadRead)
def update_lead_status(
    lead_id: int,
    payload: schemas.LeadStatusUpdate,
    actor: models.User = Depends(require_call_center),
    db: Session = Depends(get_db),
):
    lead = db.query(models.Lead).options(*_LEAD_LOAD).filter(models.Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(404, "Lid topilmadi")
    old_name = lead.stage.name if lead.stage else lead.status
    new_slug = payload.status.value
    lead.status = new_slug
    # status ustunini stage bilan sinxronlaymiz
    stage = db.query(models.LeadStage).filter(models.LeadStage.slug == new_slug).first()
    if stage:
        lead.stage_id = stage.id
    lead.callback_at = payload.callback_at
    if payload.notes is not None:
        lead.notes = payload.notes
    lead.updated_by_id = actor.id
    lead.updated_at = datetime.utcnow()
    new_name = stage.name if stage else new_slug
    if new_name != old_name:
        _log_lead_activity(
            db, lead_id=lead.id, action="stage_changed",
            description=f"Holat: {old_name} → {new_name}", author_id=actor.id,
            meta={"old": old_name, "new": new_name},
        )
    db.commit()
    db.refresh(lead)
    db.refresh(lead, attribute_names=["created_by", "updated_by", "stage", "source"])
    return _lead_read(lead)


@app.patch("/leads/{lead_id}/stage", response_model=schemas.LeadRead)
def move_lead_stage(
    lead_id: int,
    payload: schemas.LeadStageMove,
    actor: models.User = Depends(require_crm_access),
    db: Session = Depends(get_db),
):
    lead = db.query(models.Lead).options(*_LEAD_LOAD).filter(models.Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(404, "Lid topilmadi")
    if actor.role in (UserRole.hunter.value, UserRole.sales.value) and lead.created_by_id != actor.id:
        raise HTTPException(403, "Faqat o'z lidingizni o'zgartira olasiz")
    stage = db.query(models.LeadStage).filter(models.LeadStage.id == payload.stage_id).first()
    if not stage:
        raise HTTPException(404, "Bosqich topilmadi")
    old_name = lead.stage.name if lead.stage else lead.status
    lead.stage_id = stage.id
    lead.status = stage.slug
    if payload.callback_at is not None:
        lead.callback_at = payload.callback_at
    if payload.notes is not None:
        lead.notes = payload.notes
    lead.updated_by_id = actor.id
    lead.updated_at = datetime.utcnow()
    if stage.name != old_name:
        _log_lead_activity(
            db, lead_id=lead.id, action="stage_changed",
            description=f"Holat: {old_name} → {stage.name}", author_id=actor.id,
            meta={"old": old_name, "new": stage.name},
        )
    db.commit()
    db.refresh(lead)
    db.refresh(lead, attribute_names=["created_by", "updated_by", "stage", "source"])
    return _lead_read(lead)


@app.get("/leads/{lead_id}/activities", response_model=List[schemas.LeadActivityRead])
def lead_activities(
    lead_id: int,
    actor: models.User = Depends(require_crm_access),
    db: Session = Depends(get_db),
):
    lead = db.query(models.Lead).filter(models.Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(404, "Lid topilmadi")
    if actor.role in (UserRole.hunter.value, UserRole.sales.value) and lead.created_by_id != actor.id:
        raise HTTPException(403, "Ruxsat yo'q")
    acts = (
        db.query(models.LeadActivity)
        .options(joinedload(models.LeadActivity.author))
        .filter(models.LeadActivity.lead_id == lead_id)
        .order_by(models.LeadActivity.created_at.desc()).all()
    )
    return [
        schemas.LeadActivityRead(
            id=a.id, action=a.action, description=a.description,
            author_name=(a.author.full_name or a.author.username) if a.author else None,
            created_at=a.created_at,
        ) for a in acts
    ]


@app.post("/leads/{lead_id}/claim", response_model=schemas.LeadRead)
def claim_lead(
    lead_id: int,
    actor: models.User = Depends(require_crm_access),
    db: Session = Depends(get_db),
):
    """Umumiy havzadagi lidni o'ziga biriktirish."""
    lead = db.query(models.Lead).options(*_LEAD_LOAD).filter(models.Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(404, "Lid topilmadi")
    if lead.claimed_by_id and lead.claimed_by_id != actor.id:
        raise HTTPException(409, "Bu lid allaqachon band qilingan")
    lead.claimed_by_id = actor.id
    lead.claimed_at = datetime.utcnow()
    lead.updated_by_id = actor.id
    lead.updated_at = datetime.utcnow()
    actor_name = actor.full_name or actor.username
    _log_lead_activity(
        db, lead_id=lead.id, action="claimed",
        description=f"Lid band qilindi: {actor_name}", author_id=actor.id,
    )
    db.commit()
    db.refresh(lead)
    db.refresh(lead, attribute_names=["created_by", "updated_by", "claimed_by", "stage", "source", "interested_group"])
    return _lead_read(lead)


@app.post("/leads/{lead_id}/release", response_model=schemas.LeadRead)
def release_lead(
    lead_id: int,
    actor: models.User = Depends(require_crm_access),
    db: Session = Depends(get_db),
):
    """Lidni umumiy havzaga qaytarish (band qilishni bekor qilish)."""
    lead = db.query(models.Lead).options(*_LEAD_LOAD).filter(models.Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(404, "Lid topilmadi")
    is_owner = lead.claimed_by_id == actor.id
    if not is_owner and actor.role != UserRole.admin.value:
        raise HTTPException(403, "Faqat egasi yoki admin havzaga qaytara oladi")
    lead.claimed_by_id = None
    lead.claimed_at = None
    lead.is_shared = True
    lead.updated_by_id = actor.id
    lead.updated_at = datetime.utcnow()
    _log_lead_activity(
        db, lead_id=lead.id, action="released",
        description="Lid umumiy havzaga qaytarildi", author_id=actor.id,
    )
    db.commit()
    db.refresh(lead)
    db.refresh(lead, attribute_names=["created_by", "updated_by", "claimed_by", "stage", "source", "interested_group"])
    return _lead_read(lead)


@app.post("/leads/{lead_id}/share", response_model=schemas.LeadRead)
def share_lead(
    lead_id: int,
    actor: models.User = Depends(require_call_center),
    db: Session = Depends(get_db),
):
    """Lidni umumiy havzaga chiqarish (admin/call_center)."""
    lead = db.query(models.Lead).options(*_LEAD_LOAD).filter(models.Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(404, "Lid topilmadi")
    lead.is_shared = True
    lead.claimed_by_id = None
    lead.claimed_at = None
    lead.updated_by_id = actor.id
    lead.updated_at = datetime.utcnow()
    _log_lead_activity(
        db, lead_id=lead.id, action="shared",
        description="Lid umumiy havzaga chiqarildi", author_id=actor.id,
    )
    _notify(
        db, user_ids=[u[0] for u in db.query(models.User.id).filter(
            models.User.is_active == True,  # noqa: E712
            models.User.role.in_([UserRole.hunter.value, UserRole.sales.value]),
        ).all()],
        title="Umumiy havzada yangi lid",
        body=f"{lead.full_name} — band qilish uchun ochiq",
        link=f"/leads?lead={lead.id}", ntype="shared_lead",
    )
    db.commit()
    db.refresh(lead)
    db.refresh(lead, attribute_names=["created_by", "updated_by", "claimed_by", "stage", "source", "interested_group"])
    return _lead_read(lead)


@app.delete("/leads/{lead_id}", status_code=204)
def delete_lead(
    lead_id: int,
    actor: models.User = Depends(require_crm_access),
    db: Session = Depends(get_db),
):
    lead = db.query(models.Lead).filter(models.Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(404, "Lid topilmadi")
    if actor.role in (UserRole.hunter.value, UserRole.sales.value) and lead.created_by_id != actor.id:
        raise HTTPException(403, "Faqat o'z lidingizni o'chira olasiz")
    db.delete(lead)
    db.commit()


# ── Lead stages (pipeline) ────────────────────────────────────────────────────

@app.get("/lead-stages", response_model=List[schemas.LeadStageRead])
def list_lead_stages(
    include_archived: bool = Query(False),
    actor: models.User = Depends(require_crm_access),
    db: Session = Depends(get_db),
):
    q = db.query(models.LeadStage)
    if not include_archived:
        q = q.filter(models.LeadStage.is_archived == False)  # noqa: E712
    stages = q.order_by(models.LeadStage.order.asc()).all()
    counts = dict(
        db.query(models.Lead.stage_id, func.count(models.Lead.id))
        .group_by(models.Lead.stage_id).all()
    )
    out = []
    for s in stages:
        r = schemas.LeadStageRead.from_orm(s)
        r.lead_count = counts.get(s.id, 0)
        out.append(r)
    return out


@app.post("/lead-stages", response_model=schemas.LeadStageRead, status_code=201)
def create_lead_stage(
    payload: schemas.LeadStageCreate,
    actor: models.User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    slug = _slugify(payload.name)
    if db.query(models.LeadStage).filter(models.LeadStage.slug == slug).first():
        slug = f"{slug}-{int(datetime.utcnow().timestamp())}"
    max_order = db.query(func.max(models.LeadStage.order)).scalar() or 0
    stage = models.LeadStage(
        name=payload.name, slug=slug, order=max_order + 10,
        color=payload.color, icon=payload.icon, kind=payload.kind,
        created_at=datetime.utcnow(),
    )
    db.add(stage)
    db.commit()
    db.refresh(stage)
    r = schemas.LeadStageRead.from_orm(stage)
    r.lead_count = 0
    return r


@app.put("/lead-stages/reorder", response_model=List[schemas.LeadStageRead])
def reorder_lead_stages(
    payload: schemas.StageReorder,
    actor: models.User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    for idx, sid in enumerate(payload.ordered_ids):
        db.query(models.LeadStage).filter(models.LeadStage.id == sid).update({"order": (idx + 1) * 10})
    db.commit()
    return list_lead_stages(include_archived=False, actor=actor, db=db)


@app.put("/lead-stages/{stage_id}", response_model=schemas.LeadStageRead)
def update_lead_stage(
    stage_id: int,
    payload: schemas.LeadStageUpdate,
    actor: models.User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    stage = db.query(models.LeadStage).filter(models.LeadStage.id == stage_id).first()
    if not stage:
        raise HTTPException(404, "Bosqich topilmadi")
    for field, val in payload.dict(exclude_unset=True).items():
        setattr(stage, field, val)
    db.commit()
    db.refresh(stage)
    count = db.query(func.count(models.Lead.id)).filter(models.Lead.stage_id == stage.id).scalar()
    r = schemas.LeadStageRead.from_orm(stage)
    r.lead_count = count
    return r


@app.delete("/lead-stages/{stage_id}", status_code=204)
def delete_lead_stage(
    stage_id: int,
    actor: models.User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    stage = db.query(models.LeadStage).filter(models.LeadStage.id == stage_id).first()
    if not stage:
        raise HTTPException(404, "Bosqich topilmadi")
    in_use = db.query(func.count(models.Lead.id)).filter(models.Lead.stage_id == stage_id).scalar()
    if in_use:
        # Ichida lid bo'lsa o'chirmaymiz — arxivlaymiz
        stage.is_archived = True
        db.commit()
        return
    db.delete(stage)
    db.commit()


# ── Lead sources ──────────────────────────────────────────────────────────────

@app.get("/lead-sources", response_model=List[schemas.LeadSourceRead])
def list_lead_sources(
    actor: models.User = Depends(require_crm_access),
    db: Session = Depends(get_db),
):
    return (
        db.query(models.LeadSource)
        .filter(models.LeadSource.is_active == True)  # noqa: E712
        .order_by(models.LeadSource.name.asc()).all()
    )


@app.post("/lead-sources", response_model=schemas.LeadSourceRead, status_code=201)
def create_lead_source(
    payload: schemas.LeadSourceCreate,
    actor: models.User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    src = models.LeadSource(
        name=payload.name, is_campaign=payload.is_campaign,
        is_active=True, created_at=datetime.utcnow(),
    )
    db.add(src)
    db.commit()
    db.refresh(src)
    return src


@app.put("/lead-sources/{source_id}", response_model=schemas.LeadSourceRead)
def update_lead_source(
    source_id: int,
    payload: schemas.LeadSourceUpdate,
    actor: models.User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    src = db.query(models.LeadSource).filter(models.LeadSource.id == source_id).first()
    if not src:
        raise HTTPException(404, "Manba topilmadi")
    for field, val in payload.dict(exclude_unset=True).items():
        setattr(src, field, val)
    db.commit()
    db.refresh(src)
    return src


@app.delete("/lead-sources/{source_id}", status_code=204)
def delete_lead_source(
    source_id: int,
    actor: models.User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    src = db.query(models.LeadSource).filter(models.LeadSource.id == source_id).first()
    if not src:
        raise HTTPException(404, "Manba topilmadi")
    src.is_active = False
    db.commit()


# ── Reminders / tasks ─────────────────────────────────────────────────────────

def _reminder_read(r: models.Reminder) -> schemas.ReminderRead:
    now = datetime.utcnow()
    effective_due = r.snoozed_until or r.due_at
    return schemas.ReminderRead(
        id=r.id,
        lead_id=r.lead_id,
        lead_name=r.lead.full_name if r.lead else None,
        lead_phone=r.lead.phone if r.lead else None,
        assigned_to_id=r.assigned_to_id,
        assigned_to_name=(r.assigned_to.full_name or r.assigned_to.username) if r.assigned_to else None,
        due_at=r.due_at,
        body=r.body,
        kind=r.kind,
        status=r.status,
        snoozed_until=r.snoozed_until,
        is_overdue=r.status == "pending" and effective_due < now,
        created_at=r.created_at,
    )


@app.get("/reminders", response_model=List[schemas.ReminderRead])
def list_reminders(
    status: Optional[str] = Query(None),          # pending | done | dismissed
    lead_id: Optional[int] = Query(None),
    mine: bool = Query(True),
    actor: models.User = Depends(require_crm_access),
    db: Session = Depends(get_db),
):
    q = db.query(models.Reminder).options(
        joinedload(models.Reminder.lead),
        joinedload(models.Reminder.assigned_to),
    )
    if lead_id:
        q = q.filter(models.Reminder.lead_id == lead_id)
    if status:
        q = q.filter(models.Reminder.status == status)
    # Har kim o'ziga tegishli / o'zi yaratganini ko'radi; admin — hammasini
    if mine and actor.role != UserRole.admin.value:
        q = q.filter(
            (models.Reminder.assigned_to_id == actor.id) |
            (models.Reminder.created_by_id == actor.id)
        )
    reminders = q.order_by(models.Reminder.due_at.asc()).all()
    return [_reminder_read(r) for r in reminders]


@app.post("/reminders", response_model=schemas.ReminderRead, status_code=201)
def create_reminder(
    payload: schemas.ReminderCreate,
    actor: models.User = Depends(require_crm_access),
    db: Session = Depends(get_db),
):
    lead = db.query(models.Lead).filter(models.Lead.id == payload.lead_id).first()
    if not lead:
        raise HTTPException(404, "Lid topilmadi")
    r = models.Reminder(
        lead_id=payload.lead_id,
        assigned_to_id=payload.assigned_to_id or actor.id,
        created_by_id=actor.id,
        due_at=payload.due_at,
        body=payload.body,
        kind=payload.kind,
        status="pending",
        created_at=datetime.utcnow(),
    )
    db.add(r)
    _log_lead_activity(
        db, lead_id=lead.id, action="reminder",
        description=f"Eslatma: {payload.body or payload.kind} ({payload.due_at:%d.%m %H:%M})",
        author_id=actor.id,
    )
    db.commit()
    db.refresh(r)
    db.refresh(r, attribute_names=["lead", "assigned_to"])
    return _reminder_read(r)


@app.patch("/reminders/{reminder_id}", response_model=schemas.ReminderRead)
def update_reminder(
    reminder_id: int,
    payload: schemas.ReminderUpdate,
    actor: models.User = Depends(require_crm_access),
    db: Session = Depends(get_db),
):
    r = db.query(models.Reminder).options(
        joinedload(models.Reminder.lead),
        joinedload(models.Reminder.assigned_to),
    ).filter(models.Reminder.id == reminder_id).first()
    if not r:
        raise HTTPException(404, "Eslatma topilmadi")
    data = payload.dict(exclude_unset=True)
    for field, val in data.items():
        setattr(r, field, val)
    if data.get("status") == "done" and not r.done_at:
        r.done_at = datetime.utcnow()
    r.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(r)
    db.refresh(r, attribute_names=["lead", "assigned_to"])
    return _reminder_read(r)


@app.delete("/reminders/{reminder_id}", status_code=204)
def delete_reminder(
    reminder_id: int,
    actor: models.User = Depends(require_crm_access),
    db: Session = Depends(get_db),
):
    r = db.query(models.Reminder).filter(models.Reminder.id == reminder_id).first()
    if not r:
        raise HTTPException(404, "Eslatma topilmadi")
    db.delete(r)
    db.commit()


# ── Notifications ─────────────────────────────────────────────────────────────

@app.get("/notifications", response_model=List[schemas.NotificationRead])
def list_notifications(
    unread_only: bool = Query(False),
    limit: int = Query(30, le=100),
    actor: models.User = Depends(require_auth),
    db: Session = Depends(get_db),
):
    q = db.query(models.Notification).filter(models.Notification.user_id == actor.id)
    if unread_only:
        q = q.filter(models.Notification.is_read == False)  # noqa: E712
    return q.order_by(models.Notification.created_at.desc()).limit(limit).all()


@app.get("/notifications/unread-count")
def notifications_unread_count(
    actor: models.User = Depends(require_auth),
    db: Session = Depends(get_db),
):
    n = (
        db.query(func.count(models.Notification.id))
        .filter(models.Notification.user_id == actor.id, models.Notification.is_read == False)  # noqa: E712
        .scalar()
    )
    return {"count": n or 0}


@app.post("/notifications/{notification_id}/read", status_code=204)
def mark_notification_read(
    notification_id: int,
    actor: models.User = Depends(require_auth),
    db: Session = Depends(get_db),
):
    n = db.query(models.Notification).filter(
        models.Notification.id == notification_id,
        models.Notification.user_id == actor.id,
    ).first()
    if not n:
        raise HTTPException(404, "Topilmadi")
    n.is_read = True
    db.commit()


@app.post("/notifications/read-all", status_code=204)
def mark_all_notifications_read(
    actor: models.User = Depends(require_auth),
    db: Session = Depends(get_db),
):
    db.query(models.Notification).filter(
        models.Notification.user_id == actor.id,
        models.Notification.is_read == False,  # noqa: E712
    ).update({"is_read": True})
    db.commit()


# ── Lead analytics ────────────────────────────────────────────────────────────

@app.get("/leads/analytics", response_model=schemas.LeadAnalyticsRead)
def lead_analytics(
    actor: models.User = Depends(require_crm_access),
    db: Session = Depends(get_db),
):
    lead_q = db.query(models.Lead)
    if actor.role in (UserRole.hunter.value, UserRole.sales.value):
        lead_q = lead_q.filter(models.Lead.created_by_id == actor.id)
    total = lead_q.count()

    stage_counts = dict(
        lead_q.with_entities(models.Lead.stage_id, func.count(models.Lead.id))
        .group_by(models.Lead.stage_id).all()
    )
    stages = (
        db.query(models.LeadStage)
        .filter(models.LeadStage.is_archived == False)  # noqa: E712
        .order_by(models.LeadStage.order.asc()).all()
    )
    distribution = [
        schemas.FunnelStep(
            slug=s.slug, name=s.name, color=s.color,
            count=stage_counts.get(s.id, 0),
            percentage=round(stage_counts.get(s.id, 0) / total * 100, 1) if total else 0.0,
        ) for s in stages
    ]
    won = sum(stage_counts.get(s.id, 0) for s in stages if s.kind == "won")
    lost = sum(stage_counts.get(s.id, 0) for s in stages if s.kind == "lost")
    won_stage_ids = [s.id for s in stages if s.kind == "won"]

    # Manba bo'yicha konversiya
    src_total = dict(
        lead_q.with_entities(models.Lead.source_id, func.count(models.Lead.id))
        .group_by(models.Lead.source_id).all()
    )
    src_won = {}
    if won_stage_ids:
        src_won = dict(
            lead_q.filter(models.Lead.stage_id.in_(won_stage_ids))
            .with_entities(models.Lead.source_id, func.count(models.Lead.id))
            .group_by(models.Lead.source_id).all()
        )
    source_names = {s.id: s.name for s in db.query(models.LeadSource).all()}
    sources = []
    for sid, tot in sorted(src_total.items(), key=lambda kv: kv[1], reverse=True):
        enr = src_won.get(sid, 0)
        sources.append(schemas.SourceStat(
            source=source_names.get(sid, "Noma'lum"),
            total=tot, enrolled=enr,
            conversion_rate=round(enr / tot * 100, 1) if tot else 0.0,
        ))

    return schemas.LeadAnalyticsRead(
        total=total, won=won, lost=lost,
        conversion=round(won / total * 100, 1) if total else 0.0,
        distribution=distribution, sources=sources,
    )


# ── Intake forms (public lead capture) ────────────────────────────────────────

def _intake_read(f: models.IntakeForm) -> schemas.IntakeFormRead:
    return schemas.IntakeFormRead(
        id=f.id, slug=f.slug, name=f.name, title=f.title, description=f.description,
        source_id=f.source_id, source_name=f.source.name if f.source else None,
        is_active=f.is_active, submissions=f.submissions, created_at=f.created_at,
    )


@app.get("/intake-forms", response_model=List[schemas.IntakeFormRead])
def list_intake_forms(
    actor: models.User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    forms = (
        db.query(models.IntakeForm).options(joinedload(models.IntakeForm.source))
        .order_by(models.IntakeForm.created_at.desc()).all()
    )
    return [_intake_read(f) for f in forms]


@app.post("/intake-forms", response_model=schemas.IntakeFormRead, status_code=201)
def create_intake_form(
    payload: schemas.IntakeFormCreate,
    actor: models.User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    slug = _slugify(payload.name)[:50]
    if db.query(models.IntakeForm).filter(models.IntakeForm.slug == slug).first():
        slug = f"{slug}-{int(datetime.utcnow().timestamp())}"
    form = models.IntakeForm(
        slug=slug, name=payload.name, title=payload.title or payload.name,
        description=payload.description, source_id=payload.source_id,
        is_active=True, submissions=0, created_at=datetime.utcnow(),
    )
    db.add(form)
    db.commit()
    db.refresh(form)
    db.refresh(form, attribute_names=["source"])
    return _intake_read(form)


@app.put("/intake-forms/{form_id}", response_model=schemas.IntakeFormRead)
def update_intake_form(
    form_id: int,
    payload: schemas.IntakeFormUpdate,
    actor: models.User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    form = db.query(models.IntakeForm).filter(models.IntakeForm.id == form_id).first()
    if not form:
        raise HTTPException(404, "Forma topilmadi")
    for field, val in payload.dict(exclude_unset=True).items():
        setattr(form, field, val)
    db.commit()
    db.refresh(form)
    db.refresh(form, attribute_names=["source"])
    return _intake_read(form)


@app.delete("/intake-forms/{form_id}", status_code=204)
def delete_intake_form(
    form_id: int,
    actor: models.User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    form = db.query(models.IntakeForm).filter(models.IntakeForm.id == form_id).first()
    if not form:
        raise HTTPException(404, "Forma topilmadi")
    db.delete(form)
    db.commit()


# ── Public (auth talab qilmaydigan) intake endpoint'lar ───────────────────────

@app.get("/public/intake/{slug}", response_model=schemas.PublicIntakeConfig)
def public_intake_config(slug: str, db: Session = Depends(get_db)):
    form = db.query(models.IntakeForm).filter(models.IntakeForm.slug == slug).first()
    if not form or not form.is_active:
        raise HTTPException(404, "Forma topilmadi yoki faol emas")
    return schemas.PublicIntakeConfig(
        slug=form.slug, name=form.name, title=form.title or form.name,
        description=form.description, is_active=form.is_active,
    )


@app.post("/public/intake/{slug}", status_code=201)
def public_intake_submit(
    slug: str,
    payload: schemas.PublicLeadSubmit,
    request: Request,
    db: Session = Depends(get_db),
):
    ip = _client_ip(request)
    # Spam himoyasi: bir IP daqiqasiga 5 ta, soatiga 30 ta ariza
    rate_limit(f"intake-min:{ip}", limit=5, window=60)
    rate_limit(f"intake-hour:{ip}", limit=30, window=3600)

    form = db.query(models.IntakeForm).filter(models.IntakeForm.slug == slug).first()
    if not form or not form.is_active:
        raise HTTPException(404, "Forma topilmadi yoki faol emas")
    stage = _default_stage(db)
    lead = models.Lead(
        full_name=payload.full_name.strip(),
        phone=payload.phone.strip(),
        course_interest=payload.course_interest,
        parent_phone=payload.parent_phone,
        notes=payload.notes,
        source_id=form.source_id,
        status=stage.slug if stage else models.LeadStatus.new.value,
        stage_id=stage.id if stage else None,
        is_shared=True,                       # ommaviy formadan kelgan lid — umumiy havzaga
        created_by_id=_system_user_id(db),
        created_at=datetime.utcnow(),
    )
    db.add(lead)
    db.flush()
    _log_lead_activity(
        db, lead_id=lead.id, action="created",
        description=f"Ommaviy forma orqali: {form.name}", author_id=None,
    )
    form.submissions = (form.submissions or 0) + 1
    recipients = [u[0] for u in db.query(models.User.id).filter(
        models.User.is_active == True,  # noqa: E712
        models.User.role.in_([UserRole.admin.value, UserRole.call_center.value,
                              UserRole.hunter.value, UserRole.sales.value]),
    ).all()]
    _notify(
        db, user_ids=recipients,
        title="Ommaviy formadan yangi lid",
        body=f"{lead.full_name} · {form.name}",
        link=f"/leads?lead={lead.id}", ntype="new_lead",
    )
    db.commit()
    return {"ok": True, "message": "Arizangiz qabul qilindi"}


def _system_user_id(db) -> int:
    """Ommaviy formadan kelgan lidlar uchun 'egasi' — birinchi admin."""
    u = (
        db.query(models.User.id)
        .filter(models.User.role == UserRole.admin.value)
        .order_by(models.User.id.asc()).first()
    )
    if not u:
        u = db.query(models.User.id).order_by(models.User.id.asc()).first()
    return u[0]


# ═══════════════════════════════════════════════════════════════════════════
#  OTA-ONA AKKAUNTLARI — hunter/admin boshqaruvi (+ dars jadvali slotlari)
# ═══════════════════════════════════════════════════════════════════════════
import secrets as _secrets
import string as _string


def _gen_password(n: int = 8) -> str:
    alphabet = _string.ascii_letters + _string.digits
    return "".join(_secrets.choice(alphabet) for _ in range(n))


def _parent_account_read(p: models.Parent) -> schemas.ParentAccountRead:
    return schemas.ParentAccountRead(
        id=p.id, full_name=p.full_name, display_name=p.display_name,
        phone=p.phone, username=p.username, is_active=p.is_active, created_at=p.created_at,
        children=[
            schemas.ParentChildInfo(student_id=c.student_id, student_name=c.student.full_name)
            for c in p.children if c.student
        ],
    )


@app.post("/parents", response_model=schemas.ParentAccountCreated, status_code=201)
def create_parent_account(
    payload: schemas.ParentAccountCreate,
    actor: models.User = Depends(require_hunter),
    db: Session = Depends(get_db),
):
    """Ota-ona akkaunti yaratish (hunter/admin). Parol bir marta qaytariladi."""
    phone = payload.phone.strip()
    username = (payload.username or phone).strip()
    if db.query(models.Parent).filter(models.Parent.phone == phone).first():
        raise HTTPException(409, "Bu telefon bilan ota-ona allaqachon mavjud")
    if db.query(models.Parent).filter(models.Parent.username == username).first():
        raise HTTPException(409, "Bu login band")

    raw_password = payload.password or _gen_password()
    parent = models.Parent(
        full_name=payload.full_name.strip(), display_name=payload.display_name,
        phone=phone, username=username, hashed_password=hash_password(raw_password),
        is_active=True, created_by_id=actor.id, created_at=datetime.utcnow(),
    )
    db.add(parent)
    db.flush()
    for sid in dict.fromkeys(payload.student_ids):   # dublikatlarsiz
        if db.query(models.Student).filter(models.Student.id == sid).first():
            db.add(models.ParentChild(parent_id=parent.id, student_id=sid, created_at=datetime.utcnow()))
    db.commit()
    db.refresh(parent)
    out = schemas.ParentAccountCreated(**_parent_account_read(parent).dict())
    out.generated_password = raw_password
    return out


@app.get("/parents", response_model=List[schemas.ParentAccountRead])
def list_parent_accounts(
    search: Optional[str] = Query(None),
    actor: models.User = Depends(require_crm_access),
    db: Session = Depends(get_db),
):
    q = db.query(models.Parent).options(
        joinedload(models.Parent.children).joinedload(models.ParentChild.student),
    )
    if search:
        q = q.filter(
            models.Parent.full_name.ilike(f"%{search}%") |
            models.Parent.phone.ilike(f"%{search}%") |
            models.Parent.username.ilike(f"%{search}%")
        )
    return [_parent_account_read(p) for p in q.order_by(models.Parent.full_name).all()]


@app.get("/parents/{parent_id}", response_model=schemas.ParentAccountRead)
def get_parent_account(
    parent_id: int,
    actor: models.User = Depends(require_crm_access),
    db: Session = Depends(get_db),
):
    p = db.query(models.Parent).filter(models.Parent.id == parent_id).first()
    if not p:
        raise HTTPException(404, "Ota-ona topilmadi")
    return _parent_account_read(p)


@app.patch("/parents/{parent_id}", response_model=schemas.ParentAccountRead)
def update_parent_account(
    parent_id: int,
    payload: schemas.ParentAccountUpdate,
    actor: models.User = Depends(require_hunter),
    db: Session = Depends(get_db),
):
    p = db.query(models.Parent).filter(models.Parent.id == parent_id).first()
    if not p:
        raise HTTPException(404, "Ota-ona topilmadi")
    data = payload.dict(exclude_unset=True)
    if "phone" in data and data["phone"] != p.phone:
        if db.query(models.Parent).filter(models.Parent.phone == data["phone"]).first():
            raise HTTPException(409, "Bu telefon band")
    for k, v in data.items():
        setattr(p, k, v)
    db.commit()
    db.refresh(p)
    return _parent_account_read(p)


@app.post("/parents/{parent_id}/reset-password")
def reset_parent_password(
    parent_id: int,
    payload: schemas.ResetParentPassword,
    actor: models.User = Depends(require_hunter),
    db: Session = Depends(get_db),
):
    p = db.query(models.Parent).filter(models.Parent.id == parent_id).first()
    if not p:
        raise HTTPException(404, "Ota-ona topilmadi")
    raw = payload.password or _gen_password()
    p.hashed_password = hash_password(raw)
    # Barcha refresh tokenlarni bekor qilamiz (xavfsizlik)
    db.query(models.ParentRefreshToken).filter(
        models.ParentRefreshToken.parent_id == p.id,
        models.ParentRefreshToken.revoked_at.is_(None),
    ).update({"revoked_at": datetime.utcnow()})
    db.commit()
    return {"generated_password": raw}


@app.post("/parents/{parent_id}/children", response_model=schemas.ParentAccountRead)
def link_parent_child(
    parent_id: int,
    payload: schemas.LinkChildRequest,
    actor: models.User = Depends(require_hunter),
    db: Session = Depends(get_db),
):
    p = db.query(models.Parent).filter(models.Parent.id == parent_id).first()
    if not p:
        raise HTTPException(404, "Ota-ona topilmadi")
    if not db.query(models.Student).filter(models.Student.id == payload.student_id).first():
        raise HTTPException(404, "Talaba topilmadi")
    exists = db.query(models.ParentChild).filter(
        models.ParentChild.parent_id == parent_id,
        models.ParentChild.student_id == payload.student_id,
    ).first()
    if not exists:
        db.add(models.ParentChild(parent_id=parent_id, student_id=payload.student_id, created_at=datetime.utcnow()))
        db.commit()
    db.refresh(p)
    return _parent_account_read(p)


@app.delete("/parents/{parent_id}/children/{student_id}", status_code=204)
def unlink_parent_child(
    parent_id: int,
    student_id: int,
    actor: models.User = Depends(require_hunter),
    db: Session = Depends(get_db),
):
    link = db.query(models.ParentChild).filter(
        models.ParentChild.parent_id == parent_id,
        models.ParentChild.student_id == student_id,
    ).first()
    if link:
        db.delete(link)
        db.commit()


@app.post("/parents/broadcast", response_model=schemas.ParentBroadcastResult)
def broadcast_to_parents(
    payload: schemas.ParentBroadcastRequest,
    actor: models.User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Talabalarga bog'langan Telegram ID'lari (Student.telegram_user_id) orqali
    barcha ota-onalarga bir xil matnli xabar yuboradi — bitta ID bir nechta
    talabaga tegishli bo'lsa ham, unga faqat bitta xabar boradi."""
    ids = [
        row[0] for row in
        db.query(models.Student.telegram_user_id)
        .filter(models.Student.telegram_user_id.isnot(None), models.Student.telegram_user_id != "")
        .distinct()
        .all()
    ]
    safe_text = html.escape(payload.text)
    sent = 0
    for tid in ids:
        ok, _ = send_telegram_message(tid, safe_text)
        if ok:
            sent += 1
    write_audit(db, entity_type="parent_broadcast", entity_id=0, action="send",
                changed_by_id=actor.id,
                new_value={"text": payload.text, "total": len(ids), "sent": sent})
    db.commit()
    return schemas.ParentBroadcastResult(total=len(ids), sent=sent, failed=len(ids) - sent)


# ── Dars jadvali slotlari (strukturaviy jadval) ──────────────────────────────

@app.get("/groups/{group_id}/schedule-slots", response_model=List[schemas.ScheduleSlotRead])
def list_schedule_slots(
    group_id: int,
    actor: models.User = Depends(require_auth),
    db: Session = Depends(get_db),
):
    return (
        db.query(models.ScheduleSlot)
        .filter(models.ScheduleSlot.group_id == group_id)
        .order_by(models.ScheduleSlot.day_of_week, models.ScheduleSlot.start_time)
        .all()
    )


@app.post("/groups/{group_id}/schedule-slots", response_model=schemas.ScheduleSlotRead, status_code=201)
def create_schedule_slot(
    group_id: int,
    payload: schemas.ScheduleSlotCreate,
    actor: models.User = Depends(require_lms_write),
    db: Session = Depends(get_db),
):
    if not db.query(models.Group).filter(models.Group.id == group_id).first():
        raise HTTPException(404, "Guruh topilmadi")
    slot = models.ScheduleSlot(
        group_id=group_id, day_of_week=payload.day_of_week,
        start_time=payload.start_time, end_time=payload.end_time,
        room=payload.room, created_at=datetime.utcnow(),
    )
    db.add(slot)
    db.commit()
    db.refresh(slot)
    return slot


@app.delete("/schedule-slots/{slot_id}", status_code=204)
def delete_schedule_slot(
    slot_id: int,
    actor: models.User = Depends(require_lms_write),
    db: Session = Depends(get_db),
):
    slot = db.query(models.ScheduleSlot).filter(models.ScheduleSlot.id == slot_id).first()
    if slot:
        db.delete(slot)
        db.commit()
# ── Camera API ────────────────────────────────────────────────────────────────

@app.post("/students/{student_id}/face-photo")
async def upload_student_face_photo(
    student_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _: models.User = Depends(require_lms_write),
):
    s = db.query(models.Student).filter(models.Student.id == student_id).first()
    if not s:
        raise HTTPException(status_code=404, detail="Talaba topilmadi")
    ext = (file.filename or "").rsplit(".", 1)[-1].lower()
    if ext not in ("jpg", "jpeg", "png", "webp"):
        raise HTTPException(status_code=400, detail="Faqat jpg/png/webp rasm yuklang")
    filename = f"{student_id}.{ext}"
    filepath = os.path.join(STUDENT_PHOTOS_DIR, filename)
    contents = await file.read()
    with open(filepath, "wb") as f:
        f.write(contents)
    s.photo_path = f"student_photos/{filename}"
    db.commit()
    return {"photo_url": f"/uploads/student_photos/{filename}"}

def _require_camera_key(x_camera_key: Optional[str] = Header(None, alias="x-camera-key")) -> None:
    if not CAMERA_API_KEY or x_camera_key != CAMERA_API_KEY:
        raise HTTPException(status_code=403, detail="Camera API key noto'g'ri")


@app.post("/users/{user_id}/face-photo")
async def upload_user_face_photo(
    user_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _: models.User = Depends(require_admin),
):
    u = db.query(models.User).filter(models.User.id == user_id).first()
    if not u:
        raise HTTPException(status_code=404, detail="Foydalanuvchi topilmadi")
    ext = (file.filename or "").rsplit(".", 1)[-1].lower()
    if ext not in ("jpg", "jpeg", "png", "webp"):
        raise HTTPException(status_code=400, detail="Faqat jpg/png/webp rasm yuklang")
    filename = f"{user_id}.{ext}"
    filepath = os.path.join(STAFF_PHOTOS_DIR, filename)
    contents = await file.read()
    with open(filepath, "wb") as f:
        f.write(contents)
    u.face_photo_path = f"staff_photos/{filename}"
    db.commit()
    return {"face_photo_url": f"/uploads/staff_photos/{filename}"}


@app.get("/camera/students")
def camera_students(
    db: Session = Depends(get_db),
    _: None = Depends(_require_camera_key),
):
    """Kamera servisi uchun — faol o'quvchilar ro'yxati (foto, guruh, jadval, to'lov holati)."""
    now_   = datetime.utcnow()
    month_ = now_.month
    year_  = now_.year

    students = (
        db.query(models.Student)
        .options(
            selectinload(models.Student.group_memberships).selectinload(models.GroupStudent.group),
            selectinload(models.Student.payments),
        )
        .filter(models.Student.is_active == True, models.Student.is_archived == False)
        .all()
    )
    result = []
    for s in students:
        groups = [
            {"id": gs.group_id, "name": gs.group.name, "schedule": gs.group.schedule}
            for gs in s.group_memberships
            if gs.group and gs.group.is_active
        ]
        paid_this_month = sum(
            float(p.amount) for p in s.payments
            if p.month == month_ and p.year == year_
        )
        is_debtor = (paid_this_month == 0) and len(groups) > 0
        result.append({
            "id":         s.id,
            "full_name":  s.full_name,
            "telegram_id": s.telegram_id,
            "photo_url":  f"/uploads/{s.photo_path}" if s.photo_path else None,
            "groups":     groups,
            "is_debtor":  is_debtor,
        })
    return result


@app.get("/camera/staff")
def camera_staff(
    db: Session = Depends(get_db),
    _: None = Depends(_require_camera_key),
):
    """Kamera servisi uchun — yuz rasmi bor xodimlar ro'yxati."""
    staff = (
        db.query(models.User)
        .filter(
            models.User.is_active == True,
            models.User.face_photo_path.isnot(None),
        )
        .all()
    )
    return [
        {
            "id": u.id,
            "full_name": u.full_name or u.username,
            "role": u.role,
            "face_photo_url": f"/uploads/{u.face_photo_path}" if u.face_photo_path else None,
        }
        for u in staff
    ]


@app.get("/camera/photo/student/{student_id}")
def camera_student_photo(
    student_id: int,
    db: Session = Depends(get_db),
    _: None = Depends(_require_camera_key),
):
    """Kamera servisi uchun — o'quvchi yuz rasmi (bytes)."""
    from fastapi.responses import FileResponse as FR
    s = db.query(models.Student).filter(models.Student.id == student_id).first()
    if not s or not s.photo_path:
        raise HTTPException(status_code=404, detail="Rasm topilmadi")
    path = os.path.join(UPLOAD_DIR, s.photo_path)
    if not os.path.exists(path):
        raise HTTPException(status_code=404, detail="Fayl topilmadi")
    return FR(path)


@app.get("/camera/photo/staff/{user_id}")
def camera_staff_photo(
    user_id: int,
    db: Session = Depends(get_db),
    _: None = Depends(_require_camera_key),
):
    """Kamera servisi uchun — xodim yuz rasmi (bytes)."""
    from fastapi.responses import FileResponse as FR
    u = db.query(models.User).filter(models.User.id == user_id).first()
    if not u or not u.face_photo_path:
        raise HTTPException(status_code=404, detail="Rasm topilmadi")
    path = os.path.join(UPLOAD_DIR, u.face_photo_path)
    if not os.path.exists(path):
        raise HTTPException(status_code=404, detail="Fayl topilmadi")
    return FR(path)


@app.post("/camera/checkin", status_code=201)
def camera_checkin(
    payload: schemas.CameraCheckin,
    db: Session = Depends(get_db),
    _: None = Depends(_require_camera_key),
):
    """Kamera tomonidan aniqlangan keldi/ketdi hodisasini DB ga saqlaydi."""
    detected_at = payload.detected_at or datetime.utcnow()
    event_type  = "keldi" if payload.event in ("arrival", "keldi") else "ketdi"

    rec = models.CameraAttendance(
        student_id  = payload.student_id if payload.person_type == "student" else None,
        staff_id    = payload.student_id if payload.person_type == "staff"   else None,
        person_type = payload.person_type or "student",
        event_type  = event_type,
        detected_at = detected_at,
    )
    db.add(rec)
    db.commit()
    return {"status": "ok", "id": rec.id}


@app.get("/students/{student_id}/camera-attendance")
def student_camera_attendance(
    student_id: int,
    days: int = Query(30, ge=1, le=365),
    db: Session = Depends(get_db),
    _: models.User = Depends(require_staff),
):
    """O'quvchining kamera orqali keldi/ketdi tarixi (so'nggi N kun)."""
    from datetime import timedelta
    since = datetime.utcnow() - timedelta(days=days)
    rows = (
        db.query(models.CameraAttendance)
        .filter(
            models.CameraAttendance.student_id == student_id,
            models.CameraAttendance.detected_at >= since,
        )
        .order_by(models.CameraAttendance.detected_at.desc())
        .limit(200)
        .all()
    )
    return [
        {
            "id":          r.id,
            "event_type":  r.event_type,
            "detected_at": r.detected_at.isoformat(),
        }
        for r in rows
    ]


# ── Staff warnings (audit) ─────────────────────────────────────────────────────
# Audit xodimga ogohlantirish shu yerdan beradi; bot faqat xabarni Telegram orqali
# yetkazib beruvchi vositachi — botga hech qanday so'rov yozilmaydi, aksincha shu
# yerning o'zi mavjud send_telegram_message() yordamida to'g'ridan-to'g'ri yuboradi.

SEVERITY_EMOJI = {"gray": "⚪️", "yellow": "\U0001f7e1", "red": "\U0001f534"}


def _staff_name(u: Optional[models.User]) -> Optional[str]:
    if not u:
        return None
    return u.full_name or u.username


def _staff_warning_read(w: models.StaffWarning) -> schemas.StaffWarningRead:
    return schemas.StaffWarningRead(
        id=w.id, staff_id=w.staff_id, staff_name=_staff_name(w.staff),
        issued_by_id=w.issued_by_id, issued_by_name=_staff_name(w.issued_by),
        discipline_code_id=w.discipline_code_id, code=w.code, severity=w.severity,
        reason=w.reason, photo_path=w.photo_path,
        created_at=w.created_at, notified_at=w.notified_at, notify_error=w.notify_error,
        cancelled_at=w.cancelled_at, cancelled_by_name=_staff_name(w.cancelled_by),
    )


def _send_staff_warning_notification(w: models.StaffWarning, staff: models.User) -> None:
    """`w`ning notified_at/notify_error maydonlarini joyida yangilaydi — db.commit() chaqiruvchida."""
    if not staff.telegram_chat_id:
        w.notify_error = "Xodimga Telegram ID biriktirilmagan (bot orqali /start bosishi kerak)"
        return
    emoji = SEVERITY_EMOJI.get(w.severity, "⚠️")
    local_time = (w.created_at + timedelta(hours=5)).strftime("%d.%m.%Y %H:%M")
    lines = [f"{emoji} <b>Sizga ogohlantirish berildi</b>", ""]
    if w.code:
        lines.append(f"Kod: {html.escape(w.code)}")
    lines.append(f"Sabab: {html.escape(w.reason)}")
    lines.append(f"Sana: {local_time}")
    ok, err = send_telegram_message(staff.telegram_chat_id, "\n".join(lines))
    if ok:
        w.notified_at = datetime.utcnow()
        w.notify_error = None
    else:
        w.notify_error = err


@app.get("/staff-options", response_model=List[schemas.StaffOption])
def list_staff_options(db: Session = Depends(get_db), _: models.User = Depends(require_audit)):
    """Ogohlantirish beriladigan xodim tanlagichi uchun — /users kabi metodist-only emas."""
    return (
        db.query(models.User)
        .filter(models.User.is_active.is_(True))
        .order_by(models.User.full_name, models.User.username)
        .all()
    )


@app.get("/discipline-codes", response_model=List[schemas.DisciplineCodeRead])
def list_discipline_codes(db: Session = Depends(get_db), _: models.User = Depends(require_auth)):
    codes = (
        db.query(models.DisciplineCode)
        .filter(models.DisciplineCode.is_active.is_(True))
        .order_by(models.DisciplineCode.code)
        .all()
    )
    return codes


@app.get("/staff-warnings", response_model=List[schemas.StaffWarningRead])
def list_staff_warnings(
    staff_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    _: models.User = Depends(require_audit),
):
    q = db.query(models.StaffWarning).options(
        joinedload(models.StaffWarning.staff),
        joinedload(models.StaffWarning.issued_by),
        joinedload(models.StaffWarning.cancelled_by),
    )
    if staff_id:
        q = q.filter(models.StaffWarning.staff_id == staff_id)
    return [_staff_warning_read(w) for w in q.order_by(models.StaffWarning.created_at.desc()).all()]


@app.post("/staff-warnings", response_model=schemas.StaffWarningRead, status_code=201)
def create_staff_warning(
    payload: schemas.StaffWarningCreate,
    db: Session = Depends(get_db),
    actor: models.User = Depends(require_audit),
):
    staff = db.query(models.User).filter(models.User.id == payload.staff_id).first()
    if not staff:
        raise HTTPException(status_code=404, detail="Xodim topilmadi")

    code = None
    severity = payload.severity
    reason = payload.reason
    if payload.discipline_code_id:
        dc = db.query(models.DisciplineCode).filter(
            models.DisciplineCode.id == payload.discipline_code_id,
            models.DisciplineCode.is_active.is_(True),
        ).first()
        if not dc:
            raise HTTPException(status_code=404, detail="Kodeks moddasi topilmadi")
        code, severity, reason = dc.code, dc.severity, dc.text
    elif not severity or not reason:
        raise HTTPException(status_code=400, detail="Kodeks tanlang yoki daraja va sabab kiriting")

    w = models.StaffWarning(
        staff_id=staff.id, issued_by_id=actor.id,
        discipline_code_id=payload.discipline_code_id, code=code,
        severity=severity, reason=reason,
    )
    db.add(w)
    db.flush()
    _send_staff_warning_notification(w, staff)
    write_audit(db, entity_type="staff_warning", entity_id=w.id, action="create",
                changed_by_id=actor.id,
                new_value={"staff_id": staff.id, "code": code, "severity": severity, "reason": reason})
    db.commit()
    db.refresh(w)
    return _staff_warning_read(w)


@app.post("/staff-warnings/{warning_id}/resend", response_model=schemas.StaffWarningRead)
def resend_staff_warning(warning_id: int, db: Session = Depends(get_db), actor: models.User = Depends(require_audit)):
    w = db.query(models.StaffWarning).filter(models.StaffWarning.id == warning_id).first()
    if not w:
        raise HTTPException(status_code=404, detail="Ogohlantirish topilmadi")
    if w.cancelled_at:
        raise HTTPException(status_code=400, detail="Bekor qilingan ogohlantirish qayta yuborilmaydi")
    staff = db.query(models.User).filter(models.User.id == w.staff_id).first()
    _send_staff_warning_notification(w, staff)
    db.commit()
    db.refresh(w)
    return _staff_warning_read(w)


@app.post("/staff-warnings/{warning_id}/cancel", response_model=schemas.StaffWarningRead)
def cancel_staff_warning(warning_id: int, db: Session = Depends(get_db), actor: models.User = Depends(require_admin)):
    w = db.query(models.StaffWarning).filter(models.StaffWarning.id == warning_id).first()
    if not w:
        raise HTTPException(status_code=404, detail="Ogohlantirish topilmadi")
    if w.cancelled_at:
        raise HTTPException(status_code=400, detail="Ogohlantirish allaqachon bekor qilingan")
    w.cancelled_at = datetime.utcnow()
    w.cancelled_by_id = actor.id
    write_audit(db, entity_type="staff_warning", entity_id=w.id, action="cancel", changed_by_id=actor.id)
    db.commit()
    db.refresh(w)
    return _staff_warning_read(w)


@app.delete("/staff-warnings/{warning_id}", status_code=204)
def delete_staff_warning(warning_id: int, db: Session = Depends(get_db), actor: models.User = Depends(require_admin)):
    w = db.query(models.StaffWarning).filter(models.StaffWarning.id == warning_id).first()
    if not w:
        raise HTTPException(status_code=404, detail="Ogohlantirish topilmadi")
    write_audit(db, entity_type="staff_warning", entity_id=w.id, action="delete",
                changed_by_id=actor.id,
                old_value={"staff_id": w.staff_id, "code": w.code, "severity": w.severity, "reason": w.reason})
    db.delete(w)
    db.commit()
