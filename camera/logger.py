"""Keldi/Ketdi hodisalarini XLSX faylga yozadi (o'quvchilar va xodimlar alohida varaqda)."""
import os
import logging
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

import config

log = logging.getLogger("camera.logger")

_HEADERS_STUDENTS = ["Sana", "Vaqt", "Ism Familiya", "Hodisa", "Guruhlar"]
_HEADERS_STAFF    = ["Sana", "Vaqt", "Ism Familiya", "Lavozim", "Hodisa"]

_ARRIVAL_FILL   = PatternFill("solid", fgColor="C6EFCE")   # yashil
_DEPARTURE_FILL = PatternFill("solid", fgColor="FFCCCC")   # qizil


def _load_or_create() -> openpyxl.Workbook:
    path = config.EXCEL_LOG
    if os.path.exists(path):
        try:
            return openpyxl.load_workbook(path)
        except Exception as exc:
            log.warning("XLSX o'qib bo'lmadi, yangi yaratilmoqda: %s", exc)

    wb = openpyxl.Workbook()

    ws_st = wb.active
    ws_st.title = "O'quvchilar"
    _write_header(ws_st, _HEADERS_STUDENTS)

    ws_staff = wb.create_sheet("Xodimlar")
    _write_header(ws_staff, _HEADERS_STAFF)

    wb.save(path)
    return wb


def _write_header(ws, headers: list[str]) -> None:
    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="2563EB")
    ws.append(headers)
    for cell in ws[1]:
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center")


def log_event(person: dict, person_type: str, event: str) -> None:
    """
    Hodisani XLSX ga qo'shadi.
    person_type: 'student' | 'staff'
    event:       'arrival'  | 'departure'
    """
    try:
        wb = _load_or_create()
        now = datetime.now()
        date_str = now.strftime("%Y-%m-%d")
        time_str = now.strftime("%H:%M:%S")
        name = person.get("full_name", f"ID {person.get('id')}")
        event_uz = "Keldi" if event == "arrival" else "Ketdi"
        fill = _ARRIVAL_FILL if event == "arrival" else _DEPARTURE_FILL

        if person_type == "student":
            ws = wb["O'quvchilar"]
            groups = ", ".join(g["name"] for g in person.get("groups", []))
            ws.append([date_str, time_str, name, event_uz, groups])
        else:
            ws = wb["Xodimlar"]
            role_uz = _role_uz(person.get("role", ""))
            ws.append([date_str, time_str, name, role_uz, event_uz])

        # Oxirgi qatorga rang berish
        last_row = ws.max_row
        for cell in ws[last_row]:
            cell.fill = fill

        wb.save(config.EXCEL_LOG)
    except Exception as exc:
        log.error("XLSX ga yozishda xatolik: %s", exc)


def _role_uz(role: str) -> str:
    return {
        "admin":       "Administrator",
        "metodist":    "Metodist",
        "teacher":     "O'qituvchi",
        "call_center": "Call Center",
        "hunter":      "Hunter",
    }.get(role, role)
